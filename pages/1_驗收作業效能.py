from __future__ import annotations

from io import BytesIO

import streamlit as st
import pandas as pd

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

from common_ui import (
    set_page,               # set_page 內會 inject_logistics_theme()
    KPI,
    render_kpis,
    bar_topN,
    download_excel_card,    # ✅ 一行=按鈕、卡片外框、不分段
    sidebar_controls,       # ✅ 統一左側設定（TopN + 排除空窗）
    card_open,
    card_close,
    show_kpi_table,         # ✅ 整列紅/綠顯示（效率 < target 會紅）
)

from qc_core import run_qc_efficiency


# =========================================================
# 固定設定
# =========================================================

# 驗收作業上午、下午效率門檻統一為 29 筆／小時
QC_TARGET_EFFICIENCY = 29.0


# 固定休息時段：只在實際工作開始～結束時間有重疊時扣除。
# 例如 09:30-11:00 只扣 10:00-10:15；09:00-09:56 不扣。
FIXED_REST_WINDOWS = [
    {"start": "10:00", "end": "10:15", "data_entry": ""},
    {"start": "12:30", "end": "13:30", "data_entry": ""},
    {"start": "13:30", "end": "13:45", "data_entry": ""},
    {"start": "15:30", "end": "15:45", "data_entry": ""},
    {"start": "18:00", "end": "18:30", "data_entry": ""},
    {"start": "20:30", "end": "20:45", "data_entry": ""},
    {"start": "22:30", "end": "22:45", "data_entry": ""},
]


# =========================================================
# Helpers
# =========================================================
def _merge_fixed_rest_windows(exclude_windows):
    """
    將固定休息時段與側邊欄額外排除時段合併，並標記扣除來源。

    category 只用來拆欄顯示；總扣除分鐘仍等於三欄加總。
    側邊欄排除時段若有填 data_entry，歸類為「登入空窗」；
    未填 data_entry 則歸類為「空窗」。
    """
    merged = []
    seen = set()

    windows = [
        *[{**w, "category": "休息"} for w in FIXED_REST_WINDOWS],
        *[{**w, "category": "空窗"} for w in (exclude_windows or [])],
    ]

    for w in windows:
        if not isinstance(w, dict):
            continue

        data_entry = (w.get("data_entry") or "").strip()
        category = (w.get("category") or "空窗").strip()
        if category == "空窗" and data_entry:
            category = "登入空窗"

        item = {
            "start": (w.get("start") or "").strip(),
            "end": (w.get("end") or "").strip(),
            "data_entry": data_entry,
            "category": category,
        }

        key = (
            item["start"],
            item["end"],
            item["data_entry"],
        )

        if key in seen:
            continue

        seen.add(key)
        merged.append(item)

    return merged


def _adapt_exclude_windows_to_skip_rules(exclude_windows):
    """
    將 common_ui.sidebar_controls() 的 exclude_windows 格式：
      [{"start":"HH:MM","end":"HH:MM","data_entry":""}, ...]

    轉回 qc_core.run_qc_efficiency 需要的 skip_rules 格式，並保留 category
    供畫面與 Excel 分欄顯示。
    """
    skip_rules = []

    for w in exclude_windows or []:
        start_str = (w.get("start") or "").strip()
        end_str = (w.get("end") or "").strip()
        user_str = (w.get("data_entry") or "").strip()
        category = (w.get("category") or "空窗").strip()

        try:
            # 明確指定 HH:MM，避免被解析成日期
            start_time = pd.to_datetime(
                start_str,
                format="%H:%M",
            ).time()

            end_time = pd.to_datetime(
                end_str,
                format="%H:%M",
            ).time()

        except Exception:
            continue

        # 安全檢查：開始時間必須早於結束時間
        if start_time >= end_time:
            continue

        skip_rules.append(
            {
                "user": user_str,
                "t_start": start_time,
                "t_end": end_time,
                "category": category,
            }
        )

    return skip_rules


def _rule_applies_to_row(rule_user: str, row: pd.Series) -> bool:
    """
    空白 user 視為全部人適用；有填 user 時，需符合代碼或姓名欄位。
    """
    rule_user = (rule_user or "").strip()

    if not rule_user:
        return True

    for col in ("記錄輸入人", "資料輸入人", "輸入人", "姓名"):
        if col in row.index and str(row.get(col, "")).strip() == rule_user:
            return True

    return False


def _overlap_minutes_for_rules(
    first_dt,
    last_dt,
    row: pd.Series,
    skip_rules,
    category: str | None = None,
) -> float:
    """
    只扣工作區間與休息/排除區間實際重疊的分鐘數。

    category 為 None 時回傳全部扣除分鐘；指定 category 時只回傳該來源。
    """
    first_ts = pd.to_datetime(first_dt, errors="coerce")
    last_ts = pd.to_datetime(last_dt, errors="coerce")

    if pd.isna(first_ts) or pd.isna(last_ts) or last_ts <= first_ts:
        return 0.0

    total = 0.0

    for rule in skip_rules or []:
        if not isinstance(rule, dict):
            continue

        if category and rule.get("category") != category:
            continue

        if not _rule_applies_to_row(rule.get("user", ""), row):
            continue

        start_time = rule.get("t_start")
        end_time = rule.get("t_end")

        if start_time is None or end_time is None or start_time >= end_time:
            continue

        rest_start = pd.Timestamp.combine(first_ts.date(), start_time)
        rest_end = pd.Timestamp.combine(first_ts.date(), end_time)
        overlap_start = max(first_ts, rest_start)
        overlap_end = min(last_ts, rest_end)

        if overlap_end > overlap_start:
            total += (overlap_end - overlap_start).total_seconds() / 60.0

    return total


def _recalculate_rest_by_actual_overlap(
    df: pd.DataFrame,
    skip_rules,
) -> pd.DataFrame:
    """
    依每列實際工作區間重新計算：
    休息分鐘、登入空窗、空窗、總分鐘、總工時、效率。
    """
    required_cols = {
        "第一筆修訂日期",
        "最後一筆修訂日期",
        "休息分鐘",
        "總分鐘",
        "總工時",
        "效率",
    }

    if df is None or df.empty or not required_cols.issubset(df.columns):
        return df

    out = df.copy()

    for idx, row in out.iterrows():
        first_ts = pd.to_datetime(
            row.get("第一筆修訂日期"),
            errors="coerce",
        )
        last_ts = pd.to_datetime(
            row.get("最後一筆修訂日期"),
            errors="coerce",
        )

        if pd.isna(first_ts) or pd.isna(last_ts) or last_ts <= first_ts:
            continue

        rest_minutes = _overlap_minutes_for_rules(
            first_ts,
            last_ts,
            row,
            skip_rules,
            "休息",
        )
        login_idle_minutes = _overlap_minutes_for_rules(
            first_ts,
            last_ts,
            row,
            skip_rules,
            "登入空窗",
        )
        idle_minutes = _overlap_minutes_for_rules(
            first_ts,
            last_ts,
            row,
            skip_rules,
            "空窗",
        )

        excluded_minutes = rest_minutes + login_idle_minutes + idle_minutes
        raw_minutes = (last_ts - first_ts).total_seconds() / 60.0
        total_minutes = max(raw_minutes - excluded_minutes, 0.0)
        total_hours = total_minutes / 60.0

        pieces = pd.to_numeric(
            row.get("筆數"),
            errors="coerce",
        )

        out.at[idx, "休息分鐘"] = int(round(rest_minutes))
        out.at[idx, "登入空窗"] = int(round(login_idle_minutes))
        out.at[idx, "空窗"] = int(round(idle_minutes))
        out.at[idx, "總分鐘"] = round(total_minutes, 2)
        out.at[idx, "總工時"] = round(total_hours, 2)

        if pd.notna(pieces) and total_minutes > 0:
            out.at[idx, "效率"] = round(float(pieces) / total_minutes * 60.0, 2)
        else:
            out.at[idx, "效率"] = 0.0

    return out


def _recalculate_result_rest_by_actual_overlap(
    result: dict,
    skip_rules,
) -> dict:
    """
    修正 run_qc_efficiency 回傳的各個 DataFrame，避免固定休息被整班硬扣。
    """
    if not isinstance(result, dict):
        return result

    for key, value in list(result.items()):
        if isinstance(value, pd.DataFrame):
            result[key] = _recalculate_rest_by_actual_overlap(
                value,
                skip_rules,
            )

    return result


def _apply_actual_overlap_rest_to_excel(
    xlsx_bytes: bytes,
    skip_rules,
) -> bytes:
    """
    同步修正匯出 Excel 內的休息分鐘、登入空窗、空窗、總分鐘、總工時、效率。
    """
    if not xlsx_bytes:
        return xlsx_bytes

    input_buffer = BytesIO(xlsx_bytes)

    try:
        workbook = load_workbook(input_buffer)
    except Exception:
        return xlsx_bytes

    required_headers = {
        "第一筆修訂日期",
        "最後一筆修訂日期",
        "休息分鐘",
        "總分鐘",
        "總工時",
        "效率",
    }

    for worksheet in workbook.worksheets:
        header_row = None
        header_to_col = {}

        for row_idx in range(1, min(worksheet.max_row, 30) + 1):
            current = {}
            for col_idx in range(1, worksheet.max_column + 1):
                header = worksheet.cell(row=row_idx, column=col_idx).value
                header = str(header).strip() if header is not None else ""
                if header:
                    current[header] = col_idx

            if required_headers.issubset(current):
                header_row = row_idx
                header_to_col = current
                for new_header in ("空窗", "登入空窗"):
                    if new_header not in header_to_col:
                        insert_at = header_to_col["休息分鐘"] + 1
                        worksheet.insert_cols(insert_at)
                        worksheet.cell(row=row_idx, column=insert_at).value = new_header
                        for name, col in list(header_to_col.items()):
                            if col >= insert_at:
                                header_to_col[name] = col + 1
                        header_to_col[new_header] = insert_at
                break

        if header_row is None:
            continue

        for row_idx in range(header_row + 1, worksheet.max_row + 1):
            row_data = {
                header: worksheet.cell(
                    row=row_idx,
                    column=col_idx,
                ).value
                for header, col_idx in header_to_col.items()
            }

            row = pd.Series(row_data)
            first_dt = row.get("第一筆修訂日期")
            last_dt = row.get("最後一筆修訂日期")
            first_ts = pd.to_datetime(first_dt, errors="coerce")
            last_ts = pd.to_datetime(last_dt, errors="coerce")

            if pd.isna(first_ts) or pd.isna(last_ts) or last_ts <= first_ts:
                continue

            rest_minutes = _overlap_minutes_for_rules(
                first_ts,
                last_ts,
                row,
                skip_rules,
                "休息",
            )
            login_idle_minutes = _overlap_minutes_for_rules(
                first_ts,
                last_ts,
                row,
                skip_rules,
                "登入空窗",
            )
            idle_minutes = _overlap_minutes_for_rules(
                first_ts,
                last_ts,
                row,
                skip_rules,
                "空窗",
            )
            excluded_minutes = rest_minutes + login_idle_minutes + idle_minutes
            raw_minutes = (last_ts - first_ts).total_seconds() / 60.0
            total_minutes = max(raw_minutes - excluded_minutes, 0.0)
            total_hours = total_minutes / 60.0
            pieces = pd.to_numeric(row.get("筆數"), errors="coerce")
            efficiency = (
                round(float(pieces) / total_minutes * 60.0, 2)
                if pd.notna(pieces) and total_minutes > 0
                else 0.0
            )

            worksheet.cell(
                row=row_idx,
                column=header_to_col["休息分鐘"],
            ).value = int(round(rest_minutes))
            worksheet.cell(
                row=row_idx,
                column=header_to_col["登入空窗"],
            ).value = int(round(login_idle_minutes))
            worksheet.cell(
                row=row_idx,
                column=header_to_col["空窗"],
            ).value = int(round(idle_minutes))
            worksheet.cell(
                row=row_idx,
                column=header_to_col["總分鐘"],
            ).value = round(total_minutes, 2)
            worksheet.cell(
                row=row_idx,
                column=header_to_col["總工時"],
            ).value = round(total_hours, 2)
            worksheet.cell(
                row=row_idx,
                column=header_to_col["效率"],
            ).value = efficiency

    output_buffer = BytesIO()
    workbook.save(output_buffer)
    output_buffer.seek(0)

    return output_buffer.getvalue()


def _ensure_session_defaults():
    """
    建立 session_state 預設值。

    避免 Streamlit 因為調整左側設定、下載檔案或重新渲染，
    導致已計算的結果從畫面消失。
    """
    if "qc_last_result" not in st.session_state:
        st.session_state.qc_last_result = None

    if "qc_last_filename" not in st.session_state:
        st.session_state.qc_last_filename = None


def _split_am_pm(
    df: pd.DataFrame,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    依「時段」欄位分割上午與下午資料。

    支援：
    - 上午
    - 下午
    - AM
    - PM
    - AM 班
    - PM 班
    """
    if df is None or df.empty:
        return pd.DataFrame(), pd.DataFrame()

    df = df.copy()

    if "時段" not in df.columns:
        return pd.DataFrame(), pd.DataFrame()

    df["班別"] = (
        df["時段"]
        .astype(str)
        .str.strip()
        .replace(
            {
                "上午": "AM 班",
                "下午": "PM 班",
                "AM": "AM 班",
                "PM": "PM 班",
                "AM 班": "AM 班",
                "PM 班": "PM 班",
            }
        )
    )

    am_df = df[df["班別"] == "AM 班"].copy()
    pm_df = df[df["班別"] == "PM 班"].copy()

    return am_df, pm_df


def _safe_sum(
    df: pd.DataFrame,
    col: str,
) -> float:
    """
    安全加總指定欄位。
    """
    if df is None or df.empty or col not in df.columns:
        return 0.0

    values = pd.to_numeric(
        df[col],
        errors="coerce",
    ).fillna(0)

    return float(values.sum())


def _safe_mean(
    df: pd.DataFrame,
    col: str,
) -> float:
    """
    安全計算指定欄位平均值。
    """
    if df is None or df.empty or col not in df.columns:
        return 0.0

    values = pd.to_numeric(
        df[col],
        errors="coerce",
    ).dropna()

    if values.empty:
        return 0.0

    return float(values.mean())


def _safe_rate_ge(
    df: pd.DataFrame,
    col: str,
    target: float,
) -> float:
    """
    計算指定欄位大於等於門檻的達標率。
    """
    if df is None or df.empty or col not in df.columns:
        return 0.0

    values = pd.to_numeric(
        df[col],
        errors="coerce",
    ).dropna()

    if values.empty:
        return 0.0

    return float(
        (values >= float(target)).mean()
    )


def _copy_font_with_color(
    original_font: Font,
    color: str,
) -> Font:
    """
    保留原本字型設定，只修改字體顏色。
    """
    return Font(
        name=original_font.name,
        size=original_font.size,
        bold=original_font.bold,
        italic=original_font.italic,
        vertAlign=original_font.vertAlign,
        underline=original_font.underline,
        strike=original_font.strike,
        color=color,
        charset=original_font.charset,
        family=original_font.family,
        scheme=original_font.scheme,
        outline=original_font.outline,
        shadow=original_font.shadow,
        condense=original_font.condense,
        extend=original_font.extend,
    )


def _extract_numeric_value(value) -> float | None:
    """
    將 Excel 儲存格內容轉成數值。

    支援：
    - 29
    - 29.5
    - "29"
    - "29.5"
    - "29 筆/時"
    - "29%"
    - 含逗號的數字

    無法轉換則回傳 None。
    """
    if value is None:
        return None

    if isinstance(value, bool):
        return None

    if isinstance(value, (int, float)):
        return float(value)

    value_text = str(value).strip()

    if not value_text:
        return None

    value_text = (
        value_text
        .replace(",", "")
        .replace("%", "")
        .replace("筆／小時", "")
        .replace("筆/小時", "")
        .replace("筆／時", "")
        .replace("筆/時", "")
        .replace("筆", "")
        .strip()
    )

    try:
        return float(value_text)

    except (TypeError, ValueError):
        return None


def _find_efficiency_column(
    worksheet,
) -> tuple[int | None, int | None]:
    """
    搜尋工作表中的效率欄位。

    回傳：
    - 標題所在列
    - 效率欄位編號

    會優先尋找完全等於「效率」的欄名，
    再尋找包含「效率」的欄名。
    """
    if worksheet.max_row <= 0 or worksheet.max_column <= 0:
        return None, None

    # 報表上方可能有標題，因此搜尋前 30 列
    search_end_row = min(
        worksheet.max_row,
        30,
    )

    # 第一階段：優先找完全等於「效率」
    for row_idx in range(1, search_end_row + 1):
        for col_idx in range(1, worksheet.max_column + 1):
            cell_value = worksheet.cell(
                row=row_idx,
                column=col_idx,
            ).value

            header_text = (
                str(cell_value).strip()
                if cell_value is not None
                else ""
            )

            if header_text == "效率":
                return row_idx, col_idx

    # 第二階段：尋找包含「效率」的欄名
    # 排除「平均效率」等可能屬於統計摘要的欄位時，
    # 仍以第一個找到的欄位為準
    for row_idx in range(1, search_end_row + 1):
        for col_idx in range(1, worksheet.max_column + 1):
            cell_value = worksheet.cell(
                row=row_idx,
                column=col_idx,
            ).value

            header_text = (
                str(cell_value).strip()
                if cell_value is not None
                else ""
            )

            if "效率" in header_text:
                return row_idx, col_idx

    return None, None


def _is_existing_fail_fill(cell) -> bool:
    """
    判斷儲存格是否為常見的紅色未達標格式。
    """
    if cell.fill is None:
        return False

    fg_color = cell.fill.fgColor

    if fg_color is None:
        return False

    color_value = None

    if fg_color.type == "rgb":
        color_value = fg_color.rgb

    elif fg_color.type == "indexed":
        color_value = str(fg_color.indexed)

    if not color_value:
        return False

    color_value = str(color_value).upper()

    return color_value in {
        "FFC7CE",
        "00FFC7CE",
        "FFFFC7CE",
        "FF0000",
        "00FF0000",
        "FFFF0000",
        "9C0006",
        "009C0006",
        "FF9C0006",
    }


def _clear_old_fail_format(cell):
    """
    清除舊的紅色未達標格式。

    只在判斷該格屬於既有紅色格式時清除，
    避免破壞其他正常底色。
    """
    if not _is_existing_fail_fill(cell):
        return

    cell.fill = PatternFill(
        fill_type=None,
    )

    cell.font = _copy_font_with_color(
        cell.font,
        "000000",
    )


def _apply_excel_target_format(
    xlsx_bytes: bytes,
    *,
    target: float = 29.0,
) -> bytes:
    """
    將匯出的 Excel 報表重新套用效率門檻格式。

    規則：
    - 自動搜尋每個工作表中的「效率」欄位
    - 效率 < 29：整列淡紅底、深紅字
    - 效率 >= 29：清除可能存在的舊紅色未達標格式
    - 空白或非數值資料不處理
    - 不修改標題列
    """
    if not xlsx_bytes:
        return xlsx_bytes

    input_buffer = BytesIO(xlsx_bytes)

    try:
        workbook = load_workbook(input_buffer)

    except Exception:
        # 無法讀取時，維持原檔案內容
        return xlsx_bytes

    # 未達標格式：Excel 常見淡紅底、深紅字
    fail_fill = PatternFill(
        fill_type="solid",
        fgColor="FFC7CE",
    )

    fail_font_color = "9C0006"

    for worksheet in workbook.worksheets:
        header_row, efficiency_col = _find_efficiency_column(
            worksheet
        )

        if header_row is None or efficiency_col is None:
            continue

        # 從效率欄標題下一列開始
        for row_idx in range(
            header_row + 1,
            worksheet.max_row + 1,
        ):
            efficiency_cell = worksheet.cell(
                row=row_idx,
                column=efficiency_col,
            )

            efficiency_value = _extract_numeric_value(
                efficiency_cell.value
            )

            # 空白、文字或無法辨識的內容不處理
            if efficiency_value is None:
                continue

            is_failed = (
                efficiency_value < float(target)
            )

            for col_idx in range(
                1,
                worksheet.max_column + 1,
            ):
                cell = worksheet.cell(
                    row=row_idx,
                    column=col_idx,
                )

                if is_failed:
                    # 低於 29：整列紅色
                    cell.fill = fail_fill
                    cell.font = _copy_font_with_color(
                        cell.font,
                        fail_font_color,
                    )

                else:
                    # 等於或高於 29：
                    # 清除可能由舊門檻產生的紅色格式
                    _clear_old_fail_format(cell)

    output_buffer = BytesIO()

    workbook.save(output_buffer)

    output_buffer.seek(0)

    return output_buffer.getvalue()


def _render_shift_block(
    title: str,
    sdf: pd.DataFrame,
    *,
    top_n: int,
    target: float,
):
    """
    顯示單一班別的 KPI、明細表與效率排行。
    """

    # ======================
    # KPI 卡片
    # ======================
    card_open(
        f"{title}｜KPI",
        right_badge=f"門檻 {target:.0f} 筆／小時",
    )

    if sdf is None or sdf.empty:
        st.info("本班別無資料")
        card_close()
        return

    render_kpis(
        [
            KPI(
                "人數",
                f"{len(sdf):,}",
            ),
            KPI(
                "總驗收筆數",
                f"{_safe_sum(sdf, '筆數'):,.0f}",
            ),
            KPI(
                "總工時",
                f"{_safe_sum(sdf, '總工時'):.2f}",
            ),
            KPI(
                "平均效率",
                f"{_safe_mean(sdf, '效率'):.2f}",
            ),
            KPI(
                "達標率",
                f"{_safe_rate_ge(sdf, '效率', target):.0%}",
            ),
        ],
        cols=5,
    )

    card_close()

    # ======================
    # KPI 明細表
    # ======================
    card_open(
        f"{title}｜KPI 明細",
        right_badge=(
            f"紅色＜{target:.0f}｜"
            f"達標≥{target:.0f}"
        ),
    )

    show_kpi_table(
        sdf,
        eff_col="效率",
        target=float(target),
    )

    card_close()

    # ======================
    # Top N 效率排行
    # ======================
    card_open(
        f"{title}｜效率排行（Top {top_n}）",
        right_badge=f"門檻 {target:.0f}",
    )

    x_col = (
        "姓名"
        if "姓名" in sdf.columns
        else sdf.columns[0]
    )

    hover_cols = [
        col
        for col in [
            "筆數",
            "總工時",
            "休息分鐘",
            "登入空窗",
            "空窗",
        ]
        if col in sdf.columns
    ]

    bar_topN(
        sdf,
        x_col=x_col,
        y_col="效率",
        hover_cols=hover_cols,
        top_n=int(top_n),
        target=float(target),
        title=(
            f"低於 {target:.0f} 筆／小時自動標紅；"
            "虛線為達標門檻"
        ),
    )

    card_close()


# =========================================================
# Main
# =========================================================
def main():
    _ensure_session_defaults()

    set_page(
        "驗收作業效能（KPI）",
        icon="✅",
        subtitle=(
            "驗收作業｜人時效率｜AM / PM 班別｜"
            f"統一門檻 "
            f"{QC_TARGET_EFFICIENCY:.0f} 筆／小時"
        ),
    )

    # ======================
    # Sidebar：計算條件設定
    # ======================
    controls = sidebar_controls(
        default_top_n=30,
        enable_exclude_windows=True,
        state_key_prefix="qc",
    )

    top_n = int(
        controls.get(
            "top_n",
            30,
        )
    )

    exclude_windows = _merge_fixed_rest_windows(
        controls.get(
            "exclude_windows",
            [],
        )
    )

    skip_rules = _adapt_exclude_windows_to_skip_rules(
        exclude_windows
    )

    # ======================
    # Sidebar 固定門檻顯示
    # ======================
    with st.sidebar:
        st.divider()

        st.metric(
            "驗收效率門檻",
            f"{QC_TARGET_EFFICIENCY:.0f} 筆／小時",
            help="上午與下午均使用相同門檻。",
        )

        st.caption(
            "匯出 Excel 時，效率低於 29 的資料會整列顯示紅色。"
        )

    # ======================
    # 上傳資料與產出按鈕
    # ======================
    card_open(
        "📤 上傳作業原始資料（驗收）",
        right_badge="XLSX / XLS / CSV",
    )

    uploaded = st.file_uploader(
        "上傳驗收作業原始資料",
        type=[
            "xlsx",
            "xls",
            "csv",
        ],
        label_visibility="collapsed",
        key="qc_upload_file",
    )

    run_clicked = st.button(
        "🚀 產出 KPI",
        type="primary",
        disabled=(uploaded is None),
        use_container_width=True,
    )

    card_close()

    # ======================
    # 計算
    # ======================
    if run_clicked and uploaded is not None:
        try:
            with st.spinner("KPI 計算中，請稍候..."):
                result = run_qc_efficiency(
                    uploaded.getvalue(),
                    uploaded.name,
                    skip_rules,
                )

            if not result:
                st.session_state.qc_last_result = None
                st.session_state.qc_last_filename = None

                st.error(
                    "計算未產出結果，請確認上傳檔案內容與格式。"
                )

            else:
                # 強制統一上午與下午門檻為 29
                result["target_eff"] = QC_TARGET_EFFICIENCY
                result["target_eff_am"] = QC_TARGET_EFFICIENCY
                result["target_eff_pm"] = QC_TARGET_EFFICIENCY

                # 依實際工作區間重算扣休，避免未跨休息時段也被固定扣 15 分鐘。
                result = _recalculate_result_rest_by_actual_overlap(
                    result,
                    skip_rules,
                )

                # 匯出 Excel 重新套用門檻格式
                # 效率低於 29，整列顯示紅色
                if result.get("xlsx_bytes"):
                    result["xlsx_bytes"] = (
                        _apply_excel_target_format(
                            _apply_actual_overlap_rest_to_excel(
                                result["xlsx_bytes"],
                                skip_rules,
                            ),
                            target=QC_TARGET_EFFICIENCY,
                        )
                    )

                st.session_state.qc_last_result = result
                st.session_state.qc_last_filename = uploaded.name

                st.success(
                    "計算完成：上午與下午門檻均為 "
                    f"{QC_TARGET_EFFICIENCY:.0f} 筆／小時；"
                    "匯出報表低於 29 會顯示紅色。"
                )

        except Exception as exc:
            st.session_state.qc_last_result = None
            st.session_state.qc_last_filename = None

            st.error(
                f"KPI 計算失敗：{exc}"
            )

            return

    # ======================
    # 從 session_state 取用結果
    # ======================
    result = st.session_state.qc_last_result

    if not result:
        st.info(
            "請先上傳驗收作業原始資料，"
            "再點選「🚀 產出 KPI」。"
        )
        return

    df = result.get(
        "ampm_df",
        pd.DataFrame(),
    )

    # 上午、下午一律固定使用 29
    am_target = QC_TARGET_EFFICIENCY
    pm_target = QC_TARGET_EFFICIENCY

    if df is None or df.empty:
        st.warning(
            "已計算完成，但沒有產出可顯示的資料，"
            "ampm_df 為空。"
        )
        return

    if "時段" not in df.columns:
        st.error(
            "資料缺少「時段」欄位，"
            "無法區分 AM 與 PM 班別。"
        )
        return

    if "效率" not in df.columns:
        st.error(
            "資料缺少「效率」欄位，"
            "無法進行達標判斷。"
        )
        return

    am_df, pm_df = _split_am_pm(df)

    # ======================
    # 匯出 Excel
    # ======================
    if result.get("xlsx_bytes"):
        download_excel_card(
            result["xlsx_bytes"],
            result.get(
                "xlsx_name",
                "驗收作業KPI.xlsx",
            ),
            label=(
                "⬇️ 匯出 KPI 報表（Excel）"
                "｜低於 29 顯示紅色"
            ),
        )

    # ======================
    # AM / PM 兩欄呈現
    # ======================
    col_l, col_r = st.columns(
        2,
        gap="large",
    )

    with col_l:
        _render_shift_block(
            "🌓 AM 班（驗收）",
            am_df,
            top_n=top_n,
            target=am_target,
        )

    with col_r:
        _render_shift_block(
            "🌙 PM 班（驗收）",
            pm_df,
            top_n=top_n,
            target=pm_target,
        )


if __name__ == "__main__":
    main()
