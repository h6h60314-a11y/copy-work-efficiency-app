# pages/29_各時段作業效率.py
# -*- coding: utf-8 -*-
import io
import os
from io import StringIO
from datetime import datetime, date
from zoneinfo import ZoneInfo

import numpy as np
import pandas as pd
import streamlit as st
import altair as alt

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.formatting.rule import FormulaRule

# ---- 套用平台風格（有就用，沒有就退回原生）----
try:
    from common_ui import inject_logistics_theme, set_page, card_open, card_close
    HAS_COMMON_UI = True
except Exception:
    HAS_COMMON_UI = False

TPE = ZoneInfo("Asia/Taipei")

STATUS_PASS = "達標"
STATUS_FAIL = "未達標"
STATUS_NA = "未判斷"

# ✅ 特殊工時（分鐘）：12點、13點只有 30 分鐘
WORK_MINUTES_BY_HOUR = {12: 30, 13: 30}


# =============================
# 欄位整理 / 檢查
# =============================
def _norm_cols(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = [str(c).strip() for c in df.columns]
    return df


def require_columns(df: pd.DataFrame, required: list, label: str):
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(f"{label} 缺少欄位：{missing}\n目前欄位：{list(df.columns)}")


def clean_line(series: pd.Series) -> pd.Series:
    return series.astype(str).str.strip()


def clean_zone_1to4(series: pd.Series) -> pd.Series:
    z = pd.to_numeric(series, errors="coerce").astype("Int64")
    return z.where(z.between(1, 4, inclusive="both"))


def _safe_time(s: str) -> str:
    s = str(s).strip()
    if not s:
        return "08:00"
    try:
        datetime.strptime(s, "%H:%M")
        return s
    except Exception:
        return "08:00"


def _bytes_sig(b: bytes) -> str:
    if b is None:
        return "0"
    n = len(b)
    head = b[:128]
    tail = b[-128:] if n >= 128 else b
    return f"{n}-{hash(head)}-{hash(tail)}"


def _slot_minutes(hour: int) -> int:
    return int(WORK_MINUTES_BY_HOUR.get(int(hour), 60))


# =============================
# ✅ 讀檔：CSV/TSV/Excel（含 .xls 假檔）強韌讀取
#   - 你的 .xls 其實是 TSV：檔頭像 b'BOXID\\tOR...'
#   - 修正：python engine 不支援 low_memory → 不再傳入
# =============================
def _looks_like_html(raw: bytes) -> bool:
    head = raw[:4096].lstrip().lower()
    return (b"<html" in head) or (b"<!doctype" in head) or (b"<table" in head)


def _read_html_table(raw: bytes) -> pd.DataFrame:
    for enc in ["utf-8-sig", "utf-8", "cp950", "big5", "ms950", "latin1", "gb18030"]:
        try:
            text = raw.decode(enc, errors="ignore")
            tables = pd.read_html(StringIO(text))
            tables = [t for t in tables if isinstance(t, pd.DataFrame) and t.shape[1] >= 2]
            if not tables:
                continue
            return max(tables, key=lambda x: x.shape[0] * x.shape[1])
        except Exception:
            continue
    raise ValueError("偵測為 HTML 但解析表格失敗（read_html 失敗）。")


def _read_csv_guess(raw: bytes) -> pd.DataFrame:
    encodings = ["utf-8-sig", "utf-8", "cp950", "big5", "ms950", "gb18030", "latin1"]
    seps = ["\t", ",", ";", "|"]  # ✅ tab 放第一優先
    last_err = None

    for enc in encodings:
        for sep in seps:
            # 先用 C engine（最快、最穩），失敗再換 python
            for engine in ["c", "python"]:
                try:
                    kwargs = dict(encoding=enc, sep=sep, engine=engine)

                    # ✅ low_memory 只有 C engine 才能用；python engine 不能帶
                    if engine == "c":
                        kwargs["low_memory"] = False

                    df = pd.read_csv(io.BytesIO(raw), **kwargs)
                    if df.shape[1] <= 1:
                        continue
                    return df
                except Exception as e:
                    last_err = e

    # 最後：讓 pandas 自己猜分隔符（sep=None 只能用 python）
    try:
        text = raw.decode("utf-8", errors="replace")
        df = pd.read_csv(StringIO(text), sep=None, engine="python")
        if df.shape[1] <= 1:
            raise ValueError("偵測不到有效分隔符，請確認檔案內容。")
        return df
    except Exception as e:
        raise ValueError(f"讀取為 CSV/TSV 失敗：{last_err} / 最終：{e}")


def read_table_robust(file_name: str, raw: bytes, label: str = "檔案") -> pd.DataFrame:
    ext = os.path.splitext(file_name)[1].lower()
    head = raw[:2048]
    looks_tsv = (b"\t" in head) and (head.count(b"\t") >= 2)

    # ---- xlsx 類 ----
    if ext in (".xlsx", ".xlsm", ".xltx", ".xltm"):
        try:
            return pd.read_excel(io.BytesIO(raw), engine="openpyxl")
        except Exception:
            # 有些被改副檔名 → 當文字檔讀
            if _looks_like_html(raw):
                return _read_html_table(raw)
            return _read_csv_guess(raw)

    # ---- .xls：先當文字檔（你這種最常見）----
    if ext == ".xls":
        # 1) 優先 TSV/CSV（你的檔頭 BOXID\tOR...）
        try:
            if looks_tsv:
                for enc in ["utf-8-sig", "utf-8", "cp950", "big5", "ms950", "latin1", "gb18030"]:
                    try:
                        df = pd.read_csv(io.BytesIO(raw), sep="\t", encoding=enc, engine="c", low_memory=False)
                        if df.shape[1] > 1:
                            return df
                    except Exception:
                        pass
            # 不是明顯 TSV 或上面失敗 → 交給猜測器
            return _read_csv_guess(raw)
        except Exception as e_text:
            # 2) 再嘗試真正 xls（需要 xlrd，且檔案真的是 BIFF xls）
            try:
                import xlrd  # noqa: F401
                return pd.read_excel(io.BytesIO(raw), engine="xlrd")
            except Exception as e_xls:
                if _looks_like_html(raw):
                    try:
                        return _read_html_table(raw)
                    except Exception:
                        pass
                raise ValueError(
                    f"{label} 讀取 .xls 失敗：\n"
                    f"- 文字檔(TSV/CSV) 解析失敗：{e_text}\n"
                    f"- Excel(xlrd) 解析失敗：{e_xls}\n"
                    f"（此檔很可能是 TSV 文字檔，只是副檔名叫 .xls）"
                )

    # ---- 其他：先當文字檔 ----
    if _looks_like_html(raw):
        return _read_html_table(raw)
    return _read_csv_guess(raw)


# =============================
# KPI 計數（某小時）
# =============================
def _kpi_counts(dist_df: pd.DataFrame):
    if dist_df is None or dist_df.empty:
        return 0, 0, None
    p = int(dist_df.loc[dist_df["狀態"] == STATUS_PASS, "count"].sum())
    f = int(dist_df.loc[dist_df["狀態"] == STATUS_FAIL, "count"].sum())
    rate = (p / (p + f) * 100.0) if (p + f) > 0 else None
    return p, f, rate


# =============================
# Heatmap（Streamlit 顯示用）
# =============================
def render_hourly_heatmap(df_line_hourly: pd.DataFrame, hour_cols, title: str):
    if df_line_hourly is None or df_line_hourly.empty:
        st.info("沒有可呈現的圖。")
        return

    hour_cols = [int(h) for h in list(hour_cols)]
    plot = df_line_hourly.copy()

    plot["段數"] = pd.to_numeric(plot["段數"], errors="coerce").fillna(0).astype(int)
    plot["小時"] = pd.to_numeric(plot["小時"], errors="coerce").fillna(0).astype(int)
    plot["當小時加權PCS"] = pd.to_numeric(plot["當小時加權PCS"], errors="coerce").fillna(0.0)
    plot["本小時目標"] = pd.to_numeric(plot["本小時目標"], errors="coerce").fillna(0.0)

    plot["label"] = plot["段數"].astype(str) + "段｜" + plot["姓名"].astype(str)
    plot["狀態_色"] = plot["狀態"].fillna(STATUS_NA)
    plot["顯示量"] = plot["當小時加權PCS"].apply(lambda x: "" if abs(float(x)) < 1e-12 else f"{float(x):.2f}")

    order = (
        plot[["label", "段數", "姓名"]]
        .drop_duplicates()
        .sort_values(["段數", "姓名"])["label"]
        .tolist()
    )

    color_enc = alt.Color(
        "狀態_色:N",
        scale=alt.Scale(domain=[STATUS_PASS, STATUS_FAIL, STATUS_NA], range=["#2E7D32", "#C62828", "#D0D5DD"]),
        legend=alt.Legend(title="狀態"),
    )

    base = alt.Chart(plot).encode(
        x=alt.X("小時:O", sort=[str(h) for h in hour_cols], title="每小時"),
        y=alt.Y("label:N", sort=order, title="段數｜姓名"),
        tooltip=[
            alt.Tooltip("線別:N", title="線別"),
            alt.Tooltip("label:N", title="段數｜姓名"),
            alt.Tooltip("小時:O", title="小時"),
            alt.Tooltip("當小時加權PCS:Q", title="當小時加權PCS", format=",.4f"),
            alt.Tooltip("本小時目標:Q", title="本小時目標", format=",.2f"),
            alt.Tooltip("狀態:N", title="狀態"),
        ],
    )

    rect = base.mark_rect(cornerRadius=4).encode(color=color_enc)
    text = base.mark_text(fontSize=12, fontWeight=900).encode(text="顯示量:N")

    n_rows = max(1, plot["label"].nunique())
    height = min(42 * n_rows + 80, 900)
    st.altair_chart((rect + text).properties(title=title, height=height), use_container_width=True)


# =============================
# ✅ Excel：保留公式 + 色塊（條件格式）
# ✅ 不使用 LET()（舊 Excel 也能算）
# =============================
def build_excel_bytes_with_formulas_and_colors(
    detail_df: pd.DataFrame,
    roster_df: pd.DataFrame,
    hour_cols: list[int],
    target_hr: float,
    now_h: int,
    now_m: int,
) -> bytes:
    wb = Workbook()
    ws_detail = wb.active
    ws_detail.title = "完整明細_去重後"
    ws_mat = wb.create_sheet("時段量體_公式")
    ws_param = wb.create_sheet("參數")

    # 參數表
    ws_param["A1"] = "now_h"; ws_param["B1"] = int(now_h)
    ws_param["A2"] = "now_m"; ws_param["B2"] = int(now_m)
    ws_param["A3"] = "target_hr"; ws_param["B3"] = float(target_hr)
    for r in range(1, 4):
        ws_param[f"A{r}"].font = Font(bold=True)

    # Sheet1：完整明細
    cols = list(detail_df.columns)
    for c_idx, col in enumerate(cols, start=1):
        ws_detail.cell(row=1, column=c_idx, value=col).font = Font(bold=True)

    col_pack = cols.index("PACKQTY") + 1 if "PACKQTY" in cols else None
    col_w = cols.index("Cweight") + 1 if "Cweight" in cols else None
    col_aw = cols.index("加權PCS") + 1 if "加權PCS" in cols else None

    for r_idx, row in enumerate(detail_df.itertuples(index=False), start=2):
        for c_idx, col in enumerate(cols, start=1):
            v = getattr(row, col) if hasattr(row, col) else None
            ws_detail.cell(row=r_idx, column=c_idx, value=v)

        if col_pack and col_w and col_aw:
            p_cell = f"{get_column_letter(col_pack)}{r_idx}"
            w_cell = f"{get_column_letter(col_w)}{r_idx}"
            ws_detail.cell(row=r_idx, column=col_aw, value=f"={p_cell}*{w_cell}")
            ws_detail.cell(row=r_idx, column=col_aw).number_format = "0.0000"

    # SUMIFS 定位
    detail_header_to_col = {ws_detail.cell(row=1, column=i).value: i for i in range(1, ws_detail.max_column + 1)}
    need = ["線別", "段數", "小時", "加權PCS", "納入計算"]
    for k in need:
        if k not in detail_header_to_col:
            raise ValueError(f"明細缺少欄位「{k}」，無法建立 SUMIFS 公式。")

    d_line = get_column_letter(detail_header_to_col["線別"])
    d_zone = get_column_letter(detail_header_to_col["段數"])
    d_hour = get_column_letter(detail_header_to_col["小時"])
    d_aw = get_column_letter(detail_header_to_col["加權PCS"])
    d_in = get_column_letter(detail_header_to_col["納入計算"])
    d_first, d_last = 2, ws_detail.max_row

    # Sheet2：時段量體（每小時：量體/目標/狀態）
    base_cols = ["線別", "段數", "姓名", "開始時間"]
    hour_cols = [int(h) for h in hour_cols]

    headers = base_cols[:]
    for h in hour_cols:
        headers += [str(h), f"{h}_目標", f"{h}_狀態"]
    headers += ["加總", "加總目標", "加總狀態"]

    for c_idx, h in enumerate(headers, start=1):
        ws_mat.cell(row=1, column=c_idx, value=h).font = Font(bold=True)

    fill_ok = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fill_ng = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    fill_na = PatternFill(start_color="F2F4F7", end_color="F2F4F7", fill_type="solid")

    now_h_cell = "參數!$B$1"
    now_m_cell = "參數!$B$2"
    target_hr_cell = "參數!$B$3"

    for r_idx, row in enumerate(roster_df.itertuples(index=False), start=2):
        ws_mat.cell(row=r_idx, column=1, value=row.線別)
        ws_mat.cell(row=r_idx, column=2, value=int(row.段數))
        ws_mat.cell(row=r_idx, column=3, value=str(row.姓名))

        hh, mm = str(row.開始時間).split(":")
        ws_mat.cell(row=r_idx, column=4, value=f"=TIME({int(hh)},{int(mm)},0)")

        start_time_cell = f"$D{r_idx}"

        col_ptr = 5
        sum_cells = []
        tgt_cells = []

        for h in hour_cols:
            vol_col = col_ptr
            tgt_col = col_ptr + 1
            st_col = col_ptr + 2

            vol_cell = f"{get_column_letter(vol_col)}{r_idx}"
            tgt_cell = f"{get_column_letter(tgt_col)}{r_idx}"

            line_cell = f"$A{r_idx}"
            zone_cell = f"$B{r_idx}"
            vol_formula = (
                f'=SUMIFS('
                f'\'{ws_detail.title}\'!${d_aw}${d_first}:${d_aw}${d_last},'
                f'\'{ws_detail.title}\'!${d_line}${d_first}:${d_line}${d_last},{line_cell},'
                f'\'{ws_detail.title}\'!${d_zone}${d_first}:${d_zone}${d_last},{zone_cell},'
                f'\'{ws_detail.title}\'!${d_hour}${d_first}:${d_hour}${d_last},{h},'
                f'\'{ws_detail.title}\'!${d_in}${d_first}:${d_in}${d_last},TRUE)'
            )
            ws_mat.cell(row=r_idx, column=vol_col, value=vol_formula)
            ws_mat.cell(row=r_idx, column=vol_col).number_format = "0.0000"

            slot = f'IF(OR({h}=12,{h}=13),30,60)'
            endm = f'IF({h}={now_h_cell},MIN({now_m_cell},{slot}),{slot})'
            sh = f'HOUR({start_time_cell})'
            sm = f'MINUTE({start_time_cell})'
            mins = f'IF({h}>{now_h_cell},0,IF({h}<{sh},0,IF({h}={sh},MAX(0,{endm}-{sm}),{endm})))'
            tgt_formula = f'={target_hr_cell}*({mins})/60'
            ws_mat.cell(row=r_idx, column=tgt_col, value=tgt_formula)
            ws_mat.cell(row=r_idx, column=tgt_col).number_format = "0.0000"

            st_formula = f'=IF({tgt_cell}<=0,"",IF({vol_cell}>={tgt_cell},"{STATUS_PASS}","{STATUS_FAIL}"))'
            ws_mat.cell(row=r_idx, column=st_col, value=st_formula)

            sum_cells.append(vol_cell)
            tgt_cells.append(tgt_cell)
            col_ptr += 3

        sum_col = col_ptr
        sum_tgt_col = col_ptr + 1
        sum_st_col = col_ptr + 2

        sum_cell = f"{get_column_letter(sum_col)}{r_idx}"
        sum_tgt_cell = f"{get_column_letter(sum_tgt_col)}{r_idx}"

        ws_mat.cell(row=r_idx, column=sum_col, value=f"=SUM({','.join(sum_cells)})")
        ws_mat.cell(row=r_idx, column=sum_col).number_format = "0.0000"

        ws_mat.cell(row=r_idx, column=sum_tgt_col, value=f"=SUM({','.join(tgt_cells)})")
        ws_mat.cell(row=r_idx, column=sum_tgt_col).number_format = "0.0000"

        ws_mat.cell(
            row=r_idx,
            column=sum_st_col,
            value=f'=IF({sum_tgt_cell}<=0,"",IF({sum_cell}>={sum_tgt_cell},"{STATUS_PASS}","{STATUS_FAIL}"))',
        )

    # 欄寬 / 對齊
    ws_mat.column_dimensions["A"].width = 10
    ws_mat.column_dimensions["B"].width = 6
    ws_mat.column_dimensions["C"].width = 14
    ws_mat.column_dimensions["D"].width = 10

    for row in ws_mat.iter_rows(min_row=1, max_row=ws_mat.max_row, min_col=1, max_col=ws_mat.max_column):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # 隱藏目標/狀態欄，只留量體 + 加總
    start_col = 5
    for i, _h in enumerate(hour_cols):
        vol_col = start_col + i * 3
        tgt_col = vol_col + 1
        st_col = vol_col + 2
        ws_mat.column_dimensions[get_column_letter(tgt_col)].hidden = True
        ws_mat.column_dimensions[get_column_letter(st_col)].hidden = True
        ws_mat.column_dimensions[get_column_letter(vol_col)].width = 10

    sum_col = start_col + len(hour_cols) * 3
    sum_tgt_col = sum_col + 1
    sum_st_col = sum_col + 2
    ws_mat.column_dimensions[get_column_letter(sum_col)].width = 12
    ws_mat.column_dimensions[get_column_letter(sum_tgt_col)].hidden = True
    ws_mat.column_dimensions[get_column_letter(sum_st_col)].hidden = True

    # ✅ 條件格式：依隱藏狀態欄 → 把量體欄上色
    max_r = ws_mat.max_row
    for i, _h in enumerate(hour_cols):
        vol_col = start_col + i * 3
        st_col = vol_col + 2

        vol_letter = get_column_letter(vol_col)
        st_letter = get_column_letter(st_col)
        rng = f"{vol_letter}2:{vol_letter}{max_r}"

        ws_mat.conditional_formatting.add(
            rng,
            FormulaRule(formula=[f'${st_letter}2="{STATUS_PASS}"'], fill=fill_ok, stopIfTrue=True),
        )
        ws_mat.conditional_formatting.add(
            rng,
            FormulaRule(formula=[f'${st_letter}2="{STATUS_FAIL}"'], fill=fill_ng, stopIfTrue=True),
        )
        ws_mat.conditional_formatting.add(
            rng,
            FormulaRule(formula=[f'${st_letter}2=""'], fill=fill_na, stopIfTrue=True),
        )

    # 加總欄也上色
    sum_letter = get_column_letter(sum_col)
    sum_st_letter = get_column_letter(sum_st_col)
    sum_rng = f"{sum_letter}2:{sum_letter}{max_r}"

    ws_mat.conditional_formatting.add(
        sum_rng,
        FormulaRule(formula=[f'${sum_st_letter}2="{STATUS_PASS}"'], fill=fill_ok, stopIfTrue=True),
    )
    ws_mat.conditional_formatting.add(
        sum_rng,
        FormulaRule(formula=[f'${sum_st_letter}2="{STATUS_FAIL}"'], fill=fill_ng, stopIfTrue=True),
    )
    ws_mat.conditional_formatting.add(
        sum_rng,
        FormulaRule(formula=[f'${sum_st_letter}2=""'], fill=fill_na, stopIfTrue=True),
    )

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def main():
    st.set_page_config(page_title="大豐物流 - 出貨課｜各時段作業效率", page_icon="⏱️", layout="wide")
    if HAS_COMMON_UI:
        inject_logistics_theme()
        set_page("📦 出貨課", "⏱️ 29｜各時段作業效率")

    st.markdown("### ⏱️ 各時段作業效率（Excel：保留公式＋色塊自動更新；支援舊 Excel）")

    fixed_time_map = {
        "范明俊": "08:00", "阮玉名": "08:00", "李茂銓": "08:00", "河文強": "08:00",
        "蔡麗珠": "08:00", "潘文一": "08:00", "阮伊黃": "08:00", "葉欲弘": "09:00",
        "阮武玉玄": "08:00", "吳黃金珠": "08:30", "潘氏青江": "08:00", "陳國慶": "08:30",
        "楊心如": "08:00", "阮瑞美黃緣": "08:00", "周芸蓁": "08:00", "黎氏瓊": "08:00",
        "王文楷": "08:30", "潘氏慶平": "08:00", "阮氏美麗": "08:00", "岳子恆": "08:30",
        "郭雙燕": "08:30", "阮孟勇": "08:00", "廖永成": "08:30", "楊浩傑": "08:30",
        "黃日康": "08:30", "蔣金妮": "08:30", "柴家欣": "08:30", "邱思捷": "09:00",
        "王建成": "09:00",
    }

    with st.sidebar:
        st.markdown("### 設定")
        target_hr = st.number_input("每小時目標（加權PCS/小時）", min_value=1.0, value=790.0, step=10.0)
        hour_min = st.number_input("起始小時", min_value=0, max_value=23, value=8, step=1)

        use_now = st.toggle("用現在時間作為判斷截止（台北時間）", value=True)
        if use_now:
            now = datetime.now(TPE)
        else:
            t_in = st.time_input("判斷截止時間（台北時間）", value=datetime.now(TPE).time())
            now = datetime.combine(date.today(), t_in).replace(tzinfo=TPE)

        st.caption(f"目前採用時間：{now.strftime('%Y-%m-%d %H:%M:%S')} (Asia/Taipei)")
        auto_calc = st.toggle("上傳/設定變更後自動更新", value=True)

    c1, c2 = st.columns(2)
    with c1:
        prod_file = st.file_uploader("① 上傳『原始生產資料』(CSV/Excel)", type=["csv", "xlsx", "xlsm", "xls"])
    with c2:
        mem_file = st.file_uploader("② 上傳『人員名單』(CSV/Excel)", type=["csv", "xlsx", "xlsm", "xls"])

    manual = st.button("🚀 立即更新/重算", type="primary", use_container_width=True)

    if prod_file is None or mem_file is None:
        st.info("請先上傳兩個檔案：生產資料 + 人員名單。")
        return

    prod_sig = _bytes_sig(prod_file.getvalue())
    mem_sig = _bytes_sig(mem_file.getvalue())
    settings_sig = f"{target_hr}-{hour_min}-{use_now}-{now.hour}-{now.minute}"

    last = st.session_state.get("_29_last_sig", None)
    cur_sig = (prod_sig, mem_sig, settings_sig)
    should_run = manual or (auto_calc and (last != cur_sig))
    if not should_run:
        st.caption("（目前結果已是最新；如有更新檔案/設定會自動同步）")
        return
    st.session_state["_29_last_sig"] = cur_sig

    try:
        # 人員名單
        df_mem_raw = _norm_cols(read_table_robust(mem_file.name, mem_file.getvalue(), label="人員名單檔案"))

        line_col_candidates = ["LINEID", "線別", "LineID", "LINE Id", "Line Id"]
        line_col = next((c for c in line_col_candidates if c in df_mem_raw.columns), None)
        if line_col is None:
            raise ValueError("人員名單找不到線別欄位（需要 LINEID 或 線別）。")

        seg_cols = {1: "第一段", 2: "第二段", 3: "第三段", 4: "第四段"}
        for _, colname in seg_cols.items():
            if colname not in df_mem_raw.columns:
                raise ValueError(f"人員名單缺少欄位：{colname}（需要 第一段～第四段）")

        member_list = []
        for _, row in df_mem_raw.iterrows():
            line_id = str(row.get(line_col, "")).strip()
            if not line_id or line_id.lower() == "nan":
                continue
            for zid, colname in seg_cols.items():
                name = row.get(colname, None)
                if pd.notna(name) and str(name).strip() != "":
                    n_str = str(name).strip()
                    st_time = _safe_time(fixed_time_map.get(n_str, "08:00"))
                    member_list.append({"線別": line_id, "段數": zid, "姓名": n_str, "開始時間": st_time})

        roster_df = pd.DataFrame(member_list)
        if roster_df.empty:
            raise ValueError("人員名單解析後為空：請確認 第一段～第四段 內有姓名。")

        roster_df["線別"] = clean_line(roster_df["線別"])
        roster_df["段數"] = clean_zone_1to4(roster_df["段數"])
        roster_df = roster_df[roster_df["段數"].notna()].copy()
        roster_df = roster_df.drop_duplicates(["線別", "段數"], keep="first").copy()
        roster_df = roster_df[["線別", "段數", "姓名", "開始時間"]].copy()

        # 生產資料（✅ 這裡已支援 .xls 假檔 TSV）
        df_raw = read_table_robust(prod_file.name, prod_file.getvalue(), label="生產資料檔案")
        df_raw = _norm_cols(df_raw)  # ✅ 避免欄位尾巴空白

        require_columns(df_raw, ["PICKDATE", "LINEID", "ZONEID", "PACKQTY", "Cweight"], "生產資料檔案")

        df_raw["PICKDATE"] = pd.to_datetime(df_raw["PICKDATE"], errors="coerce")
        df_raw = df_raw[df_raw["PICKDATE"].notna()].copy()

        df_raw = df_raw.rename(columns={"LINEID": "線別", "ZONEID": "段數"})
        df_raw["線別"] = clean_line(df_raw["線別"])
        df_raw["段數"] = clean_zone_1to4(df_raw["段數"])
        df_raw = df_raw[df_raw["段數"].notna()].copy()

        df_raw["PACKQTY"] = pd.to_numeric(df_raw["PACKQTY"], errors="coerce").fillna(0)
        df_raw["Cweight"] = pd.to_numeric(df_raw["Cweight"], errors="coerce").fillna(0)

        # 去重
        rid_cols = [c for c in df_raw.columns if c not in ("__rid",)]
        df_raw["__rid"] = pd.util.hash_pandas_object(df_raw[rid_cols], index=False)
        df_raw = df_raw.drop_duplicates("__rid", keep="first").copy()

        df = pd.merge(df_raw, roster_df, on=["線別", "段數"], how="left", validate="m:1")
        df["姓名"] = df["姓名"].fillna("未設定")
        df["開始時間"] = df["開始時間"].fillna("08:00").map(_safe_time)

        df["小時"] = df["PICKDATE"].dt.hour
        df["PICK_MIN"] = df["PICKDATE"].dt.hour * 60 + df["PICKDATE"].dt.minute

        st_parts = df["開始時間"].astype(str).str.split(":", n=1, expand=True)
        st_h = pd.to_numeric(st_parts[0], errors="coerce").fillna(8).astype(int)
        st_m = pd.to_numeric(st_parts[1], errors="coerce").fillna(0).astype(int)
        df["開始分鐘"] = st_h * 60 + st_m

        df["納入計算"] = df["PICK_MIN"] >= df["開始分鐘"]
        df["排除原因"] = np.where(df["納入計算"], "", "早於開始時間")

        df["加權PCS"] = df["PACKQTY"] * df["Cweight"]

        # Streamlit 顯示
        cur_h, cur_m = now.hour, now.minute
        hour_cols = list(range(int(hour_min), int(cur_h) + 1)) if int(cur_h) >= int(hour_min) else [int(cur_h)]
        base_cols = ["線別", "段數", "姓名", "開始時間"]

        df_calc = df[df["納入計算"]].copy()
        hourly_sum = df_calc.groupby(base_cols + ["小時"], as_index=False)["加權PCS"].sum()
        hourly_sum = hourly_sum.rename(columns={"加權PCS": "當小時加權PCS"})

        keys = roster_df[base_cols].drop_duplicates().copy()
        grid_hours = keys.assign(_k=1).merge(pd.DataFrame({"小時": hour_cols, "_k": 1}), on="_k").drop(columns=["_k"])
        hourly_full = grid_hours.merge(hourly_sum, on=base_cols + ["小時"], how="left")
        hourly_full["當小時加權PCS"] = pd.to_numeric(hourly_full["當小時加權PCS"], errors="coerce").fillna(0.0)

        parts = hourly_full["開始時間"].astype(str).str.split(":", n=1, expand=True)
        s_h = pd.to_numeric(parts[0], errors="coerce").fillna(8).astype(int)
        s_m = pd.to_numeric(parts[1], errors="coerce").fillna(0).astype(int)
        hh = pd.to_numeric(hourly_full["小時"], errors="coerce").fillna(0).astype(int)
        slot = hh.map(lambda x: _slot_minutes(int(x))).astype(int)
        end_m = np.where(hh == cur_h, np.minimum(cur_m, slot), slot).astype(int)

        minutes_worked = np.where(
            hh > cur_h, 0,
            np.where(
                hh < s_h, 0,
                np.where(
                    hh == s_h, np.maximum(0, end_m - s_m),
                    end_m
                )
            )
        ).astype(float)

        hourly_full["本小時有效分鐘"] = minutes_worked
        hourly_full["本小時目標"] = (minutes_worked / 60.0) * float(target_hr)
        hourly_full["狀態"] = np.where(
            hourly_full["本小時有效分鐘"] <= 0,
            None,
            np.where(hourly_full["當小時加權PCS"] >= hourly_full["本小時目標"], STATUS_PASS, STATUS_FAIL)
        )

        dist = (
            hourly_full[hourly_full["狀態"].isin([STATUS_PASS, STATUS_FAIL])]
            .groupby(["線別", "小時", "狀態"], as_index=False)
            .size()
            .rename(columns={"size": "count"})
        )

        st.success("計算完成 ✅（Excel：公式＋色塊會自動更新）")

        eff_hour = int(cur_h)
        lines = sorted(keys["線別"].dropna().unique().tolist())
        for line in lines:
            if HAS_COMMON_UI:
                card_open(f"📦 {line}")
            else:
                st.markdown(f"### 📦 {line}")

            dist_now = dist[(dist["線別"] == line) & (dist["小時"] == eff_hour)]
            p, f, rate = _kpi_counts(dist_now)

            a, b, c, d = st.columns(4)
            a.metric("判斷小時", f"{eff_hour} 點")
            b.metric("達標 段數", p)
            c.metric("未達標 段數", f)
            d.metric("達標 率", (f"{rate:.1f}%" if rate is not None else "—"))

            df_line = hourly_full[hourly_full["線別"] == line][
                ["線別", "段數", "姓名", "小時", "當小時加權PCS", "本小時目標", "狀態"]
            ].copy()
            render_hourly_heatmap(df_line, hour_cols, title=f"{line}｜每小時（12/13=30分）")

            if HAS_COMMON_UI:
                card_close()

        # 明細輸出（Excel 會用公式重算加權PCS）
        detail_df = df.copy().sort_values(["線別", "段數", "PICKDATE"]).reset_index(drop=True)
        if "加權PCS" not in detail_df.columns:
            detail_df["加權PCS"] = np.nan

        xlsx_bytes = build_excel_bytes_with_formulas_and_colors(
            detail_df=detail_df,
            roster_df=roster_df,
            hour_cols=hour_cols,
            target_hr=float(target_hr),
            now_h=int(cur_h),
            now_m=int(cur_m),
        )
        filename = f"產能時段_公式_色塊_{datetime.now(TPE).strftime('%H%M')}.xlsx"
        st.download_button(
            "⬇️ 下載 Excel（保留公式＋色塊自動變色）",
            data=xlsx_bytes,
            file_name=filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

    except Exception as e:
        st.error(f"發生錯誤：{e}")


if __name__ == "__main__":
    main()
