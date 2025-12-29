# -*- coding: utf-8 -*-
import os
import re
import math
import hashlib
from io import BytesIO

import numpy as np
import pandas as pd
import streamlit as st

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import Rule
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.chart.label import DataLabelList

from common_ui import inject_logistics_theme, set_page, card_open, card_close


# =========================
# 設定
# =========================
TARGET_PER_MANHOUR_DEFAULT = 790
AM_HOURS = list(range(8, 13))     # 8-12
PM_HOURS = list(range(13, 19))    # 13-18

BASE_FONT_NAME = "微軟正黑體"
BASE_FONT_SIZE = 12
ROW_HEIGHT = 18

NUM_FMT_2_HIDE0 = "#,##0.00;-#,##0.00;;@"
NUM_FMT_4_HIDE0 = "#,##0.0000;-#,##0.0000;;@"
NUM_FMT_INT_HIDE0 = "#,##0;-#,##0;;@"
NUM_FMT_INT = "#,##0"
NUM_FMT_MONEY_HIDE0 = "#,##0;-#,##0;;@"

AM_DEFAULT_MONEY = 100
PM_DEFAULT_MONEY = 50

QCA_LIST = ["GT-B", "GT-C", "GT-D", "GT-E"]
QCB_LIST = ["GT-A", "GT-J", "GT-K"]

OLE_HEADER = b"\xD0\xCF\x11\xE0\xA1\xB1\x1A\xE1"
ZIP_HEADER = b"PK\x03\x04"


# =========================
# robust reader（支援假xls / bytes）
# =========================
def _try_read_html(raw: bytes) -> pd.DataFrame:
    tables = pd.read_html(BytesIO(raw))
    if tables:
        return tables[0]
    raise ValueError("HTML 讀取不到表格")


def _try_read_text_like(raw: bytes) -> pd.DataFrame:
    for enc in ("utf-8-sig", "utf-8", "big5", "cp950", "latin1"):
        for sep in ("\t", ",", ";", "|"):
            try:
                df = pd.read_csv(BytesIO(raw), encoding=enc, sep=sep)
                if df.shape[1] >= 2:
                    return df
            except Exception:
                continue
    raise ValueError("不是可解析的文字分隔檔")


def robust_read_bytes(filename: str, raw: bytes) -> pd.DataFrame:
    ext = os.path.splitext(filename)[1].lower()
    head = raw[:8]
    is_ole = head.startswith(OLE_HEADER)
    is_zip = head.startswith(ZIP_HEADER)

    if is_zip or ext in (".xlsx", ".xlsm", ".xltx", ".xltm"):
        return pd.read_excel(BytesIO(raw), engine="openpyxl")

    if is_ole or ext in (".xls",):
        try:
            return pd.read_excel(BytesIO(raw), engine="xlrd")
        except Exception:
            try:
                return _try_read_html(raw)
            except Exception:
                return _try_read_text_like(raw)

    if ext == ".csv":
        for enc in ("utf-8-sig", "utf-8", "big5", "cp950", "latin1"):
            try:
                return pd.read_csv(BytesIO(raw), encoding=enc)
            except Exception:
                continue
        return pd.read_csv(BytesIO(raw), encoding="utf-8", errors="replace")

    try:
        return _try_read_html(raw)
    except Exception:
        return _try_read_text_like(raw)


# =========================
# Excel style helpers
# =========================
def _clone_font(cell_font: Font, *, name=None, size=None, bold=None, color=None):
    if cell_font is None:
        cell_font = Font()
    return Font(
        name=name if name is not None else cell_font.name,
        size=size if size is not None else cell_font.size,
        bold=bold if bold is not None else cell_font.bold,
        italic=cell_font.italic,
        vertAlign=cell_font.vertAlign,
        underline=cell_font.underline,
        strike=cell_font.strike,
        color=color if color is not None else cell_font.color,
        outline=cell_font.outline,
        shadow=cell_font.shadow,
        condense=cell_font.condense,
        extend=cell_font.extend,
        charset=cell_font.charset,
        family=cell_font.family,
        scheme=cell_font.scheme,
    )


def _set_base_font(cell, *, force_color=None, force_bold=None):
    cell.font = _clone_font(
        cell.font,
        name=BASE_FONT_NAME,
        size=BASE_FONT_SIZE,
        bold=force_bold if force_bold is not None else cell.font.bold,
        color=force_color if force_color is not None else cell.font.color,
    )


def _set_row_height(ws, r):
    ws.row_dimensions[r].height = ROW_HEIGHT


# =========================
# 欄位對照
# =========================
def normalize_columns(df: pd.DataFrame):
    df = df.copy()
    df.columns = [str(c).strip() for c in df.columns]
    colmap = {str(c).strip().upper(): c for c in df.columns}

    def pick(*cands):
        for k in cands:
            if k in colmap:
                return colmap[k]
        return None

    c_pickdate = pick("PICKDATE", "PICK_DATE", "PICK DATETIME", "PICKTIME", "PICK_TIME")
    c_packqty  = pick("PACKQTY", "PACK_QTY", "PCS", "QTY")
    c_cweight  = pick("CWEIGHT", "C_WEIGHT", "C WEIGHT", "WEIGHT")
    c_lineid   = pick("LINEID", "LINE_ID", "LINE", "LINE ID")
    c_stotype  = pick("STO_TYPE", "STOTYPE", "SO_TYPE", "TYPE")

    missing = [name for name, col in [
        ("PICKDATE", c_pickdate),
        ("PACKQTY",  c_packqty),
        ("Cweight",  c_cweight),
        ("LINEID",   c_lineid),
        ("STO_TYPE", c_stotype),
    ] if col is None]

    if missing:
        raise KeyError(f"缺少必要欄位：{missing}\n目前欄位：{list(df.columns)}")

    return df, c_pickdate, c_packqty, c_cweight, c_lineid, c_stotype


# =========================
# 產出每小時彙整資料
# =========================
def build_hourly_metrics(df, c_pickdate, c_packqty, c_cweight, c_lineid, c_stotype):
    df = df.copy()
    df[c_stotype] = df[c_stotype].astype(str).str.strip()
    df[c_lineid] = df[c_lineid].astype(str).str.strip()

    df[c_pickdate] = pd.to_datetime(df[c_pickdate], errors="coerce")
    df[c_packqty] = pd.to_numeric(df[c_packqty], errors="coerce").fillna(0)
    df[c_cweight] = pd.to_numeric(df[c_cweight], errors="coerce").fillna(0)

    df["加權PCS"] = df[c_packqty] * df[c_cweight]
    df["PICK_HOUR"] = df[c_pickdate].dt.floor("h")
    df["PICK_DATE"] = df["PICK_HOUR"].dt.date
    df["HOUR"] = df["PICK_HOUR"].dt.hour

    line_base = (df.groupby(["PICK_DATE", c_lineid, "HOUR"])[[c_packqty, "加權PCS"]]
                   .sum()
                   .reset_index()
                   .rename(columns={c_packqty: "PCS"}))

    split = (df.groupby(["PICK_DATE", c_lineid, "HOUR", c_stotype])[[c_packqty, "加權PCS"]]
               .sum()
               .reset_index()
               .rename(columns={c_packqty: "PCS"}))

    return df, line_base, split


# =========================
# manpower helpers
# =========================
def _as_float(v):
    if v is None or (isinstance(v, str) and v.strip() == ""):
        return 0.0
    try:
        return float(v)
    except Exception:
        return 0.0


def _as_blank_if_zero(val):
    if val is None or (isinstance(val, float) and np.isnan(val)):
        return ""
    try:
        fv = float(val)
    except Exception:
        return ""
    if abs(fv) < 1e-12:
        return ""
    return fv


# =========================
# 寫入每小時戰情表（日期工作表）
# =========================
def write_hourly_sheet(
    wb, sheet_name, date_value, hours, lineids,
    line_base_map, split_map, manpower_map
):
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        wb.remove(ws)
    ws = wb.create_sheet(sheet_name)

    black_fill = PatternFill("solid", fgColor="111111")
    head_fill = PatternFill("solid", fgColor="F2F2F2")
    manpower_fill = PatternFill("solid", fgColor="FFF2CC")

    thin = Side(style="thin", color="D0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    right = Alignment(horizontal="right", vertical="center")

    _set_row_height(ws, 1)
    ws["A1"] = str(date_value)
    ws["A1"].alignment = center
    ws["A1"].fill = head_fill
    ws["A1"].border = border
    _set_base_font(ws["A1"], force_bold=True)

    for j, h in enumerate(hours, start=2):
        c = ws.cell(row=1, column=j, value=int(h))
        c.alignment = center
        c.fill = head_fill
        c.border = border
        _set_base_font(c, force_bold=True)

    _set_row_height(ws, 2)
    ws["A2"] = "撿貨（已撿數量）"
    ws["A2"].fill = black_fill
    ws["A2"].alignment = left
    ws["A2"].border = border
    _set_base_font(ws["A2"], force_bold=True, force_color="FFFFFF")

    for j in range(2, 2 + len(hours)):
        cell = ws.cell(row=2, column=j, value=None)
        cell.alignment = right
        cell.border = border
        cell.number_format = NUM_FMT_2_HIDE0
        _set_base_font(cell)

    r = 3
    pcs_weight_rows = []

    def write_row(label, values_by_hour=None, style_black=True, fill=None, is_manpower=False):
        nonlocal r
        _set_row_height(ws, r)

        a = ws.cell(row=r, column=1, value=label)
        a.fill = black_fill if style_black else (fill or None)
        a.alignment = left
        a.border = border
        if style_black:
            _set_base_font(a, force_bold=True, force_color="FFFFFF")
        else:
            _set_base_font(a, force_bold=True)

        for j, h in enumerate(hours, start=2):
            raw = None if values_by_hour is None else values_by_hour.get(int(h), None)

            if is_manpower:
                v = raw
                if v is None or (isinstance(v, float) and np.isnan(v)):
                    c = ws.cell(row=r, column=j, value="")
                else:
                    fv = float(v)
                    if abs(fv - round(fv)) < 1e-12:
                        c = ws.cell(row=r, column=j, value=int(round(fv)))
                    else:
                        c = ws.cell(row=r, column=j, value=fv)
                        c.number_format = "#,##0.##;-#,##0.##;;@"
            else:
                v = _as_blank_if_zero(raw)
                c = ws.cell(row=r, column=j, value=("" if v == "" else float(v)))
                if v != "":
                    c.number_format = NUM_FMT_2_HIDE0

            c.alignment = right
            c.border = border
            if fill is not None:
                c.fill = fill
                _set_base_font(c, force_bold=True)
            else:
                _set_base_font(c)

        row_idx = r
        r += 1
        return row_idx

    def write_formula_row(label, numerator_row, denom_row, number_format):
        nonlocal r
        _set_row_height(ws, r)

        a = ws.cell(row=r, column=1, value=label)
        a.fill = black_fill
        a.alignment = left
        a.border = border
        _set_base_font(a, force_bold=True, force_color="FFFFFF")

        for j in range(2, 2 + len(hours)):
            col = get_column_letter(j)
            num = f"{col}{numerator_row}"
            den = f"{col}{denom_row}"
            formula = f'=IF(OR({den}="",{den}=0),"",IF({num}/{den}=0,"",{num}/{den}))'
            c = ws.cell(row=r, column=j, value=formula)
            c.alignment = right
            c.border = border
            c.number_format = number_format
            _set_base_font(c)

        r += 1

    for lid in lineids:
        lid = str(lid)

        _set_row_height(ws, r)
        a = ws.cell(row=r, column=1, value=f"{lid} Line")
        a.fill = black_fill
        a.alignment = left
        a.border = border
        _set_base_font(a, force_bold=True, force_color="FFFFFF")
        for j in range(2, 2 + len(hours)):
            c = ws.cell(row=r, column=j, value=None)
            c.border = border
            c.alignment = Alignment(horizontal="right", vertical="center")
            _set_base_font(c)
        r += 1

        pcs_weight_map = line_base_map.get((lid, "加權PCS"), {})
        pcs_map = line_base_map.get((lid, "PCS"), {})

        gso_w = split_map.get((lid, "GSO", "加權PCS"), {})
        gxso_w = split_map.get((lid, "GXSO", "加權PCS"), {})
        gso = split_map.get((lid, "GSO", "PCS"), {})
        gxso = split_map.get((lid, "GXSO", "PCS"), {})

        row_pcs_w = write_row(f"{lid}（PCS）加權", pcs_weight_map, is_manpower=False)
        pcs_weight_rows.append(row_pcs_w)

        write_row(f"{lid}（PCS）", pcs_map, is_manpower=False)
        write_row("GSO(加權)", gso_w, is_manpower=False)
        write_row("GXSO(加權)", gxso_w, is_manpower=False)
        write_row("GSO", gso, is_manpower=False)
        write_row("GXSO", gxso, is_manpower=False)

        man_map = {int(h): manpower_map.get((date_value, lid, int(h)), np.nan) for h in hours}
        write_row(f"{lid}（人數）", man_map, fill=manpower_fill, is_manpower=True)

        # 這兩列會在彙總用到
        write_formula_row("平均產力(加權) 4", numerator_row=row_pcs_w, denom_row=row_pcs_w + 6, number_format=NUM_FMT_4_HIDE0)
        write_formula_row("平均產力(加權)", numerator_row=row_pcs_w, denom_row=row_pcs_w + 6, number_format=NUM_FMT_2_HIDE0)

        _set_row_height(ws, r)
        for j in range(1, 2 + len(hours)):
            c = ws.cell(row=r, column=j, value=None)
            c.border = border
            c.alignment = Alignment(horizontal="right", vertical="center") if j >= 2 else Alignment(horizontal="left", vertical="center")
            _set_base_font(c)
        r += 1

    # 撿貨(已撿數量) = 各 Line 的（PCS）加權加總
    for j in range(2, 2 + len(hours)):
        col = get_column_letter(j)
        refs = ",".join([f"{col}{rr}" for rr in pcs_weight_rows])
        c = ws.cell(row=2, column=j, value=f'=IF(SUM({refs})=0,"",SUM({refs}))')
        c.number_format = NUM_FMT_2_HIDE0
        c.alignment = Alignment(horizontal="right", vertical="center")
        _set_base_font(c)

    ws.column_dimensions["A"].width = 24
    for j in range(2, 2 + len(hours)):
        ws.column_dimensions[get_column_letter(j)].width = 12
    ws.freeze_panes = "B3"
    return ws


# =========================
# ✅ 達標計算
# =========================
def compute_achievers(date_value, lineids, hours, line_base_map, manpower_map, target_per_mh: int):
    am_ach, pm_ach = [], []

    for lid in lineids:
        lid = str(lid)

        am_man = sum(_as_float(manpower_map.get((date_value, lid, int(h)), 0)) for h in AM_HOURS if h in hours)
        pm_man = sum(_as_float(manpower_map.get((date_value, lid, int(h)), 0)) for h in PM_HOURS if h in hours)

        pcs_w_map = line_base_map.get((lid, "加權PCS"), {})
        am_pcs = sum(float(pcs_w_map.get(int(h), 0) or 0) for h in AM_HOURS if h in hours)
        pm_pcs = sum(float(pcs_w_map.get(int(h), 0) or 0) for h in PM_HOURS if h in hours)

        am_target = math.trunc(target_per_mh * am_man)
        pm_target = math.trunc(target_per_mh * pm_man)
        am_pcs_i = math.trunc(am_pcs)
        pm_pcs_i = math.trunc(pm_pcs)

        if am_man > 0 and am_pcs_i >= am_target:
            am_ach.append(lid)
        if pm_man > 0 and pm_pcs_i >= pm_target:
            pm_ach.append(lid)

    return am_ach, pm_ach


# =========================
# 彙總 sheet（沿用你原本邏輯）
# =========================
def _is_hour(v):
    try:
        iv = int(str(v).strip())
        return 0 <= iv <= 23
    except Exception:
        return False


def find_hour_header_row(ws):
    max_r = min(ws.max_row, 10)
    max_c = min(ws.max_column, 120)
    for r in range(1, max_r + 1):
        hour_to_col = {}
        for c in range(1, max_c + 1):
            v = ws.cell(r, c).value
            if _is_hour(v):
                hour_to_col[int(str(v).strip())] = c
        if len(hour_to_col) >= 3:
            return r, hour_to_col
    return None, {}


def parse_line_rows(ws):
    out = {}
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if not v:
            continue
        t = str(v).strip()

        m1 = re.match(r"(.+?)（PCS）加權$", t)
        if m1:
            lid = m1.group(1).strip()
            out.setdefault(lid, {})["pcs_w"] = r
            continue

        m2 = re.match(r"(.+?)（人數）$", t)
        if m2:
            lid = m2.group(1).strip()
            out.setdefault(lid, {})["man"] = r
            continue

    return {k: v for k, v in out.items() if "pcs_w" in v and "man" in v}


def sum_cells_formula(sheet, row, cols):
    if not cols:
        return "=0"
    refs = [f"'{sheet}'!{get_column_letter(c)}{row}" for c in cols]
    return f"=SUM({','.join(refs)})"


def countif_sum(range_a1: str, items: list[str]) -> str:
    return "+".join([f'COUNTIF({range_a1},"{it}")' for it in items]) if items else "0"


def build_summary_sheet_with_achievers(
    wb, summary_name, date_ws, line_map, am_cols, pm_cols,
    am_achievers, pm_achievers, target_per_mh: int
):
    if summary_name in wb.sheetnames:
        wb.remove(wb[summary_name])
    ws = wb.create_sheet(summary_name)

    thin = Side(style="thin", color="D0D0D0")
    thick = Side(style="medium", color="111111")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    box_border = Border(left=thick, right=thick, top=thick, bottom=thick)

    header_fill = PatternFill("solid", fgColor="F8CBAD")
    sub_header_fill = PatternFill("solid", fgColor="F2F2F2")

    center = Alignment(horizontal="center", vertical="center")
    right = Alignment(horizontal="right", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    headers = {
        1: "Line ID",
        2: "每小時目標(加權)",
        3: "上午人力",
        4: "上午應達成(加權)",
        5: "總PCS(加權)",
        6: "差異(加權)",
        7: "下午人力",
        8: "下午目標(加權)",
        9: "總PCS(加權)",
        10: "差異(加權)",
    }

    _set_row_height(ws, 1)
    for col, title in headers.items():
        cell = ws.cell(1, col, title)
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center
        _set_base_font(cell, force_bold=True)

    widths = {
        "A": 12, "B": 16, "C": 12, "D": 18, "E": 14, "F": 12,
        "G": 12, "H": 16, "I": 14, "J": 12,
        "K": 2,  "L": 14, "M": 18, "N": 18,
    }
    for k, v in widths.items():
        ws.column_dimensions[k].width = v

    r0 = 2
    src = date_ws.title

    def text_no_trailing_dot(expr: str):
        t = f'TEXT({expr},"#,##0.##")'
        return f'IF(RIGHT({t},1)=".",LEFT({t},LEN({t})-1),{t})'

    for i, (lid, rows) in enumerate(line_map.items()):
        r = r0 + i
        _set_row_height(ws, r)

        for col in range(1, 11):
            cell = ws.cell(r, col)
            cell.border = border
            cell.alignment = center if col in (1, 2, 3) else right
            _set_base_font(cell)

        ws.cell(r, 1, str(lid).strip())
        ws.cell(r, 2, int(target_per_mh)).number_format = NUM_FMT_INT

        man_row = rows["man"]
        pcsw_row = rows["pcs_w"]

        am_man_expr = sum_cells_formula(src, man_row, am_cols)[1:]
        am_pcs_expr = sum_cells_formula(src, pcsw_row, am_cols)[1:]
        pm_man_expr = sum_cells_formula(src, man_row, pm_cols)[1:]
        pm_pcs_expr = sum_cells_formula(src, pcsw_row, pm_cols)[1:]

        ws.cell(r, 3, f'=IF({am_man_expr}=0,"",{text_no_trailing_dot(am_man_expr)})').number_format = "@"
        ws.cell(r, 7, f'=IF({pm_man_expr}=0,"",{text_no_trailing_dot(pm_man_expr)})').number_format = "@"

        ws.cell(r, 4, f'=IF($C{r}="","",TRUNC($B{r}*VALUE(SUBSTITUTE($C{r},",","")),0))').number_format = NUM_FMT_INT_HIDE0
        ws.cell(r, 8, f'=IF($G{r}="","",TRUNC($B{r}*VALUE(SUBSTITUTE($G{r},",","")),0))').number_format = NUM_FMT_INT_HIDE0

        ws.cell(r, 5, f'=IF($C{r}="","",IF({am_pcs_expr}=0,"",TRUNC({am_pcs_expr},0)))').number_format = NUM_FMT_INT_HIDE0
        ws.cell(r, 9, f'=IF($G{r}="","",IF({pm_pcs_expr}=0,"",TRUNC({pm_pcs_expr},0)))').number_format = NUM_FMT_INT_HIDE0

        ws.cell(r, 6, f'=IF(OR($D{r}="",$E{r}=""),"",TRUNC($E{r}-$D{r},0))').number_format = NUM_FMT_INT_HIDE0
        ws.cell(r, 10, f'=IF(OR($H{r}="",$I{r}=""),"",TRUNC($I{r}-$H{r},0))').number_format = NUM_FMT_INT_HIDE0

    last_row = r0 + len(line_map) - 1
    if last_row < r0:
        last_row = r0

    # 條件格式：E / I 欄達標/未達標
    if last_row >= r0:
        red_fill = PatternFill(fill_type="solid", start_color="FFC7CE", end_color="FFC7CE")
        green_fill = PatternFill(fill_type="solid", start_color="C6EFCE", end_color="C6EFCE")
        dxf_red = DifferentialStyle(font=Font(color="9C0006"), fill=red_fill)
        dxf_green = DifferentialStyle(font=Font(color="006100"), fill=green_fill)

        rule_e_red = Rule(type="expression", dxf=dxf_red, stopIfTrue=True)
        rule_e_red.formula = [f'AND($D{r0}<>"",E{r0}<>"",E{r0}<$D{r0})']
        ws.conditional_formatting.add(f"E{r0}:E{last_row}", rule_e_red)

        rule_e_green = Rule(type="expression", dxf=dxf_green, stopIfTrue=True)
        rule_e_green.formula = [f'AND($D{r0}<>"",E{r0}<>"",E{r0}>=$D{r0})']
        ws.conditional_formatting.add(f"E{r0}:E{last_row}", rule_e_green)

        rule_i_red = Rule(type="expression", dxf=dxf_red, stopIfTrue=True)
        rule_i_red.formula = [f'AND($H{r0}<>"",I{r0}<>"",I{r0}<$H{r0})']
        ws.conditional_formatting.add(f"I{r0}:I{last_row}", rule_i_red)

        rule_i_green = Rule(type="expression", dxf=dxf_green, stopIfTrue=True)
        rule_i_green.formula = [f'AND($H{r0}<>"",I{r0}<>"",I{r0}>=$H{r0})']
        ws.conditional_formatting.add(f"I{r0}:I{last_row}", rule_i_green)

    # 達標名單（含金額可輸入）
    def draw_achiever_block(start_row, title, achievers, default_money, qc_mult, restock_mult):
        col_line = 12  # L
        col_name = 13  # M
        col_money = 14 # N

        _set_row_height(ws, start_row)
        tcell = ws.cell(start_row, col_line, title)
        tcell.alignment = left
        _set_base_font(tcell, force_bold=True)
        ws.merge_cells(start_row=start_row, start_column=col_line, end_row=start_row, end_column=col_money)

        hr = start_row + 1
        _set_row_height(ws, hr)
        h1 = ws.cell(hr, col_line, "達標 Line ID")
        h2 = ws.cell(hr, col_name, "姓名（手動輸入）")
        h3 = ws.cell(hr, col_money, "金額（可輸入/預設公式）")
        for c in (h1, h2, h3):
            c.fill = sub_header_fill
            c.border = border
            c.alignment = center
            _set_base_font(c, force_bold=True)

        rr = hr + 1
        first_list_row = rr

        def _write_row(label, money_formula_or_blank):
            nonlocal rr
            _set_row_height(ws, rr)

            c1 = ws.cell(rr, col_line, label)
            c1.alignment = center
            c1.border = border
            _set_base_font(c1, force_bold=True)

            c2 = ws.cell(rr, col_name, "")
            c2.border = box_border
            c2.alignment = left
            _set_base_font(c2)

            c3 = ws.cell(rr, col_money, money_formula_or_blank)
            c3.border = box_border
            c3.alignment = right
            c3.number_format = NUM_FMT_MONEY_HIDE0
            _set_base_font(c3)

            rr += 1

        if achievers:
            for lid in achievers:
                _write_row(str(lid), f"={int(default_money)}")
        else:
            _write_row("（無）", "")

        open_rng = f"$A$2:$A${last_row}"
        ach_rng = f"$L${first_list_row}:$L${rr-1}"

        open_qca = "(" + countif_sum(open_rng, QCA_LIST) + ")"
        ach_qca  = "(" + countif_sum(ach_rng,  QCA_LIST) + ")"
        open_qcb = "(" + countif_sum(open_rng, QCB_LIST) + ")"
        ach_qcb  = "(" + countif_sum(ach_rng,  QCB_LIST) + ")"

        open_all = f'COUNTIF({open_rng},"GT-*")'
        ach_all  = f'COUNTIF({ach_rng},"GT-*")'

        qca_formula = f'=IF({open_qca}=0,"",TRUNC({ach_qca}/{open_qca}*{qc_mult},0))'
        qcb_formula = f'=IF({open_qcb}=0,"",TRUNC({ach_qcb}/{open_qcb}*{qc_mult},0))'
        restock_formula = f'=IF({open_all}=0,"",TRUNC({ach_all}/{open_all}*{restock_mult},0))'

        _write_row("QCA", qca_formula)
        _write_row("QCB", qcb_formula)
        _write_row("補貨", restock_formula)

        return rr

    start = last_row + 3
    next_row = draw_achiever_block(
        start_row=start,
        title="上午達標名單（PCS≥上午應達成）",
        achievers=am_achievers,
        default_money=AM_DEFAULT_MONEY,
        qc_mult=50,
        restock_mult=100,
    )
    draw_achiever_block(
        start_row=next_row,
        title="下午達標名單（PCS≥下午目標）",
        achievers=pm_achievers,
        default_money=PM_DEFAULT_MONEY,
        qc_mult=100,
        restock_mult=50,
    )

    ws.freeze_panes = "A2"
    return ws


# =========================
# KPI 圖表 sheet（新增）
# =========================
def add_kpi_chart_sheet(wb: Workbook, kpi_df: pd.DataFrame):
    """
    kpi_df 欄位：
      日期, 上午達標率, 下午達標率, 上午平均產力, 下午平均產力, 上午總PCS加權, 下午總PCS加權
    """
    name = "KPI圖表"
    if name in wb.sheetnames:
        wb.remove(wb[name])
    ws = wb.create_sheet(name, 0)  # 放最前面

    header_fill = PatternFill("solid", fgColor="D9E1F2")
    thin = Side(style="thin", color="D0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    right = Alignment(horizontal="right", vertical="center")

    cols = list(kpi_df.columns)
    ws.append(cols)
    _set_row_height(ws, 1)
    for j, c in enumerate(cols, start=1):
        cell = ws.cell(1, j)
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center
        _set_base_font(cell, force_bold=True)

    for _, row in kpi_df.iterrows():
        ws.append([row.get(c, None) for c in cols])

    for r in range(2, ws.max_row + 1):
        _set_row_height(ws, r)
        for j in range(1, ws.max_column + 1):
            cell = ws.cell(r, j)
            cell.border = border
            cell.alignment = right if j >= 2 else center
            _set_base_font(cell)

    # 欄寬
    ws.column_dimensions["A"].width = 14
    for col_letter in "BCDEFG":
        ws.column_dimensions[col_letter].width = 16

    # 數字格式
    # B,C (達標率) -> %
    for col in ("B", "C"):
        for r in range(2, ws.max_row + 1):
            ws[f"{col}{r}"].number_format = "0.0%"

    # D,E (平均產力) -> 0.00
    for col in ("D", "E"):
        for r in range(2, ws.max_row + 1):
            ws[f"{col}{r}"].number_format = "#,##0.00"

    # F,G (總PCS加權) -> 整數
    for col in ("F", "G"):
        for r in range(2, ws.max_row + 1):
            ws[f"{col}{r}"].number_format = "#,##0"

    # ===== 圖表1：達標率（B/C）
    bar1 = BarChart()
    bar1.type = "col"
    bar1.title = "達標率（上午/下午）"
    bar1.y_axis.title = "達標率"
    bar1.dataLabels = DataLabelList()
    bar1.dataLabels.showVal = False

    data = Reference(ws, min_col=2, max_col=3, min_row=1, max_row=ws.max_row)
    cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
    bar1.add_data(data, titles_from_data=True)
    bar1.set_categories(cats)
    ws.add_chart(bar1, "I2")

    # ===== 圖表2：平均產力（D/E）
    line = LineChart()
    line.title = "平均產力（加權PCS/人力）"
    line.y_axis.title = "平均產力"
    line.dataLabels = DataLabelList()
    line.dataLabels.showVal = False

    data2 = Reference(ws, min_col=4, max_col=5, min_row=1, max_row=ws.max_row)
    line.add_data(data2, titles_from_data=True)
    line.set_categories(cats)
    ws.add_chart(line, "I18")

    # ===== 圖表3：總PCS加權（F/G）
    bar2 = BarChart()
    bar2.type = "col"
    bar2.title = "總PCS(加權)（上午/下午）"
    bar2.y_axis.title = "總PCS(加權)"
    bar2.dataLabels = DataLabelList()
    bar2.dataLabels.showVal = False

    data3 = Reference(ws, min_col=6, max_col=7, min_row=1, max_row=ws.max_row)
    bar2.add_data(data3, titles_from_data=True)
    bar2.set_categories(cats)
    ws.add_chart(bar2, "I34")

    ws.freeze_panes = "A2"
    return ws


# =========================
# 產出 Excel bytes（含 KPI圖表）
# =========================
def build_output_excel_bytes(df_raw: pd.DataFrame, target_per_mh: int, manpower_by_date: dict):
    df_raw, c_pickdate, c_packqty, c_cweight, c_lineid, c_stotype = normalize_columns(df_raw)
    df2, line_base, split = build_hourly_metrics(df_raw, c_pickdate, c_packqty, c_cweight, c_lineid, c_stotype)

    dates = sorted(df2["PICK_DATE"].dropna().unique())
    if not dates:
        raise ValueError("PICKDATE 解析後沒有日期資料，請確認 PICKDATE 欄位內容。")

    date_to_hours = {}
    date_to_lineids = {}
    for d in dates:
        hours = sorted(df2.loc[df2["PICK_DATE"] == d, "HOUR"].dropna().unique().tolist())
        date_to_hours[d] = [int(x) for x in hours]
        lineids = sorted(df2.loc[df2["PICK_DATE"] == d, c_lineid].dropna().astype(str).unique().tolist())
        date_to_lineids[d] = [str(x) for x in lineids]

    # manpower map (date, lineid, hour) -> value
    manpower_map = {}
    for d in dates:
        hours = date_to_hours.get(d, [])
        lineids = date_to_lineids.get(d, [])
        table = manpower_by_date.get(str(d))
        if table is None:
            continue

        # 確保是數字
        table2 = table.copy()
        for col in table2.columns:
            table2[col] = pd.to_numeric(table2[col], errors="coerce")

        for lid in lineids:
            for h in hours:
                v = np.nan
                try:
                    v = table2.loc[str(lid), str(h)]
                except Exception:
                    try:
                        v = table2.loc[str(lid), h]
                    except Exception:
                        v = np.nan
                manpower_map[(d, str(lid), int(h))] = v

    wb = Workbook()
    wb.remove(wb.active)

    summary_inputs = {}
    kpi_rows = []

    for d in dates:
        hours = date_to_hours[d]
        if not hours:
            continue
        lineids = date_to_lineids[d]

        sub_base = line_base[line_base["PICK_DATE"] == d]
        line_base_map = {}
        for lid in lineids:
            tmp = sub_base[sub_base[c_lineid].astype(str).str.strip() == str(lid)]
            line_base_map[(str(lid), "PCS")] = {int(r["HOUR"]): r["PCS"] for _, r in tmp.iterrows()}
            line_base_map[(str(lid), "加權PCS")] = {int(r["HOUR"]): r["加權PCS"] for _, r in tmp.iterrows()}

        sub_split = split[split["PICK_DATE"] == d]
        split_map = {}
        for lid in lineids:
            tmpL = sub_split[sub_split[c_lineid].astype(str).str.strip() == str(lid)]
            for t in ["GSO", "GXSO"]:
                tmpT = tmpL[tmpL[c_stotype].astype(str).str.strip() == t]
                split_map[(str(lid), t, "PCS")] = {int(r["HOUR"]): r["PCS"] for _, r in tmpT.iterrows()}
                split_map[(str(lid), t, "加權PCS")] = {int(r["HOUR"]): r["加權PCS"] for _, r in tmpT.iterrows()}

        sname = str(d)
        date_ws = write_hourly_sheet(
            wb=wb,
            sheet_name=sname,
            date_value=d,
            hours=hours,
            lineids=lineids,
            line_base_map=line_base_map,
            split_map=split_map,
            manpower_map=manpower_map,
        )

        am_ach, pm_ach = compute_achievers(d, lineids, hours, line_base_map, manpower_map, target_per_mh=target_per_mh)

        # KPI 總表用資料（Python 直接算，不靠公式）
        def _period_metrics(period_hours, achievers):
            total_man = 0.0
            total_pcs = 0.0
            open_lines = 0

            for lid in lineids:
                lid = str(lid)
                man = sum(_as_float(manpower_map.get((d, lid, int(h)), 0)) for h in period_hours if h in hours)
                if man > 0:
                    open_lines += 1
                total_man += man

                pcs_map = line_base_map.get((lid, "加權PCS"), {})
                pcs = sum(float(pcs_map.get(int(h), 0) or 0) for h in period_hours if h in hours)
                total_pcs += pcs

            ach_lines = len(achievers)
            ach_rate = (ach_lines / open_lines) if open_lines > 0 else np.nan
            avg_prod = (total_pcs / total_man) if total_man > 0 else np.nan
            return ach_rate, avg_prod, total_pcs

        am_rate, am_prod, am_pcs = _period_metrics(AM_HOURS, am_ach)
        pm_rate, pm_prod, pm_pcs = _period_metrics(PM_HOURS, pm_ach)

        kpi_rows.append({
            "日期": str(d),
            "上午達標率": am_rate,
            "下午達標率": pm_rate,
            "上午平均產力": am_prod,
            "下午平均產力": pm_prod,
            "上午總PCS加權": math.trunc(am_pcs) if not (isinstance(am_pcs, float) and np.isnan(am_pcs)) else np.nan,
            "下午總PCS加權": math.trunc(pm_pcs) if not (isinstance(pm_pcs, float) and np.isnan(pm_pcs)) else np.nan,
        })

        _, hour_to_col = find_hour_header_row(date_ws)
        line_map2 = parse_line_rows(date_ws)

        am_cols = [hour_to_col[h] for h in AM_HOURS if h in hour_to_col]
        pm_cols = [hour_to_col[h] for h in PM_HOURS if h in hour_to_col]

        summary_inputs[d] = (date_ws, line_map2, am_cols, pm_cols, am_ach, pm_ach)

    # KPI圖表 sheet
    kpi_df = pd.DataFrame(kpi_rows)
    if not kpi_df.empty:
        add_kpi_chart_sheet(wb, kpi_df)

    # 建彙總 sheet
    for d in dates:
        date_sheet = str(d)
        if date_sheet not in wb.sheetnames:
            continue
        pack = summary_inputs.get(d)
        if not pack:
            continue
        date_ws, line_map2, am_cols, pm_cols, am_ach, pm_ach = pack
        if date_ws is None or not line_map2:
            continue

        build_summary_sheet_with_achievers(
            wb=wb,
            summary_name=f"彙總_{date_sheet}",
            date_ws=date_ws,
            line_map=line_map2,
            am_cols=am_cols,
            pm_cols=pm_cols,
            am_achievers=am_ach,
            pm_achievers=pm_ach,
            target_per_mh=target_per_mh,
        )

    # 排序：KPI圖表 -> 彙總_* -> 日期
    sheet_order = []
    if "KPI圖表" in wb.sheetnames:
        sheet_order.append("KPI圖表")
    sheet_order += [sn for sn in wb.sheetnames if sn.startswith("彙總_")]
    sheet_order += [sn for sn in wb.sheetnames if (sn not in sheet_order)]
    wb._sheets = [wb[sn] for sn in sheet_order]

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue(), [str(d) for d in dates], kpi_df


# =========================
# UI helpers（修正：可一次貼上/可快速填入）
# =========================
def _hash_bytes(b: bytes) -> str:
    return hashlib.sha256(b).hexdigest()


def _init_manpower_table(lineids, hours):
    cols = [str(int(h)) for h in hours]
    idx = [str(x) for x in lineids]
    # ✅ 用可空數字型別，貼上 Excel 才不會「字串/空白」卡住
    df = pd.DataFrame(pd.NA, index=idx, columns=cols, dtype="Float64")
    return df


def _apply_fill(df: pd.DataFrame, line_id: str, hours: list[int], value: float, which: str):
    df2 = df.copy()
    if line_id not in df2.index:
        return df2

    if which == "整天":
        tgt_hours = hours
    elif which == "上午":
        tgt_hours = [h for h in hours if h in AM_HOURS]
    else:
        tgt_hours = [h for h in hours if h in PM_HOURS]

    for h in tgt_hours:
        c = str(int(h))
        if c in df2.columns:
            df2.loc[line_id, c] = value
    return df2


def _parse_clipboard_to_df(text: str) -> pd.DataFrame:
    """
    支援貼 Excel 區塊：
    - 允許第一列是小時（8 9 10...），第一欄是 line id
    """
    text = (text or "").strip()
    if not text:
        return pd.DataFrame()

    # 用 tab 分隔為主
    rows = [r for r in text.splitlines() if r.strip() != ""]
    grid = [r.split("\t") for r in rows]
    if len(grid) < 2 or len(grid[0]) < 2:
        return pd.DataFrame()

    header = [h.strip() for h in grid[0]]
    data = grid[1:]

    # 第一欄當 index
    idx = [str(r[0]).strip() for r in data]
    cols = [str(c).strip() for c in header[1:]]

    vals = []
    for r in data:
        row_vals = []
        for v in r[1:]:
            v = str(v).strip()
            if v == "":
                row_vals.append(pd.NA)
            else:
                try:
                    row_vals.append(float(v))
                except Exception:
                    row_vals.append(pd.NA)
        # 補齊長度
        if len(row_vals) < len(cols):
            row_vals += [pd.NA] * (len(cols) - len(row_vals))
        vals.append(row_vals[:len(cols)])

    out = pd.DataFrame(vals, index=idx, columns=cols, dtype="Float64")
    return out


# =========================
# Streamlit main
# =========================
def main():
    st.set_page_config(page_title="出貨作業線產能", page_icon="📦", layout="wide")
    inject_logistics_theme()
    set_page("出貨課｜出貨作業線產能", "📦")

    card_open("📥 來源檔案")
    up = st.file_uploader(
        "上傳來源檔（含 PICKDATE / PACKQTY / Cweight / LINEID / STO_TYPE）",
        type=["xlsx", "xlsm", "xltx", "xltm", "xls", "csv"],
        accept_multiple_files=False,
    )
    card_close()

    if not up:
        st.info("請先上傳檔案。")
        return

    raw = up.getvalue()
    file_sig = _hash_bytes(raw)

    card_open("⚙️ 參數設定")
    target_per_mh = st.number_input(
        "每小時目標(加權)（TARGET_PER_MANHOUR）",
        min_value=1, max_value=99999, value=int(TARGET_PER_MANHOUR_DEFAULT), step=1
    )
    card_close()

    # 解析資料（只在檔案變動時重新解析）
    if st.session_state.get("ship_line_prod_file_sig") != file_sig:
        try:
            df = robust_read_bytes(up.name, raw)
            df_norm, c_pickdate, c_packqty, c_cweight, c_lineid, c_stotype = normalize_columns(df)
            df2, _, _ = build_hourly_metrics(df_norm, c_pickdate, c_packqty, c_cweight, c_lineid, c_stotype)
        except Exception as e:
            st.error(f"讀取/解析失敗：{e}")
            return

        dates = sorted(df2["PICK_DATE"].dropna().unique())
        if not dates:
            st.error("PICKDATE 解析後沒有日期資料。")
            return

        date_to_hours = {}
        date_to_lineids = {}
        for d in dates:
            hours = sorted(df2.loc[df2["PICK_DATE"] == d, "HOUR"].dropna().unique().tolist())
            lineids = sorted(df2.loc[df2["PICK_DATE"] == d, c_lineid].dropna().astype(str).unique().tolist())
            date_to_hours[str(d)] = [int(x) for x in hours]
            date_to_lineids[str(d)] = [str(x) for x in lineids]

        st.session_state["ship_line_prod_file_sig"] = file_sig
        st.session_state["ship_line_prod_df"] = df
        st.session_state["ship_line_prod_dates"] = [str(d) for d in dates]
        st.session_state["ship_line_prod_date_to_hours"] = date_to_hours
        st.session_state["ship_line_prod_date_to_lineids"] = date_to_lineids

        # 初始化每日期人力表（可空數字表）
        for d in st.session_state["ship_line_prod_dates"]:
            key = f"mp_{d}"
            st.session_state[key] = _init_manpower_table(date_to_lineids[d], date_to_hours[d])

    dates = st.session_state["ship_line_prod_dates"]
    date_to_hours = st.session_state["ship_line_prod_date_to_hours"]
    date_to_lineids = st.session_state["ship_line_prod_date_to_lineids"]
    df_source = st.session_state["ship_line_prod_df"]

    # =========================
    # 👥 人力輸入（強化）
    # =========================
    card_open("👥 人力輸入（可直接貼 Excel 多格 / 支援快速填入）")
    st.caption("✅ 你可以從 Excel 複製一整塊（含多格）直接貼進表格；空白會保留空白。")

    tabs = st.tabs(dates)
    manpower_by_date = {}

    for i, d in enumerate(dates):
        with tabs[i]:
            hours = date_to_hours.get(d, [])
            lineids = date_to_lineids.get(d, [])
            if not hours or not lineids:
                st.warning("此日期沒有小時/Line 資料")
                continue

            key = f"mp_{d}"
            mp_df = st.session_state.get(key)
            if mp_df is None or list(mp_df.columns) != [str(int(h)) for h in hours] or list(mp_df.index) != [str(x) for x in lineids]:
                mp_df = _init_manpower_table(lineids, hours)
                st.session_state[key] = mp_df

            c1, c2, c3, c4 = st.columns([2, 2, 2, 3])
            with c1:
                sel_line = st.selectbox("快速填入｜Line ID", options=lineids, key=f"sel_line_{d}")
            with c2:
                fill_val = st.number_input("填入值（人力）", value=0.0, step=0.5, key=f"fill_val_{d}")
            with c3:
                which = st.selectbox("範圍", options=["整天", "上午", "下午"], key=f"which_{d}")
            with c4:
                if st.button("套用填入", use_container_width=True, key=f"apply_{d}"):
                    st.session_state[key] = _apply_fill(st.session_state[key], str(sel_line), hours, float(fill_val), which)

            # 可選：貼上區（解決某些瀏覽器 data_editor 貼上不穩）
            with st.expander("（備用）貼上區：把 Excel 區塊貼這裡再套用", expanded=False):
                clip = st.text_area(
                    "格式：第一列是小時，第一欄是 Line ID（Tab 分隔）",
                    height=120,
                    key=f"clip_{d}",
                )
                cc1, cc2 = st.columns([1, 1])
                with cc1:
                    if st.button("解析並套用（覆蓋同名 Line / 小時）", use_container_width=True, key=f"clip_apply_{d}"):
                        patch = _parse_clipboard_to_df(clip)
                        if patch.empty:
                            st.warning("貼上內容解析不到表格")
                        else:
                            base = st.session_state[key].copy()
                            # 對齊欄列後覆蓋
                            for ridx in patch.index:
                                if ridx in base.index:
                                    for c in patch.columns:
                                        if c in base.columns:
                                            base.loc[ridx, c] = patch.loc[ridx, c]
                            st.session_state[key] = base
                            st.success("已套用貼上內容")
                with cc2:
                    st.caption("如果你直接貼在表格會異常，就用這個備用方式。")

            # ✅ 讓 data_editor 明確是數字欄，貼上最穩
            col_cfg = {str(int(h)): st.column_config.NumberColumn(width="small") for h in hours}

            edited = st.data_editor(
                st.session_state[key],
                use_container_width=True,
                num_rows="fixed",
                column_config=col_cfg,
                key=f"editor_{d}",
            )

            st.session_state[key] = edited
            manpower_by_date[d] = edited

    card_close()

    # =========================
    # 📤 匯出 + KPI 圖表
    # =========================
    card_open("📊 KPI 總覽（頁面即時）")
    if st.button("計算 KPI 預覽", use_container_width=True):
        try:
            out_bytes, out_dates, kpi_df = build_output_excel_bytes(
                df_raw=df_source,
                target_per_mh=int(target_per_mh),
                manpower_by_date=manpower_by_date,
            )
            st.session_state["kpi_preview_df"] = kpi_df
            st.session_state["last_out_bytes"] = out_bytes
            st.session_state["last_out_name"] = f"{os.path.splitext(up.name)[0]}_出貨作業線產能_含KPI圖表.xlsx"
            st.success("KPI 已計算完成（下方有預覽與下載）")
        except Exception as e:
            st.error(f"KPI 計算失敗：{e}")

    kpi_df = st.session_state.get("kpi_preview_df")
    if isinstance(kpi_df, pd.DataFrame) and not kpi_df.empty:
        st.dataframe(kpi_df, use_container_width=True)

        # 圖表（不指定顏色，使用預設）
        chart_df = kpi_df.copy()
        chart_df = chart_df.set_index("日期")

        st.caption("達標率（上午/下午）")
        st.bar_chart(chart_df[["上午達標率", "下午達標率"]])

        st.caption("平均產力（上午/下午）")
        st.line_chart(chart_df[["上午平均產力", "下午平均產力"]])

        st.caption("總PCS(加權)（上午/下午）")
        st.bar_chart(chart_df[["上午總PCS加權", "下午總PCS加權"]])

    card_close()

    card_open("📤 匯出 Excel（含 KPI圖表 Sheet）")
    out_bytes = st.session_state.get("last_out_bytes")
    out_name = st.session_state.get("last_out_name")

    if out_bytes:
        st.download_button(
            "⬇️ 下載輸出 Excel",
            data=out_bytes,
            file_name=out_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
        st.caption("輸出包含：KPI圖表（含圖表）→ 彙總_日期 → 各日期戰情表")
    else:
        st.info("先按「計算 KPI 預覽」後就可以下載。")

    card_close()


if __name__ == "__main__":
    main()
