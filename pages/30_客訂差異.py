# -*- coding: utf-8 -*-
"""
30_客訂差異.py
Streamlit 頁面版：客訂差異分析

部署位置：pages/30_客訂差異.py
功能：
1. 上傳差異明細、訂單明細、庫存明細、儲位明細，可多檔
2. 判斷差異商品是否為客訂單
3. 依商品碼比對庫存其他儲位，並依最短效排序顯示儲位1~儲位3
4. 依完整明細的儲位比對棚別
5. 匯出 Excel，完整明細的國際條碼後 5 碼放大加粗
"""

from __future__ import annotations

from io import BytesIO
from pathlib import Path
from typing import Iterable

import pandas as pd
import streamlit as st

from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.styles import Alignment, Font, PatternFill


# =====================================================
# 頁面基礎設定：可接你現有 common_ui，沒有也不會壞
# =====================================================
try:
    from common_ui import set_page, inject_logistics_theme, card_open, card_close
except Exception:
    set_page = None
    inject_logistics_theme = None
    card_open = None
    card_close = None


def setup_page() -> None:
    if set_page:
        try:
            set_page("30_客訂差異", icon="🧾", layout="wide")
        except TypeError:
            set_page("30_客訂差異")
    else:
        st.set_page_config(page_title="30_客訂差異", page_icon="🧾", layout="wide")

    if inject_logistics_theme:
        try:
            inject_logistics_theme()
        except Exception:
            pass

    st.markdown(
        """
        <style>
        .dafeng-title {
            font-size: 1.7rem;
            font-weight: 800;
            color: #0f172a;
            margin-bottom: 0.25rem;
        }
        .dafeng-subtitle {
            color: #64748b;
            font-size: 0.95rem;
            margin-bottom: 1rem;
        }
        .hint-box {
            border: 1px solid #dbeafe;
            background: #eff6ff;
            color: #1e3a8a;
            padding: 0.8rem 1rem;
            border-radius: 14px;
            margin-bottom: 1rem;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )


# =====================================================
# 差異明細欄位對照設定
# =====================================================
COLUMN_MAP = {
    "AllDIF": "差異量",
    "SONO": "批次",
    "ORIGINAL_SKU": "商品碼",
    "BARCODE": "國際條碼",
    "DESCR": "品名",
    "PICKLOC": "儲位",
}

OUTPUT_COLUMNS = [
    "差異量",
    "儲位",
    "國際條碼",
    "品名",
    "商品碼",
    "批次",
]

FINAL_COLUMNS = [
    "差異量",
    "儲位",
    "國際條碼",
    "品名",
    "備註",
    "儲位1",
    "儲位2",
    "儲位3",
    "商品碼",
    "批次",
    "訂單類型",
    "棚別",
]

MAX_ALT_LOCS = 3


# =====================================================
# Streamlit 上傳檔案讀取
# =====================================================
def _uploaded_to_bytes(uploaded_file) -> bytes:
    uploaded_file.seek(0)
    return uploaded_file.read()


def read_file_smart(uploaded_file) -> pd.DataFrame:
    """讀取 Streamlit UploadedFile，支援 xlsx/xls/csv/txt/假 xls。"""
    file_name = getattr(uploaded_file, "name", "uploaded_file")
    data = _uploaded_to_bytes(uploaded_file)
    header = data[:8]

    # xlsx / xlsm
    if header.startswith(b"PK"):
        return pd.read_excel(BytesIO(data), dtype=str, keep_default_na=False)

    # 真正 xls：Streamlit Cloud 需 requirements.txt 有 xlrd
    if header.startswith(b"\xD0\xCF\x11\xE0"):
        try:
            return pd.read_excel(BytesIO(data), dtype=str, keep_default_na=False, engine="xlrd")
        except ImportError as exc:
            raise ImportError(
                f"{file_name} 是舊版 .xls 格式，Streamlit Cloud 需要安裝 xlrd。\n"
                "請在 requirements.txt 加上：xlrd"
            ) from exc

    # 假 xls / txt / csv
    encodings = ["utf-8-sig", "cp950", "big5", "utf-16", "latin1"]
    seps = ["\t", ",", "|", ";"]
    last_error: Exception | None = None

    for enc in encodings:
        for sep in seps:
            try:
                df = pd.read_csv(
                    BytesIO(data),
                    sep=sep,
                    dtype=str,
                    encoding=enc,
                    keep_default_na=False,
                    engine="python",
                )
                if len(df.columns) > 1:
                    return df
            except Exception as e:
                last_error = e

    raise ValueError(f"檔案讀取失敗：{file_name}\n最後錯誤：{last_error}")


def read_and_concat_files(uploaded_files: Iterable, file_type_name: str) -> pd.DataFrame:
    all_df: list[pd.DataFrame] = []

    for uploaded_file in uploaded_files or []:
        df = read_file_smart(uploaded_file)
        df.columns = df.columns.astype(str).str.strip()
        df["來源檔案"] = getattr(uploaded_file, "name", "uploaded_file")
        all_df.append(df)

    if not all_df:
        raise ValueError(f"未上傳任何{file_type_name}檔案")

    return pd.concat(all_df, ignore_index=True)


# =====================================================
# 客訂單判斷與欄位尋找
# =====================================================
def is_customer_order(order_no) -> bool:
    order_no = str(order_no).strip().upper()

    if order_no == "" or order_no.lower() == "nan":
        return False
    if order_no.startswith("U"):
        return True
    if order_no.startswith("GB"):
        return True
    if len(order_no) >= 4 and order_no[:4].isdigit() and order_no.startswith("20"):
        return True

    return False


def _find_column(df: pd.DataFrame, possible_columns: list[str], title: str) -> str:
    for col in possible_columns:
        if col in df.columns:
            return col

    raise ValueError(
        f"{title}找不到必要欄位。\n\n"
        "可接受欄位名稱包含：\n"
        + "\n".join(possible_columns)
        + "\n\n目前檔案欄位如下：\n"
        + "\n".join(df.columns.astype(str))
    )


def find_order_column(df):
    return _find_column(
        df,
        ["訂單號", "訂單編號", "貨主訂單", "貨主訂單號", "ORDER_NO", "ORDERNO", "SO_NO", "SONO"],
        "訂單明細",
    )


def find_product_column(df):
    return _find_column(
        df,
        ["商品", "商品碼", "商品編號", "品號", "SKU", "ORIGINAL_SKU", "商品號", "貨號"],
        "訂單資料",
    )


def find_inventory_product_column(df):
    return _find_column(
        df,
        ["商品號", "商品碼", "商品", "ORIGINAL_SKU", "SKU", "品號", "商品編號", "貨號"],
        "庫存明細",
    )


def find_inventory_loc_column(df):
    return _find_column(df, ["儲位", "PICKLOC", "LOCATION", "LOC", "庫位"], "庫存明細")


def find_inventory_expiry_column(df):
    for col in ["商品效期", "效期", "有效期限", "EXPDATE", "EXPIRE_DATE", "EXPIRY"]:
        if col in df.columns:
            return col
    return None


def find_inventory_days_column(df):
    for col in ["效期剩餘天數", "剩餘天數", "DAYS", "REMAIN_DAYS"]:
        if col in df.columns:
            return col
    return None


def find_inventory_qty_column(df):
    for col in ["Canuseqty", "可用量", "數量", "QTY", "庫存量"]:
        if col in df.columns:
            return col
    return None


def find_location_master_loc_column(df):
    return _find_column(df, ["儲位", "庫位", "LOCATION", "LOC", "PICKLOC"], "儲位明細")


def find_location_master_shed_column(df):
    return _find_column(df, ["棚別", "區域", "倉別", "儲區", "ZONE", "AREA"], "儲位明細")


# =====================================================
# 資料整理
# =====================================================
def clean_difference_df(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = df.columns.astype(str).str.strip()

    missing_cols = [col for col in COLUMN_MAP.keys() if col not in df.columns]
    if missing_cols:
        raise ValueError(
            "差異明細缺少必要欄位：\n"
            + "\n".join(missing_cols)
            + "\n\n目前檔案欄位如下：\n"
            + "\n".join(df.columns.astype(str))
        )

    result = df[list(COLUMN_MAP.keys())].rename(columns=COLUMN_MAP)
    result = result[OUTPUT_COLUMNS]

    for col in result.columns:
        result[col] = result[col].astype(str).str.strip()

    result["差異量"] = (
        result["差異量"].astype(str).str.replace(",", "", regex=False).str.strip()
    )
    result["差異量"] = pd.to_numeric(result["差異量"], errors="coerce").fillna(0).astype(int)
    result = result[result["差異量"] != 0].copy()

    for col in ["儲位", "國際條碼", "商品碼", "批次"]:
        result[col] = result[col].astype(str).str.strip()

    return result


def clean_order_df(order_df: pd.DataFrame):
    order_df = order_df.copy()
    order_df.columns = order_df.columns.astype(str).str.strip()

    order_col = find_order_column(order_df)
    order_df[order_col] = order_df[order_col].astype(str).str.strip()

    order_df["訂單類型"] = order_df[order_col].apply(
        lambda x: "客訂單" if is_customer_order(x) else "非客訂單"
    )
    order_df["備註"] = order_df["訂單類型"].apply(lambda x: "客訂單" if x == "客訂單" else "正常")

    customer_order_product_df = order_df[order_df["訂單類型"] == "客訂單"].copy()
    product_col = find_product_column(customer_order_product_df)
    customer_order_product_df[product_col] = customer_order_product_df[product_col].astype(str).str.strip()

    unique_customer_product_df = (
        customer_order_product_df[customer_order_product_df[product_col] != ""]
        .drop_duplicates(subset=[product_col])
        [[product_col, "訂單類型"]]
        .rename(columns={product_col: "商品"})
        .copy()
    )

    return order_df, customer_order_product_df, unique_customer_product_df, order_col, product_col


def build_inventory_location_map(inventory_df: pd.DataFrame):
    inventory_df = inventory_df.copy()
    inventory_df.columns = inventory_df.columns.astype(str).str.strip()

    product_col = find_inventory_product_column(inventory_df)
    loc_col = find_inventory_loc_column(inventory_df)
    expiry_col = find_inventory_expiry_column(inventory_df)
    days_col = find_inventory_days_column(inventory_df)
    qty_col = find_inventory_qty_column(inventory_df)

    inventory_df[product_col] = inventory_df[product_col].astype(str).str.strip()
    inventory_df[loc_col] = inventory_df[loc_col].astype(str).str.strip()

    inventory_df = inventory_df[(inventory_df[product_col] != "") & (inventory_df[loc_col] != "")].copy()

    if qty_col:
        inventory_df["_可用量排序"] = pd.to_numeric(
            inventory_df[qty_col].astype(str).str.replace(",", "", regex=False), errors="coerce"
        ).fillna(0)
        inventory_df = inventory_df[inventory_df["_可用量排序"] > 0].copy()
    else:
        inventory_df["_可用量排序"] = 0

    if expiry_col:
        inventory_df["_效期排序"] = pd.to_datetime(inventory_df[expiry_col], errors="coerce")
    else:
        inventory_df["_效期排序"] = pd.NaT

    if days_col:
        inventory_df["_剩餘天數排序"] = pd.to_numeric(
            inventory_df[days_col].astype(str).str.replace(",", "", regex=False), errors="coerce"
        )
    else:
        inventory_df["_剩餘天數排序"] = pd.NA

    inventory_df = inventory_df.sort_values(
        by=[product_col, "_效期排序", "_剩餘天數排序", loc_col],
        ascending=[True, True, True, True],
        na_position="last",
    )

    inventory_unique_loc_df = inventory_df.drop_duplicates(subset=[product_col, loc_col], keep="first").copy()

    inventory_location_map = {}
    for product, group in inventory_unique_loc_df.groupby(product_col):
        locs = group[loc_col].astype(str).str.strip().tolist()
        inventory_location_map[str(product).strip()] = locs

    return inventory_location_map, inventory_unique_loc_df.copy(), product_col, loc_col, expiry_col, days_col, qty_col


def add_other_locations_by_short_expiry(diff_result: pd.DataFrame, inventory_location_map: dict) -> pd.DataFrame:
    diff_result = diff_result.copy()

    for i in range(1, MAX_ALT_LOCS + 1):
        diff_result[f"儲位{i}"] = ""

    for idx, row in diff_result.iterrows():
        product = str(row["商品碼"]).strip()
        current_loc = str(row["儲位"]).strip()
        loc_list = inventory_location_map.get(product, [])
        other_locs = [loc for loc in loc_list if str(loc).strip() != current_loc]

        for i in range(MAX_ALT_LOCS):
            if i < len(other_locs):
                diff_result.at[idx, f"儲位{i + 1}"] = other_locs[i]

    return diff_result


def build_location_shed_map(location_df: pd.DataFrame):
    location_df = location_df.copy()
    location_df.columns = location_df.columns.astype(str).str.strip()

    loc_col = find_location_master_loc_column(location_df)
    shed_col = find_location_master_shed_column(location_df)

    location_df[loc_col] = location_df[loc_col].astype(str).str.strip()
    location_df[shed_col] = location_df[shed_col].astype(str).str.strip()

    location_df = location_df[(location_df[loc_col] != "") & (location_df[shed_col] != "")].copy()
    location_df = location_df.drop_duplicates(subset=[loc_col], keep="first").copy()

    location_shed_map = dict(zip(location_df[loc_col], location_df[shed_col]))
    location_summary_df = location_df[[loc_col, shed_col]].rename(columns={loc_col: "儲位", shed_col: "棚別"}).copy()

    return location_shed_map, location_summary_df, loc_col, shed_col


def run_analysis(diff_files, order_files, inventory_files, location_files):
    raw_diff_df = read_and_concat_files(diff_files, "差異明細")
    raw_order_df = read_and_concat_files(order_files, "訂單明細")
    raw_inventory_df = read_and_concat_files(inventory_files, "庫存明細")
    raw_location_df = read_and_concat_files(location_files, "儲位明細")

    diff_result = clean_difference_df(raw_diff_df)

    order_df, customer_order_product_df, unique_customer_product_df, order_col, product_col = clean_order_df(raw_order_df)

    customer_product_set = set(unique_customer_product_df["商品"].astype(str).str.strip().tolist())

    diff_result["訂單類型"] = diff_result["商品碼"].apply(
        lambda x: "客訂單" if str(x).strip() in customer_product_set else "非客訂單"
    )
    diff_result["備註"] = diff_result["訂單類型"].apply(lambda x: "客訂單" if x == "客訂單" else "正常")

    (
        inventory_location_map,
        inventory_summary_df,
        inv_product_col,
        inv_loc_col,
        inv_expiry_col,
        inv_days_col,
        inv_qty_col,
    ) = build_inventory_location_map(raw_inventory_df)

    diff_result = add_other_locations_by_short_expiry(diff_result, inventory_location_map)

    location_shed_map, location_summary_df, location_loc_col, location_shed_col = build_location_shed_map(raw_location_df)
    diff_result["棚別"] = diff_result["儲位"].apply(lambda x: location_shed_map.get(str(x).strip(), ""))
    diff_result = diff_result[FINAL_COLUMNS]

    stats = {
        "差異明細合併原始筆數": len(raw_diff_df),
        "訂單明細合併原始筆數": len(raw_order_df),
        "庫存明細合併原始筆數": len(raw_inventory_df),
        "儲位明細合併原始筆數": len(raw_location_df),
        "差異完整明細筆數": len(diff_result),
        "完整明細比對為客訂單筆數": int((diff_result["訂單類型"] == "客訂單").sum()),
        "完整明細比對為非客訂單筆數": int((diff_result["訂單類型"] == "非客訂單").sum()),
        "棚別成功對應筆數": int((diff_result["棚別"] != "").sum()),
        "訂單資料客訂單商品筆數": len(customer_order_product_df),
        "客訂單不重複商品數": len(unique_customer_product_df),
        "訂單號欄位": order_col,
        "訂單商品欄位": product_col,
        "庫存商品欄位": inv_product_col,
        "庫存儲位欄位": inv_loc_col,
        "庫存效期欄位": inv_expiry_col or "未找到",
        "庫存剩餘天數欄位": inv_days_col or "未找到",
        "庫存數量欄位": inv_qty_col or "未找到",
        "儲位明細儲位欄位": location_loc_col,
        "儲位明細棚別欄位": location_shed_col,
    }

    sheets = {
        "完整明細": diff_result,
        "差異明細_合併原始": raw_diff_df,
        "訂單資料_含判斷": order_df,
        "訂單資料_客訂單商品": customer_order_product_df,
        "訂單資料_客訂單商品_不重複": unique_customer_product_df,
        "庫存明細_合併原始": raw_inventory_df,
        "庫存最短效儲位整理": inventory_summary_df,
        "儲位明細_合併原始": raw_location_df,
        "儲位棚別對照": location_summary_df,
    }

    return diff_result, sheets, stats


# =====================================================
# Excel 匯出與美化
# =====================================================
def format_worksheet(ws) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    header_fill = PatternFill("solid", fgColor="1E3A8A")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="center")

    for column_cells in ws.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            cell_value = str(cell.value) if cell.value is not None else ""
            max_length = max(max_length, len(cell_value))
        ws.column_dimensions[column_letter].width = min(max_length + 4, 50)


def apply_barcode_last5_big(ws, header_name="國際條碼", normal_size=11, big_size=18) -> None:
    barcode_col = None
    for cell in ws[1]:
        if str(cell.value).strip() == header_name:
            barcode_col = cell.column
            break

    if barcode_col is None:
        return

    normal_font = InlineFont(sz=normal_size)
    big_font = InlineFont(sz=big_size, b=True)

    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=barcode_col)
        value = "" if cell.value is None else str(cell.value).strip()
        if value == "":
            continue
        if len(value) <= 5:
            cell.value = CellRichText([TextBlock(big_font, value)])
        else:
            cell.value = CellRichText([
                TextBlock(normal_font, value[:-5]),
                TextBlock(big_font, value[-5:]),
            ])


def build_excel_bytes(sheets: dict[str, pd.DataFrame]) -> bytes:
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for sheet_name, df in sheets.items():
            safe_name = sheet_name[:31]
            df.to_excel(writer, index=False, sheet_name=safe_name)

        for sheet_name in writer.book.sheetnames:
            ws = writer.book[sheet_name]
            format_worksheet(ws)
            if sheet_name == "完整明細":
                apply_barcode_last5_big(ws, header_name="國際條碼", normal_size=11, big_size=18)

    output.seek(0)
    return output.getvalue()


# =====================================================
# Streamlit UI
# =====================================================
def render_uploader(label: str, help_text: str):
    return st.file_uploader(
        label,
        type=["xls", "xlsx", "xlsm", "csv", "txt"],
        accept_multiple_files=True,
        help=help_text,
    )


def main() -> None:
    setup_page()

    st.markdown('<div class="dafeng-title">🧾 客訂差異分析</div>', unsafe_allow_html=True)
    st.markdown(
        '<div class="dafeng-subtitle">差異明細 × 訂單明細 × 庫存最短效儲位 × 儲位棚別對照</div>',
        unsafe_allow_html=True,
    )

    st.markdown(
        """
        <div class="hint-box">
        上傳 4 類檔案後，系統會產出「完整明細」，欄位順序為：
        差異量、儲位、國際條碼、品名、備註、儲位1、儲位2、儲位3、商品碼、批次、訂單類型、棚別。
        </div>
        """,
        unsafe_allow_html=True,
    )

    with st.expander("📌 比對邏輯說明", expanded=False):
        st.markdown(
            """
            - **客訂單判斷**：訂單號開頭為 `U`、`GB`、或 `20XX`。
            - **完整明細客訂比對**：用完整明細的 `商品碼` 對訂單資料中的客訂單不重複商品。
            - **其他儲位**：用 `商品碼` 對庫存商品欄位，依 `商品效期 / 效期剩餘天數 / 儲位` 排序，取其他儲位前 3 筆。
            - **棚別**：用完整明細的 `儲位` 對儲位明細的 `儲位`，帶出 `棚別`。
            - **Excel 顯示**：完整明細的 `國際條碼` 後 5 碼會放大加粗。
            """
        )

    col1, col2 = st.columns(2)
    with col1:
        diff_files = render_uploader("1️⃣ 上傳差異明細，可多選", "必要欄位：AllDIF、SONO、ORIGINAL_SKU、BARCODE、DESCR、PICKLOC")
        inventory_files = render_uploader("3️⃣ 上傳庫存明細，可多選", "用商品欄位比對其他儲位，並依最短效排序")
    with col2:
        order_files = render_uploader("2️⃣ 上傳訂單明細，可多選", "用訂單號判斷客訂單，再抓客訂商品")
        location_files = render_uploader("4️⃣ 上傳儲位明細，可多選", "用儲位比對棚別")

    ready = all([diff_files, order_files, inventory_files, location_files])

    if not ready:
        st.info("請先上傳 4 類檔案，才能開始產生客訂差異分析結果。")
        return

    if st.button("🚀 開始產生客訂差異報表", type="primary", use_container_width=True):
        try:
            with st.spinner("資料讀取與比對中，請稍候..."):
                diff_result, sheets, stats = run_analysis(diff_files, order_files, inventory_files, location_files)
                excel_bytes = build_excel_bytes(sheets)

            st.session_state["customer_diff_result"] = diff_result
            st.session_state["customer_diff_stats"] = stats
            st.session_state["customer_diff_excel_bytes"] = excel_bytes
            st.success("客訂差異報表產生完成。")

        except Exception as e:
            st.error("產生失敗，請依下方錯誤訊息檢查欄位或檔案格式。")
            st.exception(e)
            return

    if "customer_diff_result" not in st.session_state:
        return

    diff_result = st.session_state["customer_diff_result"]
    stats = st.session_state["customer_diff_stats"]
    excel_bytes = st.session_state["customer_diff_excel_bytes"]

    st.subheader("📊 統計摘要")
    m1, m2, m3, m4 = st.columns(4)
    m1.metric("完整明細筆數", f"{stats['差異完整明細筆數']:,}")
    m2.metric("客訂單筆數", f"{stats['完整明細比對為客訂單筆數']:,}")
    m3.metric("非客訂單筆數", f"{stats['完整明細比對為非客訂單筆數']:,}")
    m4.metric("棚別成功對應", f"{stats['棚別成功對應筆數']:,}")

    with st.expander("🔎 系統自動抓到的欄位", expanded=False):
        stats_df = pd.DataFrame(list(stats.items()), columns=["項目", "內容"])
        st.dataframe(stats_df, use_container_width=True, hide_index=True)

    st.subheader("📋 完整明細預覽")
    st.dataframe(diff_result.head(1000), use_container_width=True, hide_index=True)

    output_name = "30_客訂差異_多檔合併整理後_含最短效其他儲位與棚別.xlsx"
    st.download_button(
        label="📥 下載 Excel 報表",
        data=excel_bytes,
        file_name=output_name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )


if __name__ == "__main__":
    main()
