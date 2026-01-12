# pages/27_QCæœªä¸Šæ¶æ¯”å°.py
# -*- coding: utf-8 -*-
import io
from collections import defaultdict
from datetime import datetime, date
from typing import Optional, Tuple, Dict

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import copy as _copy

from common_ui import inject_logistics_theme, set_page, card_open, card_close


# =============================
# åƒæ•¸
# =============================
QC_KEY_HEADER = "å•†å“"
UN_KEY_HEADER = "å•†å“ç¢¼"
UN_DATE_HEADER = "é€²è²¨æ—¥"
UNIT_HEADER = "å¯ç§»å‹•å–®ä½"
BARCODE_HEADER = "åœ‹éš›æ¢ç¢¼"
BATCH_HEADER = "æ‰¹è™Ÿ"

MATCH_SHEET_NAME = "ç¬¦åˆæœªä¸Šæ¶æ˜ç´°"

# âœ… ä½ æŒ‡å®šè¦åˆªé™¤çš„æ¬„ä½
DELETE_HEADERS = [
    "ç§»å‹•çš„æ•¸é‡",
    "ç›®çš„å„²ä½",
    "å¯ç§»å‹•å–®ä½è‡³",
    "è¨ˆé‡å–®ä½ç”±",
    "åˆ°åŒ…è£ç¢¼",
    "å·²è©¦ç®—",
    "å·²æ€å–",
]


# =============================
# UI CSSï¼ˆèƒŒæ™¯ + ä¸Šå‚³æ¡†é¢¨æ ¼ï¼‰
# =============================
def _page_css():
    st.markdown(
        r"""
<style>
div[data-testid="stAppViewContainer"]{
  background: linear-gradient(
    180deg,
    rgba(232,245,255,1) 0%,
    rgba(244,250,255,1) 34%,
    rgba(255,255,255,1) 100%
  ) !important;
}

.qc-chips{
  margin-top: 4px;
  font-size: 12.5px;
  font-weight: 800;
  color: rgba(15,23,42,.62);
}
.qc-chips .sep{ margin: 0 8px; opacity:.55; }

.qc-u-label{
  font-size: 13.5px;
  font-weight: 900;
  color: rgba(15,23,42,.86);
  margin: 4px 0 6px 0;
}

section[data-testid="stFileUploadDropzone"]{
  border: 1px solid rgba(148,163,184,.35) !important;
  border-radius: 14px !important;
  background: rgba(255,255,255,1) !important;
  padding: 12px 14px !important;
}
section[data-testid="stFileUploadDropzone"]:hover{
  border-color: rgba(59,130,246,.35) !important;
  box-shadow: 0 6px 18px rgba(59,130,246,.08);
}

section[data-testid="stFileUploadDropzone"] button{
  border-radius: 10px !important;
  font-weight: 900 !important;
}

div[data-testid="stButton"] > button{
  border-radius: 12px !important;
  font-weight: 900 !important;
  padding: 9px 14px !important;
}

.qc-banner{
  background: rgba(219, 234, 254, .9);
  border: 1px solid rgba(59,130,246,.18);
  color: rgba(15,23,42,.86);
  border-radius: 10px;
  padding: 10px 12px;
  font-weight: 900;
  font-size: 13px;
  margin-top: 14px;
}
</style>
""",
        unsafe_allow_html=True,
    )


# =============================
# å·¥å…·ï¼šå®šä½æ¬„ä½ / æ–‡å­—æ ¼å¼
# =============================
def get_ws(wb, sheet_name: Optional[str]):
    return wb[sheet_name] if sheet_name else wb.worksheets[0]


def find_header_col(ws, header_name: str, header_row: int = 1) -> Optional[int]:
    # exact
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=c).value
        if isinstance(v, str) and v.strip() == header_name:
            return c
    # contains
    target = header_name.strip()
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=c).value
        if isinstance(v, str) and target in v.strip():
            return c
    return None


def zero_run_width(number_format: str) -> int:
    if not number_format:
        return 0
    fmt = number_format.split(";")[0]
    best = cur = 0
    for ch in fmt:
        if ch == "0":
            cur += 1
            best = max(best, cur)
        else:
            cur = 0
    return best


def normalize_code(value, fmt: str, fallback_width: int = 0) -> str:
    """åªç”¨æ–¼æ¯”å°ï¼ˆä¸å›å¯«ï¼‰"""
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")

    width = max(zero_run_width(fmt or ""), fallback_width)

    if isinstance(value, int):
        s = str(value)
        return s.zfill(width) if width >= 2 else s

    if isinstance(value, float):
        if abs(value - round(value)) < 1e-9:
            iv = int(round(value))
            s = str(iv)
            return s.zfill(width) if width >= 2 else s
        return str(value)

    return str(value).strip()


def normalize_unit(value) -> str:
    """åªç”¨æ–¼æ¯”å°ï¼ˆä¸å›å¯«ï¼‰"""
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, float) and abs(value - round(value)) < 1e-9:
        return str(int(round(value)))
    if isinstance(value, int):
        return str(value)
    return str(value).strip()


def _infer_digit_width(ws, col_idx: int, scan_limit: int = 50000) -> int:
    """æ¨æ–·ç¢¼é•·ï¼ˆæœªä¸Šæ¶æ˜ç´°ï¼‰"""
    if col_idx is None:
        return 0
    w = 0
    end_r = min(ws.max_row, scan_limit)
    for r in range(2, end_r + 1):
        cell = ws.cell(row=r, column=col_idx)
        v = cell.value
        if v is None:
            continue

        fmt = getattr(cell, "number_format", "") or ""
        w = max(w, zero_run_width(fmt))

        if isinstance(v, str):
            s = v.strip()
            if s.isdigit():
                w = max(w, len(s))
    return w


def _pad_digits_for_compare(s: str, width: int) -> str:
    """åªæ‹¿ä¾†æ¯”å°ï¼Œä¸å›å¯«åˆ° QC"""
    if not s:
        return ""
    if width >= 2 and s.isdigit():
        return s.zfill(width)
    return s


def format_date_value(v) -> str:
    if v is None:
        return ""
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    if not s:
        return ""
    try:
        dtv = pd.to_datetime(s)
        return pd.Timestamp(dtv).strftime("%Y-%m-%d")
    except Exception:
        return s


# =============================
# âœ… è¼¸å‡ºç”¨ï¼šå¯ç§»å‹•å–®ä½è£œæ»¿ 10 ç¢¼ï¼ˆä¸è¶³è£œ0ï¼‰+ è¨­æ–‡å­—
# =============================
def pad_unit_to_10(ws, unit_col: int, start_row: int = 2, width: int = 10):
    if not unit_col:
        return
    for r in range(start_row, ws.max_row + 1):
        cell = ws.cell(row=r, column=unit_col)
        v = cell.value
        if v is None:
            continue

        if isinstance(v, float) and abs(v - round(v)) < 1e-9:
            s = str(int(round(v)))
        else:
            s = str(v).strip()

        if not s:
            continue
        if s.isdigit():
            s = s.zfill(width)

        cell.value = s
        cell.number_format = "@"
        if cell.alignment is None:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


# =============================
# âœ… è¼¸å‡ºç”¨ï¼šæŠŠæŒ‡å®šæ¬„ä½å¼·åˆ¶è½‰æ–‡å­—ï¼ˆé¿å…ç§‘å­¸è¨˜è™Ÿ/æ‰0ï¼‰
# - barcode ä¸è£œ0ï¼Œåªè½‰æˆç´”æ•¸å­—å­—ä¸²
# - batch ä¾åŸæœ¬æ ¼å¼/å­—ä¸²é•·åº¦æ¨å¯¬åº¦å¾Œè£œ0
# =============================
def force_text_digits(ws, col_idx: int, pad_width: int = 0, start_row: int = 2):
    if not col_idx:
        return
    for r in range(start_row, ws.max_row + 1):
        cell = ws.cell(row=r, column=col_idx)
        v = cell.value
        if v is None:
            continue

        if isinstance(v, float) and abs(v - round(v)) < 1e-9:
            s = str(int(round(v)))
        elif isinstance(v, int):
            s = str(v)
        else:
            s = str(v).strip()

        if not s:
            continue

        # å¦‚æœæ˜¯ç´”æ•¸å­—ï¼Œæ‰è£œ0
        if pad_width >= 2 and s.isdigit():
            s = s.zfill(pad_width)

        cell.value = s
        cell.number_format = "@"
        if cell.alignment is None:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def infer_col_digit_width(ws, col_idx: int, header_row: int = 1, scan_limit: int = 50000) -> int:
    """ä¾ number_format çš„ 0000 æˆ–å­—ä¸²é•·åº¦ æ¨æ–·è©²æ¬„æœ€é•·ç¢¼é•·"""
    if not col_idx:
        return 0
    w = 0
    end_r = min(ws.max_row, scan_limit)
    for r in range(header_row + 1, end_r + 1):
        cell = ws.cell(row=r, column=col_idx)
        v = cell.value
        if v is None:
            continue
        fmt = getattr(cell, "number_format", "") or ""
        w = max(w, zero_run_width(fmt))
        if isinstance(v, str):
            s = v.strip()
            if s.isdigit():
                w = max(w, len(s))
    return w


# =============================
# âœ… åˆªé™¤æŒ‡å®šæ¬„ä½ï¼ˆè¼¸å‡ºæª”ç”¨ï¼‰
# =============================
def delete_columns_by_headers(ws, headers_to_delete, header_row: int = 1):
    if not headers_to_delete:
        return

    headers = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=c).value
        headers.append("" if v is None else str(v).strip())

    to_delete_cols = set()
    for target in headers_to_delete:
        t = str(target).strip()
        if not t:
            continue
        for idx, h in enumerate(headers, start=1):
            if not h:
                continue
            if h == t or (t in h):
                to_delete_cols.add(idx)

    for col_idx in sorted(to_delete_cols, reverse=True):
        ws.delete_cols(col_idx, 1)

    try:
        if getattr(ws, "auto_filter", None) and ws.auto_filter.ref:
            ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    except Exception:
        pass


def delete_cols_after_header(ws, header_name: str, header_row: int = 1):
    """æŠŠæŸå€‹è¡¨é ­å³é‚Šå…¨éƒ¨æ¬„ä½åˆªæ‰ï¼ˆå«é€²è²¨æ—¥å³é‚Šå…¨æ¸…ç©ºï¼‰"""
    col = find_header_col(ws, header_name, header_row)
    if col is None:
        return
    if ws.max_column > col:
        ws.delete_cols(col + 1, ws.max_column - col)


# =============================
# è½‰æ›ï¼šDataFrames -> openpyxl Workbookï¼ˆxls/xlsb ç”¨ï¼‰
# âœ… ç”¨ dtype=str å·²åœ¨è®€å–è™•è™•ç†ï¼Œé€™è£¡åªè² è²¬å¯«å…¥
# =============================
def _dfs_to_workbook(sheets: Dict[str, pd.DataFrame]) -> Workbook:
    wb = Workbook()
    if wb.worksheets:
        wb.remove(wb.worksheets[0])

    for sheet_name, df in sheets.items():
        name = str(sheet_name)[:31] if sheet_name else "Sheet1"
        ws = wb.create_sheet(title=name)

        headers = [("" if c is None else str(c)) for c in df.columns.tolist()]
        ws.append(headers)

        for row in df.itertuples(index=False, name=None):
            out_row = []
            for v in row:
                if v is None:
                    out_row.append(None)
                elif isinstance(v, float) and pd.isna(v):
                    out_row.append(None)
                else:
                    out_row.append(v)
            ws.append(out_row)

        # converted ä¾†æºï¼šå…ˆæŠŠåœ‹éš›æ¢ç¢¼/æ‰¹è™Ÿ/å¯ç§»å‹•å–®ä½éƒ½ç•¶æ–‡å­—ï¼ˆé¿å… E+12 / æ‰0ï¼‰
        hmap = {str(h).strip(): i + 1 for i, h in enumerate(headers)}
        if BARCODE_HEADER in hmap:
            force_text_digits(ws, hmap[BARCODE_HEADER], pad_width=0)
        if BATCH_HEADER in hmap:
            # æ‰¹è™Ÿåœ¨ converted ä¾†æºï¼šç›´æ¥ä¾å­—ä¸²é•·åº¦æ¨æœ€é•·è£œ0
            w = 0
            col = df[BATCH_HEADER].astype(str).tolist()
            for s in col:
                s = (s or "").strip()
                if s.isdigit():
                    w = max(w, len(s))
            force_text_digits(ws, hmap[BATCH_HEADER], pad_width=w)
        if UNIT_HEADER in hmap:
            pad_unit_to_10(ws, hmap[UNIT_HEADER], width=10)

    return wb


# =============================
# è®€å–ï¼šæ”¯æ´ xlsx/xlsm/xls/xlsbï¼ˆå›å‚³ modeï¼‰
# =============================
def _load_wb_from_upload(uploaded_file) -> Tuple[str, Workbook, str]:
    name = uploaded_file.name
    ext = (name.split(".")[-1] or "").lower()
    raw = uploaded_file.getvalue()
    bio = io.BytesIO(raw)

    if ext in ("xlsx", "xlsm"):
        keep_vba = (ext == "xlsm")
        wb = load_workbook(bio, keep_vba=keep_vba)
        return name, wb, "native"

    if ext == "xlsb":
        try:
            sheets = pd.read_excel(
                io.BytesIO(raw),
                engine="pyxlsb",
                sheet_name=None,
                dtype=str,
                keep_default_na=False,
            )
        except Exception as e:
            raise ValueError(f"è®€å– .xlsb å¤±æ•—ï¼š{e}\nè«‹ç¢ºèª requirements.txt æœ‰ pyxlsb")
        wb = _dfs_to_workbook(sheets)
        return name, wb, "converted"

    if ext == "xls":
        try:
            sheets = pd.read_excel(
                io.BytesIO(raw),
                engine="xlrd",
                sheet_name=None,
                dtype=str,
                keep_default_na=False,
            )
        except ModuleNotFoundError:
            raise ValueError(
                "ç›®å‰ç’°å¢ƒç¼ºå°‘ xlrdï¼Œç„¡æ³•è®€å– .xlsã€‚\n"
                "è«‹åœ¨ requirements.txt åŠ ä¸Šï¼šxlrd==2.0.1\n"
                "æˆ–å…ˆç”¨ Excel å¦å­˜ç‚º .xlsx å†ä¸Šå‚³ã€‚"
            )
        except Exception as e:
            raise ValueError(f"è®€å– .xls å¤±æ•—ï¼š{e}\nå»ºè­°å…ˆç”¨ Excel å¦å­˜ .xlsx å†ä¸Šå‚³ã€‚")
        wb = _dfs_to_workbook(sheets)
        return name, wb, "converted"

    raise ValueError(f"ä¸æ”¯æ´çš„æª”æ¡ˆæ ¼å¼ï¼š.{ext}\næ”¯æ´ï¼š.xlsx / .xlsm / .xls / .xlsb")


# =============================
# å°å·¥å…·ï¼šåˆªé™¤éç¬¦åˆåˆ—ï¼ˆä¿ç•™ headerï¼‰
# =============================
def _delete_non_matched_rows(ws, keep_rows, header_rows: int = 1):
    max_row = ws.max_row
    start = header_rows + 1
    keep_set = set(r for r in keep_rows if start <= r <= max_row)
    keep_sorted = sorted(keep_set)

    segs = []
    cur = start
    for k in keep_sorted:
        if cur < k:
            segs.append((cur, k - 1))
        cur = k + 1
    if cur <= max_row:
        segs.append((cur, max_row))

    for s, e in reversed(segs):
        ws.delete_rows(s, e - s + 1)

    try:
        if getattr(ws, "auto_filter", None) and ws.auto_filter.ref:
            ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    except Exception:
        pass


# =============================
# ä¸»æµç¨‹
# =============================
def process_wb(
    qc_wb,
    un_wb,
    qc_sheet_name: Optional[str] = None,
    un_sheet_name: Optional[str] = None,
) -> Tuple[int, bytes]:
    qc_ws = get_ws(qc_wb, qc_sheet_name)
    un_ws = get_ws(un_wb, un_sheet_name)

    qc_key_col = find_header_col(qc_ws, QC_KEY_HEADER, 1)
    qc_unit_col = find_header_col(qc_ws, UNIT_HEADER, 1)

    qc_barcode_col = find_header_col(qc_ws, BARCODE_HEADER, 1)
    qc_batch_col = find_header_col(qc_ws, BATCH_HEADER, 1)

    un_key_col = find_header_col(un_ws, UN_KEY_HEADER, 1)
    un_unit_col = find_header_col(un_ws, UNIT_HEADER, 1)
    un_date_col = find_header_col(un_ws, UN_DATE_HEADER, 1)

    if qc_key_col is None:
        raise ValueError(f"QC æ‰¾ä¸åˆ°æ¬„ä½ï¼š{QC_KEY_HEADER}")
    if qc_unit_col is None:
        raise ValueError(f"QC æ‰¾ä¸åˆ°æ¬„ä½ï¼š{UNIT_HEADER}")

    if un_key_col is None:
        raise ValueError(f"æœªä¸Šæ¶æ˜ç´°æ‰¾ä¸åˆ°æ¬„ä½ï¼š{UN_KEY_HEADER}")
    if un_unit_col is None:
        raise ValueError(f"æœªä¸Šæ¶æ˜ç´°æ‰¾ä¸åˆ°æ¬„ä½ï¼š{UNIT_HEADER}")
    if un_date_col is None:
        raise ValueError(f"æœªä¸Šæ¶æ˜ç´°æ‰¾ä¸åˆ°æ¬„ä½ï¼š{UN_DATE_HEADER}")

    # 1) æ¨ä¼°å•†å“ç¢¼é•·ï¼ˆæ¯”å°ç”¨ï¼‰
    code_len = 0
    for r in range(2, un_ws.max_row + 1):
        cell = un_ws.cell(row=r, column=un_key_col)
        if isinstance(cell.value, str):
            s = cell.value.strip()
            if s.isdigit():
                code_len = max(code_len, len(s))
        else:
            code_len = max(code_len, zero_run_width(getattr(cell, "number_format", "") or ""))
    fallback_width = code_len or 6

    # 2) æ¨ä¼°å¯ç§»å‹•å–®ä½ç¢¼é•·ï¼ˆæ¯”å°ç”¨ï¼‰
    unit_width = _infer_digit_width(un_ws, un_unit_col)

    # 3) å»ºç´¢å¼•ï¼š(å•†å“ç¢¼, å¯ç§»å‹•å–®ä½) -> é€²è²¨æ—¥
    date_sets = defaultdict(set)
    for r in range(2, un_ws.max_row + 1):
        code_cell = un_ws.cell(row=r, column=un_key_col)
        code = normalize_code(code_cell.value, getattr(code_cell, "number_format", ""), fallback_width)
        if code and code.isdigit():
            code = code.zfill(fallback_width)

        unit_cell = un_ws.cell(row=r, column=un_unit_col)
        unit = normalize_unit(unit_cell.value)
        unit = _pad_digits_for_compare(unit, unit_width)

        d_cell = un_ws.cell(row=r, column=un_date_col)
        d_str = format_date_value(d_cell.value)

        if code and unit and d_str:
            date_sets[(code, unit)].add(d_str)

    date_map: Dict[Tuple[str, str], str] = {k: "ã€".join(sorted(v)) for k, v in date_sets.items()}

    # 4) æ–°å¢/å®šä½ã€Œé€²è²¨æ—¥ã€
    qc_date_col = find_header_col(qc_ws, "é€²è²¨æ—¥", 1)
    if qc_date_col is None:
        qc_date_col = qc_ws.max_column + 1
        hdr = qc_ws.cell(row=1, column=qc_date_col, value="é€²è²¨æ—¥")
        src_hdr = qc_ws.cell(row=1, column=qc_unit_col)
        try:
            hdr._style = _copy.copy(src_hdr._style)
        except Exception:
            pass
        hdr.alignment = _copy.copy(getattr(src_hdr, "alignment", Alignment(horizontal="center", vertical="center")))

    # 5) å¡«å…¥é€²è²¨æ—¥ + æ”¶é›† match rows
    match_rows = []
    for r in range(2, qc_ws.max_row + 1):
        code_cell = qc_ws.cell(row=r, column=qc_key_col)
        code = normalize_code(code_cell.value, getattr(code_cell, "number_format", ""), fallback_width)
        if code and code.isdigit():
            code = code.zfill(fallback_width)

        unit_cell = qc_ws.cell(row=r, column=qc_unit_col)
        unit = normalize_unit(unit_cell.value)
        unit = _pad_digits_for_compare(unit, unit_width)

        d_str = date_map.get((code, unit), "")

        out_cell = qc_ws.cell(row=r, column=qc_date_col)
        out_cell.value = d_str
        out_cell.number_format = "@"
        out_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        if d_str:
            match_rows.append(r)

    # âœ… 5.5) å¯ç§»å‹•å–®ä½ï¼šè£œæ»¿ 10 ç¢¼ + è¨­æ–‡å­—
    pad_unit_to_10(qc_ws, qc_unit_col, start_row=2, width=10)

    # âœ… 5.6) åœ‹éš›æ¢ç¢¼ï¼šå¼·åˆ¶æ–‡å­—ï¼ˆé¿å… E+12ï¼‰
    if qc_barcode_col is not None:
        force_text_digits(qc_ws, qc_barcode_col, pad_width=0)

    # âœ… 5.7) æ‰¹è™Ÿï¼šä¾æ¬„ä½æœ€é•·ç¢¼/number_format æ¨å¯¬åº¦å¾Œè£œ0 + è¨­æ–‡å­—
    if qc_batch_col is not None:
        batch_w = infer_col_digit_width(qc_ws, qc_batch_col)
        force_text_digits(qc_ws, qc_batch_col, pad_width=batch_w)

    # âœ… 5.8) åˆªé™¤ä½ æŒ‡å®šçš„æ¬„ä½
    delete_columns_by_headers(qc_ws, DELETE_HEADERS, header_row=1)

    # âœ… 5.9) åªä¿ç•™åˆ°ã€Œé€²è²¨æ—¥ã€ç‚ºæ­¢ï¼ˆå³é‚Šå…¨éƒ¨åˆªæ‰ â†’ è¼¸å‡ºå°±æœƒåƒä½ åœ–ï¼‰
    delete_cols_after_header(qc_ws, "é€²è²¨æ—¥", header_row=1)

    # 6) ç”¢ç”Ÿç¬¦åˆåˆ†é ï¼šè¤‡è£½ QC â†’ åˆªé™¤ä¸ç¬¦åˆåˆ—ï¼ˆç‰ˆé¢æœ€å¤§ç¨‹åº¦ä¸€è‡´ï¼‰
    if MATCH_SHEET_NAME in qc_wb.sheetnames:
        del qc_wb[MATCH_SHEET_NAME]

    mws = qc_wb.copy_worksheet(qc_ws)
    mws.title = MATCH_SHEET_NAME
    _delete_non_matched_rows(mws, keep_rows=match_rows, header_rows=1)

    out = io.BytesIO()
    qc_wb.save(out)
    out.seek(0)
    return len(match_rows), out.getvalue()


# =============================
# Streamlit UI
# =============================
st.set_page_config(page_title="å¤§è±ç‰©æµ - é€²è²¨èª²ï½œQCæœªä¸Šæ¶æ¯”å°", page_icon="ğŸ§¾", layout="wide")
inject_logistics_theme()
_page_css()

set_page(
    "QC æœªä¸Šæ¶æ¯”å°",
    icon="ğŸ§¾",
    subtitle="æ¯”å°ï¼šQCã€Œå•†å“+å¯ç§»å‹•å–®ä½ã€= æœªä¸Šæ¶ã€Œå•†å“ç¢¼+å¯ç§»å‹•å–®ä½ã€ï¼›è¼¸å‡ºï¼šå¯ç§»å‹•å–®ä½è£œ10ç¢¼ã€åˆªæŒ‡å®šæ¬„ä½ã€åªä¿ç•™åˆ°é€²è²¨æ—¥ã€‚",
)

st.markdown(
    '<div class="qc-chips">é›™æ¢ä»¶æ¯”å°<span class="sep">ï½œ</span>å¯ç§»å‹•å–®ä½è£œ10ç¢¼<span class="sep">ï½œ</span>åœ‹éš›æ¢ç¢¼/æ‰¹è™Ÿæ–‡å­—åŒ–<span class="sep">ï½œ</span>åªä¿ç•™åˆ°é€²è²¨æ—¥</div>',
    unsafe_allow_html=True,
)

card_open("ğŸ“ æª”æ¡ˆä¸Šå‚³")

st.markdown('<div class="qc-u-label">QC æ˜ç´°ï¼ˆæ”¯æ´ï¼š.xlsx / .xlsm / .xls / .xlsbï¼‰</div>', unsafe_allow_html=True)
qc_file = st.file_uploader(
    "QC æ˜ç´°",
    type=["xlsx", "xlsm", "xls", "xlsb"],
    accept_multiple_files=False,
    label_visibility="collapsed",
    key="qc_file",
)

st.markdown('<div style="height:10px"></div>', unsafe_allow_html=True)

st.markdown('<div class="qc-u-label">æœªä¸Šæ¶æ˜ç´°ï¼ˆæ”¯æ´ï¼š.xlsx / .xlsm / .xls / .xlsbï¼‰</div>', unsafe_allow_html=True)
un_file = st.file_uploader(
    "æœªä¸Šæ¶æ˜ç´°",
    type=["xlsx", "xlsm", "xls", "xlsb"],
    accept_multiple_files=False,
    label_visibility="collapsed",
    key="un_file",
)

qc_sheet_name = None
un_sheet_name = None

with st.expander("é€²éšè¨­å®šï¼ˆå·¥ä½œè¡¨é¸æ“‡ï¼‰", expanded=False):
    c1, c2 = st.columns(2)
    with c1:
        if qc_file:
            try:
                _, qc_wb_preview, _ = _load_wb_from_upload(qc_file)
                qc_sheet_name = st.selectbox("QC å·¥ä½œè¡¨", options=qc_wb_preview.sheetnames, index=0)
            except Exception as e:
                st.error(str(e))
    with c2:
        if un_file:
            try:
                _, un_wb_preview, _ = _load_wb_from_upload(un_file)
                un_sheet_name = st.selectbox("æœªä¸Šæ¶æ˜ç´° å·¥ä½œè¡¨", options=un_wb_preview.sheetnames, index=0)
            except Exception as e:
                st.error(str(e))

ready = bool(qc_file and un_file)
run = st.button("ğŸš€ ç”¢å‡ºæ¯”å°", disabled=not ready)

card_close()

status_msg = "è«‹ä¾åºä¸Šå‚³ï¼šQC æ˜ç´° + æœªä¸Šæ¶æ˜ç´°"
xlsx_bytes = None
matched = None

if ready:
    status_msg = "æª”æ¡ˆå·²å°±ç·’ï¼Œå¯æŒ‰ã€Œç”¢å‡ºæ¯”å°ã€"

if run:
    try:
        with st.spinner("è™•ç†ä¸­â€¦"):
            _, qc_wb, _ = _load_wb_from_upload(qc_file)
            _, un_wb, _ = _load_wb_from_upload(un_file)

            matched, xlsx_bytes = process_wb(
                qc_wb=qc_wb,
                un_wb=un_wb,
                qc_sheet_name=qc_sheet_name,
                un_sheet_name=un_sheet_name,
            )
    except Exception as e:
        st.error(f"âŒ åŸ·è¡Œå¤±æ•—ï¼š{e}")

if xlsx_bytes is not None:
    card_open("âœ… ç”¢å‡ºçµæœ")
    st.success(f"å®Œæˆï¼ç¬¦åˆç­†æ•¸ï¼š{matched}")
    out_name = f"QCæœªä¸Šæ¶æ¯”å°_è¼¸å‡º_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    st.download_button(
        label="ğŸ“¥ ä¸‹è¼‰è¼¸å‡º Excel",
        data=xlsx_bytes,
        file_name=out_name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    card_close()

st.markdown(f'<div class="qc-banner">{status_msg}</div>', unsafe_allow_html=True)
