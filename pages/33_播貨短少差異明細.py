import pandas as pd
import streamlit as st
from datetime import datetime
from io import BytesIO

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

from common_ui import inject_logistics_theme, set_page, card_open, card_close


st.set_page_config(
    page_title="播貨短少差異明細",
    page_icon="📍",
    layout="wide"
)

inject_logistics_theme()
set_page("播貨短少差異明細", "上傳短少明細、庫存明細、儲位明細，自動產出揀差異明細")


# =========================
# 基本工具
# =========================

def read_excel_first_sheet(uploaded_file):
    return pd.read_excel(uploaded_file, dtype=str)


def clean_columns(df):
    df = df.copy()
    df.columns = [
        str(c).strip().replace("\n", "").replace("\r", "").replace(" ", "")
        for c in df.columns
    ]
    return df


def find_col(df, candidates, required=True):
    cols = list(df.columns)

    for name in candidates:
        if name in cols:
            return name

    for col in cols:
        for name in candidates:
            if name in col:
                return col

    if required:
        raise ValueError(f"找不到欄位：{candidates}")

    return None


def clean_text(value):
    if pd.isna(value):
        return ""

    text = str(value).strip()

    if text.endswith(".0"):
        text = text[:-2]

    return text


def to_number(value):
    if pd.isna(value):
        return 0

    text = str(value).replace(",", "").strip()

    if text == "":
        return 0

    try:
        return float(text)
    except Exception:
        return 0


def to_date(value):
    if pd.isna(value):
        return pd.Timestamp.max

    try:
        return pd.to_datetime(value)
    except Exception:
        return pd.Timestamp.max


def parse_report_date(value):
    if pd.isna(value):
        return pd.NaT

    text = str(value).strip()
    text = text.replace("上午", "AM")
    text = text.replace("下午", "PM")

    return pd.to_datetime(text, errors="coerce")


def format_qty(value):
    num = to_number(value)

    if float(num).is_integer():
        return int(num)

    return num


# =========================
# 儲位挑選邏輯
# =========================

def pick_locations(stock_rows, col_location, col_canuseqty, col_calculated, col_expiry):
    """
    儲位規則：

    儲位1：
        Canuseqty = 0，短效期優先

    儲位2：
        短效期優先後，取 已試算 最大值

    儲位3：
        短效期優先後，取 已試算 第二大值

    若前面沒有符合，就往前遞補。
    """

    selected = []

    if stock_rows.empty:
        return ["", "", ""]

    data = stock_rows.copy()

    data["_Canuseqty_num"] = data[col_canuseqty].apply(to_number)
    data["_已試算_num"] = data[col_calculated].apply(to_number)

    if col_expiry:
        data["_效期_sort"] = data[col_expiry].apply(to_date)
    else:
        data["_效期_sort"] = pd.Timestamp.max

    data[col_location] = data[col_location].fillna("").astype(str).str.strip()

    rule1 = data[
        (data["_Canuseqty_num"] == 0) &
        (data[col_location] != "")
    ].copy()

    if not rule1.empty:
        rule1 = rule1.sort_values(
            by=["_效期_sort", "_已試算_num"],
            ascending=[True, False]
        )

        loc = clean_text(rule1.iloc[0][col_location])

        if loc not in selected:
            selected.append(loc)

    rule23 = data[data[col_location] != ""].copy()

    if not rule23.empty:
        rule23 = rule23.sort_values(
            by=["_效期_sort", "_已試算_num"],
            ascending=[True, False]
        )

        for _, row in rule23.iterrows():
            loc = clean_text(row[col_location])

            if loc == "":
                continue

            if loc in selected:
                continue

            selected.append(loc)

            if len(selected) >= 3:
                break

    while len(selected) < 3:
        selected.append("")

    return selected[:3]


# =========================
# Excel 輸出格式
# =========================

def format_excel_bytes(df):
    output = BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="揀差異明細")

    output.seek(0)

    wb = load_workbook(output)
    ws = wb["揀差異明細"]

    max_row = ws.max_row
    max_col = ws.max_column

    header_fill = PatternFill("solid", fgColor="0F766E")
    header_font = Font(color="FFFFFF", bold=True)
    row_fill = PatternFill("solid", fgColor="EAF1DD")
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row in range(2, max_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = border
            cell.alignment = Alignment(vertical="center")

            if row % 2 == 0:
                cell.fill = row_fill

    widths = {
        "A": 12,
        "B": 18,
        "C": 12,
        "D": 45,
        "E": 14,
        "F": 14,
        "G": 14,
        "H": 14,
        "I": 12,
    }

    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    if max_row >= 2:
        table_ref = f"A1:{get_column_letter(max_col)}{max_row}"
        tab = Table(displayName="PickDiffTable", ref=table_ref)

        style = TableStyleInfo(
            name="TableStyleMedium4",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )

        tab.tableStyleInfo = style
        ws.add_table(tab)

    ws.freeze_panes = "A2"

    final_output = BytesIO()
    wb.save(final_output)
    final_output.seek(0)

    return final_output


# =========================
# 主處理
# =========================

def build_pick_diff(shortage_file, stock_file, location_file):
    shortage_df = clean_columns(read_excel_first_sheet(shortage_file))
    stock_df = clean_columns(read_excel_first_sheet(stock_file))
    location_df = clean_columns(read_excel_first_sheet(location_file))

    # 短少明細欄位
    shortage_time_col = find_col(shortage_df, ["回報時間", "時間"])
    shortage_qty_col = find_col(shortage_df, ["數量", "短少數量", "Qty"])
    shortage_item_col = find_col(shortage_df, ["商品碼", "商品號", "品號"])
    shortage_barcode_col = find_col(shortage_df, ["條碼", "國際條碼"], required=False)

    # 庫存明細欄位
    stock_item_col = find_col(stock_df, ["商品號", "商品碼", "商品編號", "品號"])
    stock_name_col = find_col(stock_df, ["全名", "商品名稱"])
    stock_location_col = find_col(stock_df, ["儲位", "儲位號", "儲位名稱"])
    stock_canuseqty_col = find_col(stock_df, ["Canuseqty", "可用量", "可用庫存"])
    stock_calculated_col = find_col(stock_df, ["已試算", "試算"])
    stock_expiry_col = find_col(stock_df, ["效期", "有效日期", "保存期限", "到期日"], required=False)

    # 儲位明細欄位
    location_loc_col = find_col(location_df, ["儲位", "儲位號", "儲位名稱"])
    location_shed_col = find_col(location_df, ["棚別", "棚"])

    # 只抓當天日期
    today = datetime.today().date()

    shortage_df["_回報日期"] = shortage_df[shortage_time_col].apply(parse_report_date).dt.date
    shortage_df = shortage_df[shortage_df["_回報日期"] == today].copy()

    if shortage_df.empty:
        return pd.DataFrame(), today

    # 同一天同商品碼加總
    shortage_df[shortage_item_col] = shortage_df[shortage_item_col].apply(clean_text)
    shortage_df[shortage_qty_col] = shortage_df[shortage_qty_col].apply(to_number)

    agg_dict = {
        shortage_qty_col: "sum"
    }

    if shortage_barcode_col:
        shortage_df[shortage_barcode_col] = shortage_df[shortage_barcode_col].apply(clean_text)
        agg_dict[shortage_barcode_col] = "first"

    shortage_df = (
        shortage_df
        .groupby(shortage_item_col, as_index=False)
        .agg(agg_dict)
    )

    # 清理庫存與儲位資料
    stock_df[stock_item_col] = stock_df[stock_item_col].apply(clean_text)
    stock_df[stock_location_col] = stock_df[stock_location_col].fillna("").astype(str).str.strip()

    location_df[location_loc_col] = location_df[location_loc_col].fillna("").astype(str).str.strip()
    location_df[location_shed_col] = location_df[location_shed_col].fillna("").astype(str).str.strip()

    shed_map = dict(zip(location_df[location_loc_col], location_df[location_shed_col]))

    result_rows = []

    for _, row in shortage_df.iterrows():
        item_no = clean_text(row[shortage_item_col])
        shortage_qty = format_qty(row[shortage_qty_col])

        if shortage_barcode_col:
            barcode = clean_text(row[shortage_barcode_col])
        else:
            barcode = ""

        if item_no == "":
            continue

        matched_stock = stock_df[
            stock_df[stock_item_col] == item_no
        ].copy()

        if matched_stock.empty:
            result_rows.append({
                "短少數量": shortage_qty,
                "條碼": barcode,
                "商品號": item_no,
                "全名": "",
                "儲位1": "",
                "儲位2": "",
                "儲位3": "",
                "儲位1棚別": "",
                "比對結果": "未比對",
            })
            continue

        product_name = matched_stock.iloc[0][stock_name_col]

        loc1, loc2, loc3 = pick_locations(
            stock_rows=matched_stock,
            col_location=stock_location_col,
            col_canuseqty=stock_canuseqty_col,
            col_calculated=stock_calculated_col,
            col_expiry=stock_expiry_col,
        )

        shed1 = shed_map.get(loc1, "")
        compare_result = "已比對" if loc1 != "" else "未比對"

        result_rows.append({
            "短少數量": shortage_qty,
            "條碼": barcode,
            "商品號": item_no,
            "全名": product_name,
            "儲位1": loc1,
            "儲位2": loc2,
            "儲位3": loc3,
            "儲位1棚別": shed1,
            "比對結果": compare_result,
        })

    result_df = pd.DataFrame(result_rows)

    output_columns = [
        "短少數量",
        "條碼",
        "商品號",
        "全名",
        "儲位1",
        "儲位2",
        "儲位3",
        "儲位1棚別",
        "比對結果",
    ]

    result_df = result_df[output_columns]

    return result_df, today


# =========================
# Streamlit 畫面
# =========================

card_open("📍 播貨短少差異明細")

st.markdown(
    """
請依序上傳三個檔案，系統會自動：

1. 只抓取今天回報的短少資料  
2. 同商品碼自動加總短少數量  
3. 用商品碼比對庫存明細  
4. 依規則產出儲位1、儲位2、儲位3  
5. 依儲位1帶出棚別  
"""
)

col1, col2, col3 = st.columns(3)

with col1:
    shortage_file = st.file_uploader(
        "① 上傳短少明細",
        type=["xlsx", "xls"],
        key="shortage_file",
    )

with col2:
    stock_file = st.file_uploader(
        "② 上傳庫存明細",
        type=["xlsx", "xls"],
        key="stock_file",
    )

with col3:
    location_file = st.file_uploader(
        "③ 上傳儲位明細",
        type=["xlsx", "xls"],
        key="location_file",
    )

st.divider()

if st.button("產生揀差異明細", type="primary", use_container_width=True):
    if not shortage_file or not stock_file or not location_file:
        st.warning("請先上傳【短少明細】、【庫存明細】、【儲位明細】三個檔案。")
    else:
        try:
            result_df, today = build_pick_diff(
                shortage_file=shortage_file,
                stock_file=stock_file,
                location_file=location_file,
            )

            if result_df.empty:
                st.warning(f"短少明細中沒有找到今天日期的資料：{today}")
            else:
                st.success(f"完成！已產生 {len(result_df):,} 筆揀差異明細。")

                st.dataframe(
                    result_df,
                    use_container_width=True,
                    hide_index=True,
                )

                excel_file = format_excel_bytes(result_df)

                st.download_button(
                    label="下載揀差異明細 Excel",
                    data=excel_file,
                    file_name=f"揀差異明細_{today}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )

        except Exception as e:
            st.error(f"產生失敗：{e}")

card_close()
