# pages/3_總揀達標.py
# ------------------------------------------------------------
#  總揀達標獎金計算報表（合併版：上午 + 下午同頁呈現）
#  - 第一階段：上午（<=12:30，休息 10:00-10:15）
#  - 第二階段：下午（13:30-18:00，休息 15:30-15:45）
#  - 版面：同一個 Sheet1 上下分段（符合你截圖的呈現方式）
#  - 匯出：openpyxl（避免 Streamlit Cloud 缺 xlsxwriter）
#  v2025-12-22 (streamlit)
# ------------------------------------------------------------

from __future__ import annotations

import io
from dataclasses import dataclass
from datetime import datetime, timedelta
from typing import Dict, List, Tuple, Optional

import pandas as pd
import streamlit as st

from common_ui import inject_logistics_theme, set_page, card_open, card_close

# ---------- 可視參數 -------------------------------------------------
# 上午
MORNING_END = datetime.strptime("12:30:00", "%H:%M:%S").time()
M_REST_START = datetime.strptime("10:00:00", "%H:%M:%S").time()
M_REST_END = datetime.strptime("10:15:00", "%H:%M:%S").time()

# 下午
AFTERNOON_START = datetime.strptime("13:30:00", "%H:%M:%S").time()
AFTERNOON_END = datetime.strptime("18:00:00", "%H:%M:%S").time()
A_REST_START = datetime.strptime("15:30:00", "%H:%M:%S").time()
A_REST_END = datetime.strptime("15:45:00", "%H:%M:%S").time()

IDLE_THRESHOLD = timedelta(minutes=10)  # 空窗門檻
default_start_time_str = "08:05:00"

# ---------- 揀貨人預設資料 ------------------------------------------
# 若「區域」留空 → 以「低空」處理
# （此段來自你提供的合併版，原樣保留）
preset_picker_info: Dict[str, Dict[str, str]] = {
    "20230412002": {"姓名": "吳秉丞", "起始時間": "8:05:00", "區域": "低空"},
    "20200812002": {"姓名": "彭慈暉", "起始時間": "7:05:00", "區域": "低空"},
    "20210104001": {"姓名": "楊承珉", "起始時間": "7:05:00", "區域": "低空"},
    "20201109001": {"姓名": "梁冠如", "起始時間": "8:05:00", "區域": "低空"},
    "20201109003": {"姓名": "吳振凱", "起始時間": "8:05:00", "區域": "低空"},
    "20231226003": {"姓名": "顏秀菁", "起始時間": "8:05:00", "區域": "低空"},
    "20200922002": {"姓名": "葉欲弘", "起始時間": "8:05:00", "區域": "低空"},
    "20200924001": {"姓名": "黃雅君", "起始時間": "8:05:00", "區域": "低空"},
    "20201019001": {"姓名": "邱清瑞", "起始時間": "8:05:00", "區域": "低空"},
    "20220526001": {"姓名": "黃芷憶", "起始時間": "8:05:00", "區域": "低空"},
    "20240221003": {"姓名": "呂治明", "起始時間": "8:05:00", "區域": "低空"},
    "20240909001": {"姓名": "蔡麗珠", "起始時間": "8:05:00", "區域": "低空"},
    "20240926001": {"姓名": "陳莉娜", "起始時間": "8:05:00", "區域": "低空"},
    "20241011002": {"姓名": "林雙慧", "起始時間": "8:05:00", "區域": "低空"},
    "20250326001": {"姓名": "王大中", "起始時間": "8:05:00", "區域": "低空"},
    "20250303002": {"姓名": "周映華", "起始時間": "8:05:00", "區域": "低空"},
    "20250311001": {"姓名": "徐欣", "起始時間": "8:05:00", "區域": "低空"},
    "20250226002": {"姓名": "阮黃英", "起始時間": "7:05:00", "區域": "低空"},
    "20250901009": {"姓名": "張寶萱", "起始時間": "8:35:00", "區域": "低空"},
    "20250226010": {"姓名": "楊心如", "起始時間": "7:05:00", "區域": "低空"},
    "20250226011": {"姓名": "阮武玉玄", "起始時間": "7:05:00", "區域": "低空"},
    "20250226016": {"姓名": "阮氏美麗", "起始時間": "7:05:00", "區域": "低空"},
    "20250226018": {"姓名": "阮瑞美黃緣", "起始時間": "7:05:00", "區域": "低空"},
    "20250226020": {"姓名": "潘氏慶平", "起始時間": "7:05:00", "區域": "低空"},
    "20250226021": {"姓名": "潘氏青江", "起始時間": "7:05:00", "區域": "低空"},
    "20250923019": {"姓名": "阮氏紅深", "起始時間": "8:05:00", "區域": "低空"},
    "20250226026": {"姓名": "黎氏瓊", "起始時間": "7:05:00", "區域": "低空"},
    "20191205002": {"姓名": "阮功水", "起始時間": "8:05:00", "區域": "低空"},
    "20230119001": {"姓名": "陶春青", "起始時間": "7:05:00", "區域": "高空"},
    "20210318001": {"姓名": "陳文勇", "起始時間": "8:05:00", "區域": "低空"},
    "20210805001": {"姓名": "郭中合", "起始時間": "8:05:00", "區域": "低空"},
    "20220421002": {"姓名": "楊文點", "起始時間": "8:05:00", "區域": "低空"},
    "20220505001": {"姓名": "阮伊黃", "起始時間": "8:05:00", "區域": "低空"},
    "20220505002": {"姓名": "阮文青明", "起始時間": "7:05:00", "區域": "高空"},
    "20221222005": {"姓名": "謝忠龍", "起始時間": "8:05:00", "區域": "高空"},
    "20221222009": {"姓名": "潘文一", "起始時間": "8:05:00", "區域": "低空"},
    "20221221001": {"姓名": "阮文全", "起始時間": "7:05:00", "區域": "高空"},
    "20230504001": {"姓名": "黃文重", "起始時間": "8:05:00", "區域": "低空"},
    "20230511003": {"姓名": "范日明", "起始時間": "7:05:00", "區域": "低空"},
    "20230810003": {"姓名": "范明俊", "起始時間": "8:05:00", "區域": "低空"},
    "20231211004": {"姓名": "河文南", "起始時間": "8:05:00", "區域": "低空"},
    "20231218004": {"姓名": "河文強", "起始時間": "8:05:00", "區域": "低空"},
    "20240107001": {"姓名": "范文春", "起始時間": "8:05:00", "區域": "低空"},
    "20240313001": {"姓名": "陳文越", "起始時間": "8:05:00", "區域": "低空"},
    "20240313003": {"姓名": "阮曰忠", "起始時間": "7:05:00", "區域": "高空"},
    "20240730001": {"姓名": "阮文忠", "起始時間": "8:05:00", "區域": "低空"},
    "20241204005": {"姓名": "阮春水", "起始時間": "7:05:00", "區域": "低空"},
    "20241204007": {"姓名": "阮玉名", "起始時間": "8:05:00", "區域": "低空"},
    "20241204009": {"姓名": "阮長文", "起始時間": "7:05:00", "區域": "低空"},
    "20220421001": {"姓名": "阮德平", "起始時間": "8:05:00", "區域": "高空"},
    "20250502001": {"姓名": "吳詩敏", "起始時間": "8:05:00", "區域": "低空"},
    "20250617003": {"姓名": "喬家寶", "起始時間": "8:05:00", "區域": "低空"},
    "20250901011": {"姓名": "章愛玲", "起始時間": "8:35:00", "區域": "低空"},
    "20250617001": {"姓名": "阮文譚", "起始時間": "7:05:00", "區域": "高空"},
    "09963": {"姓名": "黃謙凱", "起始時間": "8:05:00", "區域": "低空"},
    "11399": {"姓名": "陳哲沅", "起始時間": "8:05:00", "區域": "低空"},
}

# =========================================================
# Utils
# =========================================================
def parse_tw_datetime(series: pd.Series) -> pd.Series:
    """
    同時支援：
      1. 2025/06/26 上午 09:35:01  （中文 AM/PM）
      2. 2025/6/30 10:37:51       （24h 制字串）
      3. 45549.435694444          （Excel 浮點序列）
    解析失敗 → NaT
    """
    if pd.api.types.is_datetime64_any_dtype(series):
        return series

    s = series.astype(str).str.strip()
    out = pd.Series(pd.NaT, index=s.index, dtype="datetime64[ns]")

    # Excel 浮點
    num_mask = s.str.match(r"^\d+(\.\d+)?$")
    if num_mask.any():
        out.loc[num_mask] = pd.to_datetime(
            s[num_mask].astype(float), unit="d", origin="1899-12-30"
        )

    # 字串解析
    str_mask = ~num_mask
    if str_mask.any():
        tmp = s[str_mask]
        pm_mask = tmp.str.contains("下午")

        tmp = (
            tmp.str.replace("上午", "", regex=False)
            .str.replace("下午", "", regex=False)
            .str.replace(r"\s+", " ", regex=True)
            .str.strip()
        )

        parsed = pd.to_datetime(tmp, format="%Y/%m/%d %H:%M:%S", errors="coerce")
        need_fallback = parsed.isna()
        if need_fallback.any():
            parsed.loc[need_fallback] = pd.to_datetime(tmp[need_fallback], errors="coerce")

        # 原字串有「下午」且解析結果 <12 點 → +12h
        if pm_mask.any():
            pm_idx = pm_mask[pm_mask].index
            adjust_idx = pm_idx[parsed.loc[pm_idx].dt.hour < 12]
            parsed.loc[adjust_idx] += pd.Timedelta(hours=12)

        out.loc[str_mask] = parsed

    return out


def _get_region(picker: str, mapping: Dict[str, Dict[str, str]]) -> str:
    return (mapping.get(picker, {}).get("區域", "低空") or "低空").strip() or "低空"


def _get_name(picker: str, mapping: Dict[str, Dict[str, str]]) -> str:
    return (mapping.get(picker, {}).get("姓名", picker) or picker).strip() or picker


def _get_start_time_str(picker: str, mapping: Dict[str, Dict[str, str]]) -> str:
    return (mapping.get(picker, {}).get("起始時間", default_start_time_str) or default_start_time_str).strip()


def _storage_area_str(working_records: pd.DataFrame) -> str:
    storage_prefixes: List[str] = []
    for loc in working_records.get("儲位", []):
        prefix = str(loc)[:3]
        if prefix not in storage_prefixes:
            storage_prefixes.append(prefix)
    return ",".join(storage_prefixes)


def split_idle_segment(seg_start, seg_end, rest_start, rest_end):
    segments = []
    if seg_end <= rest_start or seg_start >= rest_end:
        segments.append((seg_start, seg_end))
    else:
        if seg_start < rest_start:
            segments.append((seg_start, rest_start))
        if seg_end > rest_end:
            segments.append((rest_end, seg_end))
    return segments


def get_effective_idle_segments(seg_start, seg_end, rest_start, rest_end, threshold=IDLE_THRESHOLD):
    segments = split_idle_segment(seg_start, seg_end, rest_start, rest_end)
    return [(s, e) for s, e in segments if (e - s) >= threshold]


def ensure_datetime(full_df: pd.DataFrame) -> pd.DataFrame:
    if not pd.api.types.is_datetime64_any_dtype(full_df["揀貨完成時間"]):
        full_df = full_df.copy()
        full_df["揀貨完成時間"] = parse_tw_datetime(full_df["揀貨完成時間"])
    return full_df


def remove_boxed_rows(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    if "成箱箱號" in df.columns:
        df["成箱箱號"] = df["成箱箱號"].astype(str).str.strip()
        df = df[df["成箱箱號"] == ""]
    return df


def combine_rows(df: pd.DataFrame) -> pd.DataFrame:
    # 與你合併版一致：同儲位/商品/揀貨人/完成時間 → 數量加總
    group_cols = ["儲位", "商品", "揀貨人", "揀貨完成時間"]
    if "數量" not in df.columns:
        # 若來源沒有數量，仍保留原筆數統計（不影響你原本主邏輯：以筆數做效率）
        df = df.copy()
        df["數量"] = 1
    combined_df = df.groupby(group_cols, as_index=False).agg({"數量": "sum"})
    return combined_df


def filter_morning_period(df: pd.DataFrame) -> pd.DataFrame:
    dtv = parse_tw_datetime(df["揀貨完成時間"])
    df = df.assign(揀貨完成時間=dtv).dropna(subset=["揀貨完成時間"])
    df = df[df["揀貨完成時間"].dt.time <= MORNING_END]
    return df


def filter_afternoon_period(df: pd.DataFrame) -> pd.DataFrame:
    dtv = parse_tw_datetime(df["揀貨完成時間"])
    df = df.assign(揀貨完成時間=dtv).dropna(subset=["揀貨完成時間"])
    df = df[(df["揀貨完成時間"].dt.time >= AFTERNOON_START) & (df["揀貨完成時間"].dt.time <= AFTERNOON_END)]
    return df


# =========================================================
# 計算：上午 / 下午（保留你合併版邏輯）
# =========================================================
def calculate_statistics_morning(morning_df: pd.DataFrame, full_df: pd.DataFrame, mapping: Dict[str, Dict[str, str]]) -> pd.DataFrame:
    """
    早班（保留你原本上午版邏輯）：
    - effective_start = min(第一筆, 設定起始時間)
    - effective_end：若有下午紀錄 -> 12:30；否則 min(最後一筆, 12:30)
    - 休息：10:00-10:15（交集才扣）
    - 空窗：>=10 分，且切掉休息重疊
    - 結尾空窗：只有「有下午紀錄」才補到 12:30
    """
    full_df = ensure_datetime(full_df)

    if morning_df.empty:
        raise ValueError("早班未找到任何有效紀錄，請檢查時間格式或篩選條件")

    stats = []
    pickers = morning_df["揀貨人"].unique()

    for picker in pickers:
        picker_morning = morning_df[morning_df["揀貨人"] == picker].sort_values("揀貨完成時間")
        times = list(picker_morning["揀貨完成時間"])
        if not times:
            continue

        config_time = datetime.strptime(_get_start_time_str(picker, mapping), "%H:%M:%S").time()
        first_record = times[0]
        last_record = times[-1]

        configured_start = datetime.combine(first_record.date(), config_time)
        effective_start = min(first_record, configured_start)

        morning_end_dt = datetime.combine(first_record.date(), MORNING_END)

        picker_full = full_df[full_df["揀貨人"] == picker]
        has_afternoon = any(rec.time() > MORNING_END for rec in picker_full["揀貨完成時間"])
        effective_end = morning_end_dt if has_afternoon else min(last_record, morning_end_dt)

        rest_start_dt = datetime.combine(effective_start.date(), M_REST_START)
        rest_end_dt = datetime.combine(effective_start.date(), M_REST_END)
        overlap_start = max(effective_start, rest_start_dt)
        overlap_end = min(effective_end, rest_end_dt)
        rest_duration = (overlap_end - overlap_start) if overlap_end > overlap_start else timedelta(0)

        work_duration = (effective_end - effective_start) - rest_duration
        total_minutes = round(work_duration.total_seconds() / 60, 2)

        num_records = len(picker_morning)

        idle_segments: List[Tuple[datetime, datetime]] = []
        # 開頭空窗
        if times[0] > effective_start:
            idle_segments.extend(split_idle_segment(effective_start, times[0], rest_start_dt, rest_end_dt))
        # 中間空窗 >=10
        for i in range(1, len(times)):
            idle_segments.extend(get_effective_idle_segments(times[i - 1], times[i], rest_start_dt, rest_end_dt))
        # 結尾空窗：只有有下午紀錄才補到 12:30
        if last_record < morning_end_dt:
            overall_max = picker_full["揀貨完成時間"].max()
            if last_record != overall_max:
                idle_segments.extend(get_effective_idle_segments(last_record, morning_end_dt, rest_start_dt, rest_end_dt))

        idle_minutes = round(sum((e - s).total_seconds() for s, e in idle_segments) / 60, 2)
        efficiency = round((num_records / total_minutes * 60) if total_minutes else 0, 2)

        time_period_str = f"{effective_start.strftime('%H:%M:%S')} ~ {effective_end.strftime('%H:%M:%S')}"
        idle_segments_str = "; ".join(f"{s.strftime('%H:%M:%S')} ~ {e.strftime('%H:%M:%S')}" for s, e in idle_segments)

        working_records = picker_morning[
            (picker_morning["揀貨完成時間"] >= effective_start) & (picker_morning["揀貨完成時間"] <= effective_end)
        ]
        storage_area_str = _storage_area_str(working_records)
        region = _get_region(picker, mapping)

        stats.append(
            {
                "區域": region,
                "揀貨人": picker,
                "姓名": _get_name(picker, mapping),
                "筆數": num_records,
                "工作區間": time_period_str,
                "總分鐘": total_minutes,
                "效率": efficiency,
                "空窗分鐘": idle_minutes,
                "儲位區域": storage_area_str,
                "空窗時間段": idle_segments_str,
            }
        )

    statistics_df = pd.DataFrame(stats)
    statistics_df["區域"] = pd.Categorical(statistics_df["區域"], categories=["低空", "高空"], ordered=True)
    statistics_df = statistics_df.sort_values(by=["區域", "揀貨人"])

    columns_order = ["區域", "揀貨人", "姓名", "筆數", "工作區間", "總分鐘", "效率", "空窗分鐘", "儲位區域", "空窗時間段"]
    return statistics_df[columns_order]


def calculate_statistics_afternoon(afternoon_df: pd.DataFrame, full_df: pd.DataFrame, mapping: Dict[str, Dict[str, str]]) -> pd.DataFrame:
    """
    下午（對稱邏輯）：
    - effective_start：max(第一筆, 13:30)
    - effective_end：若有 18:00 後紀錄 -> 18:00；否則 min(最後一筆, 18:00)
    - 休息：15:30-15:45（交集才扣）
    - 空窗：>=10 分，切掉休息重疊
    - 結尾空窗：計算到 effective_end
    """
    full_df = ensure_datetime(full_df)

    columns_order = ["區域", "揀貨人", "姓名", "筆數", "工作區間", "總分鐘", "效率", "空窗分鐘", "儲位區域", "空窗時間段"]
    if afternoon_df.empty:
        return pd.DataFrame(columns=columns_order)

    stats = []
    pickers = afternoon_df["揀貨人"].unique()

    for picker in pickers:
        picker_a = afternoon_df[afternoon_df["揀貨人"] == picker].sort_values("揀貨完成時間")
        times = list(picker_a["揀貨完成時間"])
        if not times:
            continue

        first_record = times[0]
        last_record = times[-1]

        start_dt = datetime.combine(first_record.date(), AFTERNOON_START)
        end_dt = datetime.combine(first_record.date(), AFTERNOON_END)
        effective_start = max(first_record, start_dt)

        picker_full = full_df[full_df["揀貨人"] == picker]
        has_after_end = any(rec.time() > AFTERNOON_END for rec in picker_full["揀貨完成時間"])
        effective_end = end_dt if has_after_end else min(last_record, end_dt)

        rest_start_dt = datetime.combine(first_record.date(), A_REST_START)
        rest_end_dt = datetime.combine(first_record.date(), A_REST_END)

        overlap_start = max(effective_start, rest_start_dt)
        overlap_end = min(effective_end, rest_end_dt)
        rest_duration = (overlap_end - overlap_start) if overlap_end > overlap_start else timedelta(0)

        work_duration = (effective_end - effective_start) - rest_duration
        total_minutes = round(work_duration.total_seconds() / 60, 2)

        num_records = len(picker_a)

        idle_segments: List[Tuple[datetime, datetime]] = []
        # 開頭空窗
        if times[0] > effective_start:
            idle_segments.extend(split_idle_segment(effective_start, times[0], rest_start_dt, rest_end_dt))
        # 中間空窗 >=10
        for i in range(1, len(times)):
            idle_segments.extend(get_effective_idle_segments(times[i - 1], times[i], rest_start_dt, rest_end_dt))
        # 結尾空窗：到 effective_end
        if last_record < effective_end:
            idle_segments.extend(get_effective_idle_segments(last_record, effective_end, rest_start_dt, rest_end_dt))

        idle_minutes = round(sum((e - s).total_seconds() for s, e in idle_segments) / 60, 2)
        efficiency = round((num_records / total_minutes * 60) if total_minutes else 0, 2)

        time_period_str = f"{effective_start.strftime('%H:%M:%S')} ~ {effective_end.strftime('%H:%M:%S')}"
        idle_segments_str = "; ".join(f"{s.strftime('%H:%M:%S')} ~ {e.strftime('%H:%M:%S')}" for s, e in idle_segments)

        working_records = picker_a[
            (picker_a["揀貨完成時間"] >= effective_start) & (picker_a["揀貨完成時間"] <= effective_end)
        ]
        storage_area_str = _storage_area_str(working_records)

        region = _get_region(picker, mapping)

        stats.append(
            {
                "區域": region,
                "揀貨人": picker,
                "姓名": _get_name(picker, mapping),
                "筆數": num_records,
                "工作區間": time_period_str,
                "總分鐘": total_minutes,
                "效率": efficiency,
                "空窗分鐘": idle_minutes,
                "儲位區域": storage_area_str,
                "空窗時間段": idle_segments_str,
            }
        )

    statistics_df = pd.DataFrame(stats)
    statistics_df["區域"] = pd.Categorical(statistics_df["區域"], categories=["低空", "高空"], ordered=True)
    statistics_df = statistics_df.sort_values(by=["區域", "揀貨人"])
    return statistics_df[columns_order]


# =========================================================
# 匯出 Excel（openpyxl）：同一張 Sheet 上下分段 + 達標紅綠底色
# =========================================================
def build_export_xlsx_bytes(
    title: str,
    morning_df: pd.DataFrame,
    afternoon_df: pd.DataFrame,
    low_threshold: float = 48.0,
    high_threshold: float = 20.0,
) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import FormulaRule

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    title_font = Font(name="新細明體", size=18, bold=True)
    stage_font = Font(name="新細明體", size=16, bold=True)
    header_font = Font(name="新細明體", size=12, bold=True)
    body_font = Font(name="新細明體", size=12)

    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    fill_green = PatternFill("solid", fgColor="C6EFCE")
    fill_red = PatternFill("solid", fgColor="FFC7CE")

    def write_table(start_row: int, df: pd.DataFrame) -> Tuple[int, int]:
        """回傳 (first_data_row, last_data_row)"""
        if df is None:
            df = pd.DataFrame()

        # header
        for c, col in enumerate(df.columns, start=1):
            cell = ws.cell(row=start_row, column=c, value=col)
            cell.font = header_font
            cell.alignment = align_center
            cell.border = border

        # body
        for r, row in enumerate(df.itertuples(index=False), start=start_row + 1):
            for c, value in enumerate(row, start=1):
                cell = ws.cell(row=r, column=c, value=value)
                cell.font = body_font
                cell.alignment = align_left if c in (10,) else align_center  # 空窗時間段較長
                cell.border = border

        first_data = start_row + 1
        last_data = start_row + len(df)
        return first_data, last_data

    def set_col_width(df: pd.DataFrame):
        if df is None or df.empty:
            return
        widths = {}
        for idx, col in enumerate(df.columns, start=1):
            # 簡單抓長度
            max_len = max([len(str(col))] + [len(str(x)) for x in df[col].astype(str).head(80).tolist()])
            widths[idx] = min(max(10, max_len + 2), 60)

        for col_idx, w in widths.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = w

    def add_conditional_format(first_data_row: int, last_data_row: int, max_col: int):
        """
        整列著色（依區域 + 效率門檻）
        - 高空：效率>=20 綠 / <20 紅
        - 低空：效率>=48 綠 / <48 紅
        欄位位置依你合併版：A=區域, G=效率
        """
        if last_data_row < first_data_row:
            return

        # Excel row number is 1-based already in openpyxl
        # 套用範圍：A..最後欄，從 first_data_row 到 last_data_row
        start_cell = f"A{first_data_row}"
        end_cell = f"{get_column_letter(max_col)}{last_data_row}"
        rng = f"{start_cell}:{end_cell}"

        # 注意：公式要用「範圍左上角那列」來參照，openpyxl 會自動相對套用
        row_ref = first_data_row

        # 高空綠/紅
        ws.conditional_formatting.add(
            rng,
            FormulaRule(
                formula=[f'=AND($A{row_ref}="高空",$G{row_ref}>={high_threshold})'],
                fill=fill_green,
            ),
        )
        ws.conditional_formatting.add(
            rng,
            FormulaRule(
                formula=[f'=AND($A{row_ref}="高空",$G{row_ref}<{high_threshold})'],
                fill=fill_red,
            ),
        )
        # 低空綠/紅
        ws.conditional_formatting.add(
            rng,
            FormulaRule(
                formula=[f'=AND($A{row_ref}="低空",$G{row_ref}>={low_threshold})'],
                fill=fill_green,
            ),
        )
        ws.conditional_formatting.add(
            rng,
            FormulaRule(
                formula=[f'=AND($A{row_ref}="低空",$G{row_ref}<{low_threshold})'],
                fill=fill_red,
            ),
        )

    # 決定欄數（以 morning 的欄位為主；若 morning 空則用 afternoon）
    base_df = morning_df if (morning_df is not None and not morning_df.empty) else afternoon_df
    max_col = max(1, len(base_df.columns)) if base_df is not None else 1

    # Title row (row 1)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    c = ws.cell(row=1, column=1, value=title)
    c.font = title_font
    c.alignment = align_center
    c.border = border

    # Stage 1 row (row 2)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_col)
    c = ws.cell(row=2, column=1, value="第一階段（上午）")
    c.font = stage_font
    c.alignment = align_center
    c.border = border

    # Morning table starts at row 3 (header)
    header_row_1 = 3
    first_data_1, last_data_1 = write_table(header_row_1, morning_df)
    if morning_df is not None and not morning_df.empty:
        add_conditional_format(first_data_1, last_data_1, max_col)

    # Gap + Stage 2
    gap = 2
    stage_row_2 = last_data_1 + gap + 1 if (morning_df is not None and not morning_df.empty) else (header_row_1 + 2)
    ws.merge_cells(start_row=stage_row_2, start_column=1, end_row=stage_row_2, end_column=max_col)
    c = ws.cell(row=stage_row_2, column=1, value="第二階段（下午）")
    c.font = stage_font
    c.alignment = align_center
    c.border = border

    header_row_2 = stage_row_2 + 1
    first_data_2, last_data_2 = write_table(header_row_2, afternoon_df)
    if afternoon_df is not None and not afternoon_df.empty:
        add_conditional_format(first_data_2, last_data_2, max_col)

    # Row heights (讓標題更像你原版)
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 24
    ws.row_dimensions[stage_row_2].height = 24

    # 欄寬
    set_col_width(base_df)

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


# =========================================================
# Streamlit UI
# =========================================================
def _load_uploaded_files(files: List) -> pd.DataFrame:
    dfs = []
    for f in files:
        name = (f.name or "").lower()
        if name.endswith(".csv"):
            df = pd.read_csv(f)
        else:
            df = pd.read_excel(f)
        dfs.append(df)
    if not dfs:
        return pd.DataFrame()
    return pd.concat(dfs, ignore_index=True)


def _init_mapping_state():
    if "pick_map" not in st.session_state:
        # 以 preset 為基底，但允許 UI 編輯（姓名可輸入中文）
        st.session_state.pick_map = {k: dict(v) for k, v in preset_picker_info.items()}


def _mapping_editor():
    """
    ✅ 姓名可輸入中文（text）
    ✅ 不改計算邏輯：只是更新 mapping 供 _get_name/_get_start_time_str/_get_region 使用
    """
    _init_mapping_state()
    with st.sidebar:
        st.subheader("🧑‍🤝‍🧑 揀貨人資料（可編輯）")
        st.caption("姓名可直接輸入中文；未填區域視為低空。")

        # 轉成 DataFrame 供編輯
        rows = []
        for emp_id, info in st.session_state.pick_map.items():
            rows.append(
                {
                    "員編": emp_id,
                    "姓名": info.get("姓名", ""),
                    "起始時間": info.get("起始時間", default_start_time_str),
                    "區域": info.get("區域", "低空") or "低空",
                }
            )
        edit_df = pd.DataFrame(rows)

        edited = st.data_editor(
            edit_df,
            use_container_width=True,
            num_rows="dynamic",
            hide_index=True,
            column_config={
                "員編": st.column_config.TextColumn("員編"),
                "姓名": st.column_config.TextColumn("姓名（中文可輸入）"),
                "起始時間": st.column_config.TextColumn("起始時間（HH:MM:SS）"),
                "區域": st.column_config.TextColumn("區域（低空/高空）"),
            },
            key="pick_map_editor",
        )

        # 回寫 session mapping
        new_map: Dict[str, Dict[str, str]] = {}
        for _, r in edited.iterrows():
            emp = str(r.get("員編", "")).strip()
            if not emp:
                continue
            new_map[emp] = {
                "姓名": str(r.get("姓名", "")).strip(),
                "起始時間": str(r.get("起始時間", default_start_time_str)).strip() or default_start_time_str,
                "區域": (str(r.get("區域", "低空")).strip() or "低空"),
            }
        st.session_state.pick_map = new_map


def main():
    inject_logistics_theme()
    set_page("總揀達標（上午＋下午）", icon="🧾", subtitle="同一張報表 Sheet 上下分段｜達標紅綠底色｜姓名可中文輸入")

    # sidebar controls
    _mapping_editor()
    with st.sidebar:
        st.divider()
        st.subheader("⚙️ 報表設定")
        report_title = st.text_input("報表標題", value="總揀達標獎金計算報表（合併版）")
        st.caption("達標門檻（沿用你原本條件）：高空 20、低空 48")
        high_threshold = st.number_input("高空達標（效率）", min_value=0.0, max_value=9999.0, value=20.0, step=1.0)
        low_threshold = st.number_input("低空達標（效率）", min_value=0.0, max_value=9999.0, value=48.0, step=1.0)

    # upload
    card_open("📤 上傳原始資料（可多檔合併）")
    files = st.file_uploader(
        "上傳 Excel / CSV",
        type=["xlsx", "xls", "csv"],
        accept_multiple_files=True,
        label_visibility="collapsed",
    )
    run = st.button("🚀 產出報表", type="primary", disabled=not files)
    card_close()

    if "picking_result" not in st.session_state:
        st.session_state.picking_result = None

    if run:
        # 計算一次後存 session，避免你按匯出導致 KPI 畫面消失
        with st.spinner("計算中，請稍候..."):
            raw_df = _load_uploaded_files(files)
            if raw_df.empty:
                st.error("未讀到任何資料，請確認檔案內容。")
                return

            # 合併版邏輯：去成箱 + 合併列 + 解析時間
            df = remove_boxed_rows(raw_df)
            full_df = combine_rows(df)
            full_df = ensure_datetime(full_df).dropna(subset=["揀貨完成時間"])

            # 篩上午/下午 + 計算
            morning_df = filter_morning_period(full_df)
            afternoon_df = filter_afternoon_period(full_df)

            mapping = st.session_state.pick_map

            morning_stats = calculate_statistics_morning(morning_df, full_df, mapping)
            afternoon_stats = calculate_statistics_afternoon(afternoon_df, full_df, mapping)

            # 產出 xlsx bytes（同一張 Sheet 上下分段）
            xlsx_bytes = build_export_xlsx_bytes(
                title=report_title.strip() or "總揀達標獎金計算報表（合併版）",
                morning_df=morning_stats,
                afternoon_df=afternoon_stats,
                low_threshold=float(low_threshold),
                high_threshold=float(high_threshold),
            )

            st.session_state.picking_result = {
                "report_title": report_title.strip() or "總揀達標獎金計算報表（合併版）",
                "morning_stats": morning_stats,
                "afternoon_stats": afternoon_stats,
                "xlsx_bytes": xlsx_bytes,
            }

    # render result (persist)
    result = st.session_state.picking_result
    if not result:
        st.info("請先上傳檔案並點「產出報表」。")
        return

    morning_stats = result["morning_stats"]
    afternoon_stats = result["afternoon_stats"]

    # KPI畫面：上午/下午上下顯示（你也可以改成左右 columns）
    card_open("📊 第一階段（上午）")
    if morning_stats is None or morning_stats.empty:
        st.info("上午無資料")
    else:
        st.dataframe(morning_stats, use_container_width=True, hide_index=True)
    card_close()

    card_open("📊 第二階段（下午）")
    if afternoon_stats is None or afternoon_stats.empty:
        st.info("下午無資料")
    else:
        st.dataframe(afternoon_stats, use_container_width=True, hide_index=True)
    card_close()

    # ✅ 匯出按鈕：直接是按鈕（不再用一個卡片標題分開）
    st.download_button(
        label="⬇️ 匯出報表（Excel）",
        data=result["xlsx_bytes"],
        file_name=f"{result['report_title']}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=False,
    )


if __name__ == "__main__":
    main()
