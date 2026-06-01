# pages/3_總揀達標.py
from __future__ import annotations

import io
from datetime import datetime, timedelta, time
from typing import Dict, List, Tuple, Optional

import pandas as pd
import streamlit as st

from common_ui import inject_logistics_theme, set_page, card_open, card_close


# =========================================================
# 參數（保留你原本邏輯）
# =========================================================
MORNING_END = datetime.strptime("12:30:00", "%H:%M:%S").time()
M_REST_START = datetime.strptime("10:00:00", "%H:%M:%S").time()
M_REST_END = datetime.strptime("10:15:00", "%H:%M:%S").time()

AFTERNOON_START = datetime.strptime("13:30:00", "%H:%M:%S").time()
AFTERNOON_END = datetime.strptime("18:00:00", "%H:%M:%S").time()
A_REST_START = datetime.strptime("15:30:00", "%H:%M:%S").time()
A_REST_END = datetime.strptime("15:45:00", "%H:%M:%S").time()

IDLE_THRESHOLD = timedelta(minutes=10)
default_start_time_str = "08:05:00"


# =========================================================
# 預設揀貨人資料（原樣保留；你可放完整名單）
# =========================================================
preset_picker_info: Dict[str, Dict[str, str]] = {
    "20230412002": {"姓名": "吳秉丞", "起始時間": "8:05:00", "區域": "低空"},
    "20200812002": {"姓名": "彭慈暉", "起始時間": "7:05:00", "區域": "低空"},
    "20210104001": {"姓名": "楊承珉", "起始時間": "7:05:00", "區域": "低空"},
    "20201109001": {"姓名": "梁冠如", "起始時間": "8:05:00", "區域": "低空"},
    "20201109003": {"姓名": "吳振凱", "起始時間": "8:05:00", "區域": "低空"},
        "20231226003": {"姓名": "顏秀菁", "起始時間": "8:05:00", "區域": "低空"},
    "20200922002": {"姓名": "葉欲弘", "起始時間": "8:05:00", "區域": "低空"},
    "20200924001": {"姓名": "黃雅君", "起始時間": "8:05:00", "區域": "低空"},
    "20201019001": {"姓名": "邱清瑞", "起始時間": "8:05:00", "區域": "低空"},
    "20220526001": {"姓名": "黃芷憶", "起始時間": "8:05:00", "區域": "低空"},
    "20240221003": {"姓名": "呂治明", "起始時間": "8:05:00", "區域": "低空"},
    "20240909001": {"姓名": "蔡麗珠", "起始時間": "8:05:00", "區域": "低空"},
    "20240926001": {"姓名": "陳莉娜", "起始時間": "8:05:00", "區域": "低空"},
    "20241011002": {"姓名": "林雙慧", "起始時間": "8:05:00", "區域": "低空"},
    "20250326001": {"姓名": "王大中", "起始時間": "8:05:00", "區域": "低空"},
    "20250303002": {"姓名": "周映華", "起始時間": "8:05:00", "區域": "低空"},
    "20250311001": {"姓名": "徐欣", "起始時間": "8:05:00", "區域": "低空"},
    "20250226002": {"姓名": "阮黃英", "起始時間": "7:05:00", "區域": "低空"},
    "20250901009": {"姓名": "張寶萱", "起始時間": "8:35:00", "區域": "低空"},
    "20250226010": {"姓名": "楊心如", "起始時間": "7:05:00", "區域": "低空"},
    "20250226011": {"姓名": "阮武玉玄", "起始時間": "7:05:00", "區域": "低空"},
    "20250226016": {"姓名": "阮氏美麗", "起始時間": "7:05:00", "區域": "低空"},
    "20250226018": {"姓名": "阮瑞美黃緣", "起始時間": "7:05:00", "區域": "低空"},
    "20250226020": {"姓名": "潘氏慶平", "起始時間": "7:05:00", "區域": "低空"},
    "20250226021": {"姓名": "潘氏青江", "起始時間": "7:05:00", "區域": "低空"},
    "20250923019": {"姓名": "阮氏紅深", "起始時間": "8:05:00", "區域": "低空"},
    "20250226026": {"姓名": "黎氏瓊", "起始時間": "7:05:00", "區域": "低空"},
    "20191205002": {"姓名": "阮功水", "起始時間": "8:05:00", "區域": "低空"},
    "20230119001": {"姓名": "陶春青", "起始時間": "7:05:00", "區域": "高空"},
    "20210318001": {"姓名": "陳文勇", "起始時間": "8:05:00", "區域": "低空"},
    "20210805001": {"姓名": "郭中合", "起始時間": "8:05:00", "區域": "低空"},
    "20220421002": {"姓名": "楊文點", "起始時間": "8:05:00", "區域": "低空"},
    "20220505001": {"姓名": "阮伊黃", "起始時間": "8:05:00", "區域": "低空"},
    "20220505002": {"姓名": "阮文青明", "起始時間": "7:05:00", "區域": "高空"},
    "001": {"姓名": "阮孟雄", "起始時間": "7:05:00", "區域": "高空"},
    "20221222005": {"姓名": "謝忠龍", "起始時間": "8:05:00", "區域": "高空"},
    "20221222009": {"姓名": "潘文一", "起始時間": "8:05:00", "區域": "低空"},
    "20221221001": {"姓名": "阮文全", "起始時間": "7:05:00", "區域": "高空"},
    "20230504001": {"姓名": "黃文重", "起始時間": "8:05:00", "區域": "低空"},
    "20230511003": {"姓名": "范日明", "起始時間": "7:05:00", "區域": "低空"},
    "20230810003": {"姓名": "范明俊", "起始時間": "8:05:00", "區域": "低空"},
    "20231211004": {"姓名": "河文南", "起始時間": "8:05:00", "區域": "低空"},
    "20231218004": {"姓名": "河文強", "起始時間": "8:05:00", "區域": "低空"},
    "20240107001": {"姓名": "范文春", "起始時間": "8:05:00", "區域": "低空"},
    "20240313001": {"姓名": "陳文越", "起始時間": "8:05:00", "區域": "低空"},
    "20240313003": {"姓名": "阮曰忠", "起始時間": "7:05:00", "區域": "高空"},
    "20240730001": {"姓名": "阮文忠", "起始時間": "8:05:00", "區域": "低空"},
    "20241204005": {"姓名": "阮春水", "起始時間": "7:05:00", "區域": "低空"},
    "20241204007": {"姓名": "阮玉名", "起始時間": "8:05:00", "區域": "低空"},
    "20241204009": {"姓名": "阮長文", "起始時間": "7:05:00", "區域": "低空"},
    "20220421001": {"姓名": "阮德平", "起始時間": "8:05:00", "區域": "高空"},
    "20250502001": {"姓名": "吳詩敏", "起始時間": "8:05:00", "區域": "低空"},
    "20250617003": {"姓名": "喬家寶", "起始時間": "8:05:00", "區域": "低空"},
    "20250901011": {"姓名": "章愛玲", "起始時間": "8:35:00", "區域": "低空"},
    "20250617001": {"姓名": "阮文譚", "起始時間": "7:05:00", "區域": "高空"},
    "09963": {"姓名": "黃謙凱", "起始時間": "8:05:00", "區域": "低空"},
    "11399": {"姓名": "陳哲沅", "起始時間": "8:05:00", "區域": "低空"},
}


# =========================================================
# 取 mapping：姓名/區域/起始時間（保留原邏輯）
# =========================================================
def _get_name(picker_id: str, mapping: Dict[str, Dict[str, str]]) -> str:
    if picker_id in mapping and (mapping[picker_id].get("姓名") or "").strip():
        return str(mapping[picker_id].get("姓名")).strip()
    if picker_id in preset_picker_info:
        return str(preset_picker_info[picker_id].get("姓名", "")).strip()
    return ""


def _get_start_time(picker_id: str, mapping: Dict[str, Dict[str, str]]) -> str:
    if picker_id in mapping and (mapping[picker_id].get("起始時間") or "").strip():
        return str(mapping[picker_id].get("起始時間")).strip()
    if picker_id in preset_picker_info:
        return str(preset_picker_info[picker_id].get("起始時間", default_start_time_str)).strip()
    return default_start_time_str


def _get_region(picker_id: str, mapping: Dict[str, Dict[str, str]]) -> str:
    if picker_id in mapping and (mapping[picker_id].get("區域") or "").strip():
        return str(mapping[picker_id].get("區域")).strip()
    if picker_id in preset_picker_info:
        return str(preset_picker_info[picker_id].get("區域", "低空")).strip() or "低空"
    return "低空"


def _storage_area_str(records: pd.DataFrame) -> str:
    if records is None or records.empty or "儲位" not in records.columns:
        return ""
    vals = records["儲位"].dropna().astype(str).str.strip()
    vals = vals[vals != ""]
    if vals.empty:
        return ""
    head = vals.str[:2].value_counts()
    top = head.head(8).index.tolist()
    return ",".join(top)


# =========================================================
# 時間解析（保留既有容錯）
# =========================================================
def parse_tw_datetime(series: pd.Series) -> pd.Series:
    return pd.to_datetime(series, errors="coerce")


def ensure_datetime(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    if "揀貨完成時間" in df.columns:
        df["揀貨完成時間"] = parse_tw_datetime(df["揀貨完成時間"])
    return df


# =========================================================
# ✅ 手動空窗：時間字串解析（HH:MM 或 HH:MM:SS）
# =========================================================
def parse_time_str(s: str) -> Optional[time]:
    s = (s or "").strip()
    if not s:
        return None
    fmts = ["%H:%M:%S", "%H:%M"]
    for f in fmts:
        try:
            return datetime.strptime(s, f).time()
        except Exception:
            continue
    return None


# =========================================================
# ✅ 區間工具：切掉多個排除區間（休息 + 手動空窗）
# =========================================================
def subtract_intervals(seg_start: datetime, seg_end: datetime, excludes: List[Tuple[datetime, datetime]]) -> List[Tuple[datetime, datetime]]:
    """
    回傳 seg_start~seg_end 扣掉 excludes 後剩餘的子區間。
    excludes 必須是同一天 datetime 且 start<end。
    """
    if seg_end <= seg_start:
        return []
    if not excludes:
        return [(seg_start, seg_end)]

    # normalize + sort
    ex = [(s, e) for s, e in excludes if e > s]
    ex.sort(key=lambda x: x[0])

    # 合併重疊
    merged: List[Tuple[datetime, datetime]] = []
    for s, e in ex:
        if not merged or s > merged[-1][1]:
            merged.append((s, e))
        else:
            merged[-1] = (merged[-1][0], max(merged[-1][1], e))

    cur_s = seg_start
    out: List[Tuple[datetime, datetime]] = []
    for s, e in merged:
        if e <= cur_s:
            continue
        if s >= seg_end:
            break
        if s > cur_s:
            out.append((cur_s, min(s, seg_end)))
        cur_s = max(cur_s, e)
        if cur_s >= seg_end:
            break

    if cur_s < seg_end:
        out.append((cur_s, seg_end))

    return [(s, e) for s, e in out if e > s]


def overlap_minutes(a_start: datetime, a_end: datetime, b_start: datetime, b_end: datetime) -> float:
    s = max(a_start, b_start)
    e = min(a_end, b_end)
    if e <= s:
        return 0.0
    return (e - s).total_seconds() / 60.0


# =========================================================
# 讀檔/前處理（保留原邏輯）
# =========================================================
def _load_uploaded_files(files: List[st.runtime.uploaded_file_manager.UploadedFile]) -> pd.DataFrame:
    frames: List[pd.DataFrame] = []
    for f in files:
        name = (f.name or "").lower()
        b = f.getvalue()
        try:
            if name.endswith(".csv"):
                frames.append(pd.read_csv(io.BytesIO(b)))
            else:
                frames.append(pd.read_excel(io.BytesIO(b)))
        except Exception:
            continue
    if not frames:
        return pd.DataFrame()
    return pd.concat(frames, ignore_index=True)


def remove_boxed_rows(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return df
    if "成箱箱號" in df.columns:
        tmp = df.copy()
        tmp["成箱箱號"] = tmp["成箱箱號"].astype(str).fillna("").str.strip()
        tmp = tmp[tmp["成箱箱號"] == ""]
        return tmp
    return df


def combine_rows(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return df
    group_cols = ["儲位", "商品", "揀貨人", "揀貨完成時間"]
    for c in group_cols:
        if c not in df.columns:
            return df
    if "數量" not in df.columns:
        df = df.copy()
        df["數量"] = 1
    combined_df = df.groupby(group_cols, as_index=False).agg({"數量": "sum"})
    return combined_df


def filter_morning_period(df: pd.DataFrame) -> pd.DataFrame:
    dtv = parse_tw_datetime(df["揀貨完成時間"])
    df = df.assign(揀貨完成時間=dtv).dropna(subset=["揀貨完成時間"])
    df = df[df["揀貨完成時間"].dt.time <= MORNING_END]
    return df


def filter_afternoon_period(df: pd.DataFrame) -> pd.DataFrame:
    dtv = parse_tw_datetime(df["揀貨完成時間"])
    df = df.assign(揀貨完成時間=dtv).dropna(subset=["揀貨完成時間"])
    df = df[(df["揀貨完成時間"].dt.time >= AFTERNOON_START) & (df["揀貨完成時間"].dt.time <= AFTERNOON_END)]
    return df


# =========================================================
# ✅ 手動空窗：將 session 設定轉成該日 datetime 排除區間
# =========================================================
def build_manual_excludes_for_day(
    day: datetime.date,
    picker: str,
    manual_excludes: List[Dict[str, str]],
) -> List[Tuple[datetime, datetime]]:
    out: List[Tuple[datetime, datetime]] = []
    for r in manual_excludes or []:
        who = (r.get("picker") or "").strip()
        # who 空白 → 全體；有填 → 只套用該 picker
        if who and who != picker:
            continue

        ts = parse_time_str(r.get("start", ""))
        te = parse_time_str(r.get("end", ""))
        if ts is None or te is None:
            continue

        sdt = datetime.combine(day, ts)
        edt = datetime.combine(day, te)
        if edt <= sdt:
            # 不做跨日，避免破壞原邏輯（需要跨日再另外擴充）
            continue
        out.append((sdt, edt))
    return out


# =========================================================
# 空窗判斷（保留原邏輯，但「切掉排除區間」）
# =========================================================
def get_effective_idle_segments(
    prev_t: datetime,
    cur_t: datetime,
    excludes: List[Tuple[datetime, datetime]],
) -> List[Tuple[datetime, datetime]]:
    if cur_t <= prev_t:
        return []
    if (cur_t - prev_t) < IDLE_THRESHOLD:
        return []
    segs = subtract_intervals(prev_t, cur_t, excludes)
    return [(s, e) for s, e in segs if (e - s) >= IDLE_THRESHOLD]


# =========================================================
# 計算：上午（保留邏輯 + 加入手動空窗扣除）
# =========================================================
def calculate_statistics_morning(
    morning_df: pd.DataFrame,
    full_df: pd.DataFrame,
    mapping: Dict[str, Dict[str, str]],
    manual_excludes: List[Dict[str, str]],
) -> pd.DataFrame:
    columns_order = ["區域", "揀貨人", "姓名", "筆數", "工作區間", "總分鐘", "效率", "空窗分鐘", "儲位區域", "空窗時間段"]
    if morning_df is None or morning_df.empty:
        return pd.DataFrame(columns=columns_order)

    stats: List[Dict[str, object]] = []
    morning_df = ensure_datetime(morning_df).dropna(subset=["揀貨完成時間"])
    full_df = ensure_datetime(full_df).dropna(subset=["揀貨完成時間"])

    for picker in sorted(morning_df["揀貨人"].dropna().astype(str).unique()):
        picker_m = morning_df[morning_df["揀貨人"].astype(str) == picker].sort_values("揀貨完成時間")
        if picker_m.empty:
            continue

        first_record = picker_m["揀貨完成時間"].iloc[0].to_pydatetime()
        last_record = picker_m["揀貨完成時間"].iloc[-1].to_pydatetime()

        # 起始時間（你原本邏輯：可被設定覆蓋）
        start_time_str = _get_start_time(picker, mapping) or default_start_time_str
        st_time = parse_time_str(start_time_str) or parse_time_str(default_start_time_str)  # type: ignore
        start_dt = datetime.combine(first_record.date(), st_time)  # type: ignore
        end_dt = datetime.combine(first_record.date(), MORNING_END)

        effective_start = min(first_record, start_dt)

        # 若有下午紀錄 → 上午結束用 12:30；否則用 min(最後一筆, 12:30)
        picker_full = full_df[full_df["揀貨人"].astype(str) == picker]
        has_afternoon = any(rec.time() >= AFTERNOON_START for rec in picker_full["揀貨完成時間"])
        effective_end = end_dt if has_afternoon else min(last_record, end_dt)

        # 排除區間 = 休息 + 手動空窗（同一天）
        rest_start_dt = datetime.combine(first_record.date(), M_REST_START)
        rest_end_dt = datetime.combine(first_record.date(), M_REST_END)

        manual_dt_ex = build_manual_excludes_for_day(first_record.date(), picker, manual_excludes)
        excludes = [(rest_start_dt, rest_end_dt)] + manual_dt_ex

        # ✅ 總分鐘：原本扣休息，再額外扣手動空窗（只扣與工作區間交集）
        total_range_minutes = (effective_end - effective_start).total_seconds() / 60.0

        rest_minutes = overlap_minutes(effective_start, effective_end, rest_start_dt, rest_end_dt)
        manual_minutes = sum(overlap_minutes(effective_start, effective_end, s, e) for s, e in manual_dt_ex)

        total_minutes = round(max(0.0, total_range_minutes - rest_minutes - manual_minutes), 2)

        # 空窗計算（原本：>=10 分鐘；現在會切掉 excludes）
        times = picker_m["揀貨完成時間"].dt.to_pydatetime().tolist()
        idle_segments: List[Tuple[datetime, datetime]] = []

        # 開頭空窗
        if times and times[0] > effective_start:
            segs = subtract_intervals(effective_start, times[0], excludes)
            idle_segments.extend([(s, e) for s, e in segs if (e - s) >= IDLE_THRESHOLD])

        # 中間空窗
        for i in range(1, len(times)):
            idle_segments.extend(get_effective_idle_segments(times[i - 1], times[i], excludes))

        # 結尾空窗（依你原本邏輯：到 effective_end）
        if last_record < effective_end:
            idle_segments.extend(get_effective_idle_segments(last_record, effective_end, excludes))

        idle_minutes = round(sum((e - s).total_seconds() for s, e in idle_segments) / 60.0, 2)

        num_records = len(picker_m)
        efficiency = round((num_records / total_minutes * 60) if total_minutes else 0, 2)

        time_period_str = f"{effective_start.strftime('%H:%M:%S')} ~ {effective_end.strftime('%H:%M:%S')}"
        idle_segments_str = "; ".join(f"{s.strftime('%H:%M:%S')} ~ {e.strftime('%H:%M:%S')}" for s, e in idle_segments)

        working_records = picker_m[(picker_m["揀貨完成時間"] >= effective_start) & (picker_m["揀貨完成時間"] <= effective_end)]
        storage_area_str = _storage_area_str(working_records)
        region = _get_region(picker, mapping)

        stats.append(
            {
                "區域": region,
                "揀貨人": picker,
                "姓名": _get_name(picker, mapping),
                "筆數": num_records,
                "工作區間": time_period_str,
                "總分鐘": total_minutes,
                "效率": efficiency,
                "空窗分鐘": idle_minutes,
                "儲位區域": storage_area_str,
                "空窗時間段": idle_segments_str,
            }
        )

    df = pd.DataFrame(stats)
    if df.empty:
        return pd.DataFrame(columns=columns_order)
    df["區域"] = pd.Categorical(df["區域"], categories=["低空", "高空"], ordered=True)
    df = df.sort_values(by=["區域", "揀貨人"])
    return df[columns_order]


# =========================================================
# 計算：下午（保留邏輯 + 加入手動空窗扣除）
# =========================================================
def calculate_statistics_afternoon(
    afternoon_df: pd.DataFrame,
    full_df: pd.DataFrame,
    mapping: Dict[str, Dict[str, str]],
    manual_excludes: List[Dict[str, str]],
) -> pd.DataFrame:
    columns_order = ["區域", "揀貨人", "姓名", "筆數", "工作區間", "總分鐘", "效率", "空窗分鐘", "儲位區域", "空窗時間段"]
    if afternoon_df is None or afternoon_df.empty:
        return pd.DataFrame(columns=columns_order)

    stats: List[Dict[str, object]] = []
    afternoon_df = ensure_datetime(afternoon_df).dropna(subset=["揀貨完成時間"])
    full_df = ensure_datetime(full_df).dropna(subset=["揀貨完成時間"])

    for picker in sorted(afternoon_df["揀貨人"].dropna().astype(str).unique()):
        picker_a = afternoon_df[afternoon_df["揀貨人"].astype(str) == picker].sort_values("揀貨完成時間")
        if picker_a.empty:
            continue

        first_record = picker_a["揀貨完成時間"].iloc[0].to_pydatetime()
        last_record = picker_a["揀貨完成時間"].iloc[-1].to_pydatetime()

        start_dt = datetime.combine(first_record.date(), AFTERNOON_START)
        end_dt = datetime.combine(first_record.date(), AFTERNOON_END)

        # 你原本邏輯是 min(first_record, start_dt)；我保持不動
        effective_start = min(first_record, start_dt)

        picker_full = full_df[full_df["揀貨人"].astype(str) == picker]
        has_after_end = any(rec.time() > AFTERNOON_END for rec in picker_full["揀貨完成時間"])
        effective_end = end_dt if has_after_end else min(last_record, end_dt)

        rest_start_dt = datetime.combine(first_record.date(), A_REST_START)
        rest_end_dt = datetime.combine(first_record.date(), A_REST_END)

        manual_dt_ex = build_manual_excludes_for_day(first_record.date(), picker, manual_excludes)
        excludes = [(rest_start_dt, rest_end_dt)] + manual_dt_ex

        total_range_minutes = (effective_end - effective_start).total_seconds() / 60.0
        rest_minutes = overlap_minutes(effective_start, effective_end, rest_start_dt, rest_end_dt)
        manual_minutes = sum(overlap_minutes(effective_start, effective_end, s, e) for s, e in manual_dt_ex)

        total_minutes = round(max(0.0, total_range_minutes - rest_minutes - manual_minutes), 2)

        times = picker_a["揀貨完成時間"].dt.to_pydatetime().tolist()
        idle_segments: List[Tuple[datetime, datetime]] = []

        if times and times[0] > effective_start:
            segs = subtract_intervals(effective_start, times[0], excludes)
            idle_segments.extend([(s, e) for s, e in segs if (e - s) >= IDLE_THRESHOLD])

        for i in range(1, len(times)):
            idle_segments.extend(get_effective_idle_segments(times[i - 1], times[i], excludes))

        if last_record < effective_end:
            idle_segments.extend(get_effective_idle_segments(last_record, effective_end, excludes))

        idle_minutes = round(sum((e - s).total_seconds() for s, e in idle_segments) / 60.0, 2)

        num_records = len(picker_a)
        efficiency = round((num_records / total_minutes * 60) if total_minutes else 0, 2)

        time_period_str = f"{effective_start.strftime('%H:%M:%S')} ~ {effective_end.strftime('%H:%M:%S')}"
        idle_segments_str = "; ".join(f"{s.strftime('%H:%M:%S')} ~ {e.strftime('%H:%M:%S')}" for s, e in idle_segments)

        working_records = picker_a[(picker_a["揀貨完成時間"] >= effective_start) & (picker_a["揀貨完成時間"] <= effective_end)]
        storage_area_str = _storage_area_str(working_records)
        region = _get_region(picker, mapping)

        stats.append(
            {
                "區域": region,
                "揀貨人": picker,
                "姓名": _get_name(picker, mapping),
                "筆數": num_records,
                "工作區間": time_period_str,
                "總分鐘": total_minutes,
                "效率": efficiency,
                "空窗分鐘": idle_minutes,
                "儲位區域": storage_area_str,
                "空窗時間段": idle_segments_str,
            }
        )

    df = pd.DataFrame(stats)
    if df.empty:
        return pd.DataFrame(columns=columns_order)
    df["區域"] = pd.Categorical(df["區域"], categories=["低空", "高空"], ordered=True)
    df = df.sort_values(by=["區域", "揀貨人"])
    return df[columns_order]


# =========================================================
# 匯出 Excel：同一張 Sheet 上下分段 + 紅綠底（保留你原匯出邏輯）
# =========================================================
def build_export_xlsx_bytes(
    title: str,
    morning_df: pd.DataFrame,
    afternoon_df: pd.DataFrame,
    low_threshold: float = 57.0,
    high_threshold: float = 21.0,
) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    thin = Side(style="thin", color="333333")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    fill_green = PatternFill("solid", fgColor="C6EFCE")
    fill_red = PatternFill("solid", fgColor="FFC7CE")
    font_green = Font(color="006100")
    font_red = Font(color="9C0006")

    def autosize_cols(start_row: int, end_row: int, start_col: int, end_col: int):
        widths = {}
        for r in range(start_row, end_row + 1):
            for c in range(start_col, end_col + 1):
                v = ws.cell(r, c).value
                if v is None:
                    continue
                widths[c] = max(widths.get(c, 0), len(str(v)))
        for c, w in widths.items():
            ws.column_dimensions[get_column_letter(c)].width = min(max(10, w + 2), 48)

    def write_block(block_title: str, df: pd.DataFrame, start_row: int) -> int:
        max_cols = max(1, len(df.columns) if df is not None and not df.empty else 1)
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=max_cols)
        c = ws.cell(start_row, 1, block_title)
        c.font = Font(size=14, bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[start_row].height = 22

        if df is None or df.empty:
            ws.cell(start_row + 1, 1, "（本段無資料）")
            return start_row + 3

        header_row = start_row + 1
        for j, col in enumerate(df.columns, start=1):
            cell = ws.cell(header_row, j, col)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

        for i, row in enumerate(df.itertuples(index=False), start=header_row + 1):
            for j, v in enumerate(row, start=1):
                cell = ws.cell(i, j, v)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = border

            try:
                region = str(df.iloc[i - (header_row + 1)]["區域"])
                eff = float(df.iloc[i - (header_row + 1)]["效率"])
            except Exception:
                region, eff = "", 0.0

            ok = False
            if region == "高空":
                ok = eff >= float(high_threshold)
            elif region == "低空":
                ok = eff >= float(low_threshold)

            for j in range(1, len(df.columns) + 1):
                if ok:
                    ws.cell(i, j).fill = fill_green
                    ws.cell(i, j).font = font_green
                else:
                    ws.cell(i, j).fill = fill_red
                    ws.cell(i, j).font = font_red

        end_row = header_row + len(df)
        autosize_cols(header_row, end_row, 1, len(df.columns))
        return end_row + 2

    max_cols = max(
        1,
        (len(morning_df.columns) if morning_df is not None and not morning_df.empty else 1),
        (len(afternoon_df.columns) if afternoon_df is not None and not afternoon_df.empty else 1),
    )
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_cols)
    t = ws.cell(1, 1, title)
    t.font = Font(size=18, bold=True)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    next_row = 3
    next_row = write_block("第一階段（上午）", morning_df, next_row)
    next_row = write_block("第二階段（下午）", afternoon_df, next_row)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# =========================================================
# sidebar：揀貨人設定（姓名可中文輸入；保留原邏輯）
# =========================================================
def _mapping_editor():
    if "pick_map" not in st.session_state:
        st.session_state.pick_map = {}

    with st.sidebar:
        st.subheader("👤 揀貨人設定（可中文姓名）")
        st.caption("輸入「揀貨人」代碼後，可設定姓名/起始時間/區域（高空/低空）。")

        picker_id = st.text_input("揀貨人代碼（可貼上）", value="")
        col1, col2 = st.columns(2)
        with col1:
            name = st.text_input("姓名（中文可輸入）", value="")
        with col2:
            region = st.selectbox("區域", options=["低空", "高空"], index=0)

        start_time = st.text_input("起始時間（HH:MM:SS）", value=default_start_time_str)

        if st.button("➕ 新增 / 更新"):
            pid = (picker_id or "").strip()
            if not pid:
                st.warning("請輸入揀貨人代碼")
            else:
                st.session_state.pick_map[pid] = {
                    "姓名": (name or "").strip(),
                    "起始時間": (start_time or "").strip(),
                    "區域": (region or "").strip(),
                }
                st.success("已更新")

        if st.session_state.pick_map:
            st.divider()
            st.caption("目前設定：")
            mdf = pd.DataFrame([{"揀貨人": k, **v} for k, v in st.session_state.pick_map.items()])
            st.dataframe(mdf, use_container_width=True, hide_index=True)


# =========================================================
# ✅ sidebar：手動空窗扣除（新增）
# =========================================================
def _manual_excludes_editor():
    if "manual_excludes" not in st.session_state:
        st.session_state.manual_excludes = []

    with st.sidebar:
        st.divider()
        st.subheader("⛔ 手動空窗扣除（會從總分鐘扣除）")
        st.caption("時間格式：HH:MM 或 HH:MM:SS。可指定某位揀貨人，留空則全體套用。")

        who = st.text_input("指定揀貨人代碼（可留空=全體）", value="", key="mx_who")
        c1, c2 = st.columns(2)
        with c1:
            s = st.text_input("開始時間", value="", placeholder="例如 11:10 或 11:10:00", key="mx_start")
        with c2:
            e = st.text_input("結束時間", value="", placeholder="例如 11:25 或 11:25:00", key="mx_end")

        add = st.button("➕ 新增空窗扣除")
        if add:
            ts = parse_time_str(s)
            te = parse_time_str(e)
            if ts is None or te is None:
                st.error("時間格式錯誤，請用 HH:MM 或 HH:MM:SS")
            else:
                st.session_state.manual_excludes.append(
                    {"picker": (who or "").strip(), "start": s.strip(), "end": e.strip()}
                )
                st.success("已新增空窗扣除")

        if st.session_state.manual_excludes:
            st.caption("目前空窗扣除清單：")
            st.dataframe(pd.DataFrame(st.session_state.manual_excludes), use_container_width=True, hide_index=True)

            colx1, colx2 = st.columns(2)
            with colx1:
                if st.button("🧹 清空全部空窗扣除"):
                    st.session_state.manual_excludes = []
                    st.success("已清空")
            with colx2:
                st.caption("（建議：每次產出前檢查一次）")


# =========================================================
# ✅ 畫面 KPI 整列紅/綠底（依區域+效率門檻）
# =========================================================
def _style_kpi_rows(df: pd.DataFrame, low_threshold: float, high_threshold: float) -> "pd.io.formats.style.Styler":
    if df is None or df.empty:
        return df.style

    def _row_style(row: pd.Series) -> List[str]:
        try:
            region = str(row.get("區域", ""))
            eff = float(row.get("效率", 0))
        except Exception:
            region, eff = "", 0.0

        ok = False
        if region == "高空":
            ok = eff >= float(high_threshold)
        elif region == "低空":
            ok = eff >= float(low_threshold)

        if ok:
            bg, fg = "#C6EFCE", "#006100"
        else:
            bg, fg = "#FFC7CE", "#9C0006"
        return [f"background-color: {bg}; color: {fg};" for _ in row.index]

    return df.style.apply(_row_style, axis=1)


# =========================================================
# Main
# =========================================================
def main():
    inject_logistics_theme()
    set_page(
        "總揀達標（合併版）",
        icon="🧾",
        subtitle="同一張報表 Sheet 上下分段｜達標紅綠底色｜姓名可中文輸入｜可手動扣除空窗",
    )

    # Sidebar
    _mapping_editor()
    _manual_excludes_editor()

    with st.sidebar:
        st.divider()
        st.subheader("⚙️ 報表設定")
        report_title = st.text_input("報表標題", value="總揀達標獎金計算報表（合併版）")
        st.caption("達標門檻（沿用你原本條件）：高空 21、低空 57")
        high_threshold = st.number_input("高空達標（效率）", min_value=0.0, max_value=9999.0, value=21.0, step=1.0)
        low_threshold = st.number_input("低空達標（效率）", min_value=0.0, max_value=9999.0, value=57.0, step=1.0)

    # Upload
    card_open("📤 上傳原始資料（可多檔合併）")
    files = st.file_uploader(
        "上傳 Excel / CSV",
        type=["xlsx", "xls", "csv"],
        accept_multiple_files=True,
        label_visibility="collapsed",
    )
    run = st.button("🚀 產出報表", type="primary", disabled=not files)
    card_close()

    if "picking_result" not in st.session_state:
        st.session_state.picking_result = None

    if run:
        with st.spinner("計算中，請稍候..."):
            raw_df = _load_uploaded_files(files)
            if raw_df.empty:
                st.error("未讀到任何資料，請確認檔案內容。")
                return

            df = remove_boxed_rows(raw_df)
            full_df = combine_rows(df)
            full_df = ensure_datetime(full_df).dropna(subset=["揀貨完成時間"])

            morning_df = filter_morning_period(full_df)
            afternoon_df = filter_afternoon_period(full_df)

            mapping = st.session_state.pick_map
            manual_excludes = st.session_state.manual_excludes

            morning_stats = calculate_statistics_morning(morning_df, full_df, mapping, manual_excludes)
            afternoon_stats = calculate_statistics_afternoon(afternoon_df, full_df, mapping, manual_excludes)

            xlsx_bytes = build_export_xlsx_bytes(
                title=report_title.strip() or "總揀達標獎金計算報表（合併版）",
                morning_df=morning_stats,
                afternoon_df=afternoon_stats,
                low_threshold=float(low_threshold),
                high_threshold=float(high_threshold),
            )

            st.session_state.picking_result = {
                "report_title": report_title.strip() or "總揀達標獎金計算報表（合併版）",
                "morning_stats": morning_stats,
                "afternoon_stats": afternoon_stats,
                "xlsx_bytes": xlsx_bytes,
                "low_threshold": float(low_threshold),
                "high_threshold": float(high_threshold),
            }

    result = st.session_state.picking_result
    if not result:
        st.info("請先上傳檔案並點「產出報表」。")
        return

    morning_stats = result["morning_stats"]
    afternoon_stats = result["afternoon_stats"]
    low_thr = float(result.get("low_threshold", 57.0))
    high_thr = float(result.get("high_threshold", 21.0))

    # KPI 表格（整列紅綠底）
    card_open("📊 第一階段（上午）")
    if morning_stats is None or morning_stats.empty:
        st.info("上午無資料")
    else:
        st.dataframe(_style_kpi_rows(morning_stats, low_thr, high_thr), use_container_width=True, hide_index=True)
    card_close()

    card_open("📊 第二階段（下午）")
    if afternoon_stats is None or afternoon_stats.empty:
        st.info("下午無資料")
    else:
        st.dataframe(_style_kpi_rows(afternoon_stats, low_thr, high_thr), use_container_width=True, hide_index=True)
    card_close()

    # 匯出（按了也不會清掉 KPI，因為資料存在 session_state）
    st.download_button(
        label="⬇️ 匯出報表（Excel）",
        data=result["xlsx_bytes"],
        file_name=f"{result['report_title']}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


if __name__ == "__main__":
    main()
