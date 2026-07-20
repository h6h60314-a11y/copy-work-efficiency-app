from __future__ import annotations

import io
import re
import zipfile
from typing import BinaryIO

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


PAGE_TITLE = "月出貨量與產力"
DELETE_KEYWORDS = ("FT03", "FT04", "FT05", "FT06", "FT07", "FT08", "FT09")
ENCODINGS = ("utf-8-sig", "cp950", "big5", "gb18030", "utf-16", "utf-16le", "utf-16be")
GREEN = "1B7F4B"
LIGHT_GREEN = "EAF5EF"


def _setup_page() -> None:
    """沿用平台共用版型；單獨執行時仍可正常顯示。"""
    try:
        from common_ui import set_page

        set_page(PAGE_TITLE)
    except Exception:
        try:
            st.set_page_config(page_title=PAGE_TITLE, page_icon="📦", layout="wide")
        except Exception:
            pass

    st.markdown(
        """
        <style>
        .df-title {font-size: 1.75rem; font-weight: 800; color: #175c3a; margin-bottom: .15rem;}
        .df-subtitle {color: #607268; margin-bottom: 1.25rem;}
        [data-testid="stMetric"] {background: #f7fbf8; border: 1px solid #dcebe2;
            border-radius: 12px; padding: 14px 16px;}
        div.stButton > button[kind="primary"], div.stDownloadButton > button[kind="primary"] {
            background: #1b7f4b; border-color: #1b7f4b;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )


def normalize_key(value) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.lower() in {"nan", "none"}:
        return ""
    return text[:-2].strip() if text.endswith(".0") else text


def normalize_text(value) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return text[:-2] if text.endswith(".0") else text


def to_number(value):
    if value is None:
        return None
    try:
        text = str(value).replace(",", "").strip()
        if not text or text.lower() in {"nan", "none"}:
            return None
        return float(text)
    except (TypeError, ValueError):
        return None


def calc_unit_qty(packqty, qty_per_box):
    """GCOR 排除資料：單位入數 = packqty / 揀貨入數。"""
    packqty_num = to_number(packqty)
    qty_num = to_number(qty_per_box)
    if packqty_num is None or qty_num in (None, 0):
        return ""
    result = packqty_num / qty_num
    return int(result) if result.is_integer() else result


def calc_normal_unit_qty(packqty, qty_per_box):
    """一般資料：若 packqty / 入數不為整數，改採原 packqty。"""
    packqty_num = to_number(packqty)
    qty_num = to_number(qty_per_box)
    if packqty_num is None or qty_num in (None, 0):
        return ""
    result = packqty_num / qty_num
    if not result.is_integer():
        result = packqty_num
    return int(result) if float(result).is_integer() else result


def should_delete_vehicle(value) -> bool:
    text = normalize_text(value).upper()
    return any(keyword in text for keyword in DELETE_KEYWORDS)


def is_gcor_packqty_100(buyersreference_value, packqty_value) -> bool:
    buyer_text = normalize_text(buyersreference_value).upper()
    packqty_num = to_number(packqty_value)
    return buyer_text == "GCOR" and packqty_num is not None and packqty_num >= 100


def clean_illegal_excel_text(value):
    if isinstance(value, str):
        return re.sub(r"[\000-\010]|[\013-\014]|[\016-\037]", "", value)
    return value


def load_master_mapping(master_file: BinaryIO) -> tuple[dict[str, str], str]:
    master_file.seek(0)
    all_sheets = pd.read_excel(master_file, sheet_name=None, dtype=str)
    for sheet_name, df in all_sheets.items():
        df.columns = df.columns.astype(str).str.strip()
        if {"商品代號", "揀貨入數"}.issubset(df.columns):
            codes = df["商品代號"].map(normalize_key)
            quantities = df["揀貨入數"].map(normalize_key)
            mapping: dict[str, str] = {}
            for product_code, pick_qty in zip(codes, quantities):
                if product_code and product_code not in mapping:
                    mapping[product_code] = pick_qty
            return mapping, str(sheet_name)
    raise ValueError("商品主檔中找不到欄位「商品代號」與「揀貨入數」")


def detect_encoding(raw: bytes) -> str:
    for encoding in ENCODINGS:
        try:
            raw[:10000].decode(encoding, errors="strict")
            return encoding
        except (UnicodeDecodeError, LookupError):
            continue
    return "cp950"


def detect_separator(text: str) -> str:
    lines = [line for line in text.splitlines()[:30] if line.strip()]
    sample = "\n".join(lines)
    candidates = {"\t": sample.count("\t"), ",": sample.count(","), "|": sample.count("|"), ";": sample.count(";")}
    best_separator = max(candidates, key=candidates.get)
    return best_separator if candidates[best_separator] > 0 else r"\s{2,}"


def read_txt_smart(uploaded_file: BinaryIO) -> pd.DataFrame:
    uploaded_file.seek(0)
    raw = uploaded_file.read()
    encoding = detect_encoding(raw)
    text = raw.decode(encoding, errors="replace")
    separator = detect_separator(text)
    try:
        df = pd.read_csv(io.StringIO(text), sep=separator, dtype=str, engine="python", on_bad_lines="skip")
    except Exception:
        df = pd.read_fwf(io.StringIO(text), dtype=str)
    df.columns = df.columns.astype(str).str.strip()
    for column in df.columns:
        df[column] = df[column].astype(str).str.strip().replace({"nan": "", "None": "", "NaN": "", "NAN": ""})
    return df


def add_pick_qty_by_sku(df: pd.DataFrame, master_mapping: dict[str, str]):
    df = df.copy()
    df.columns = df.columns.astype(str).str.strip()
    if "SKU" not in df.columns:
        df["揀貨入數"] = ""
        return df, 0, len(df)
    df["SKU"] = df["SKU"].map(normalize_key)
    pick_values = df["SKU"].map(master_mapping).fillna("")
    matched_count = int(pick_values.ne("").sum())
    if "揀貨入數" in df.columns:
        df["揀貨入數"] = pick_values
    elif "入數" in df.columns:
        df.insert(list(df.columns).index("入數") + 1, "揀貨入數", pick_values)
    else:
        df["揀貨入數"] = pick_values
    return df, matched_count, len(df) - matched_count


def process_dataframe(df: pd.DataFrame, master_mapping: dict[str, str]):
    df = df.copy()
    df.columns = df.columns.astype(str).str.strip()
    original_count = len(df)
    if "載具號" in df.columns:
        df = df.loc[~df["載具號"].map(should_delete_vehicle)].copy()
    deleted_count = original_count - len(df)

    df, matched, unmatched = add_pick_qty_by_sku(df, master_mapping)
    for column in ("BOXTYPE", "buyersreference", "packqty", "入數", "揀貨入數"):
        if column not in df.columns:
            df[column] = ""

    df["_是否GCOR排除"] = df.apply(
        lambda row: is_gcor_packqty_100(row.get("buyersreference", ""), row.get("packqty", "")), axis=1
    )
    units = df.apply(
        lambda row: calc_unit_qty(row.get("packqty", ""), row.get("揀貨入數", ""))
        if row["_是否GCOR排除"]
        else calc_normal_unit_qty(row.get("packqty", ""), row.get("入數", "")),
        axis=1,
    )
    if "單位入數" in df.columns:
        df["單位入數"] = units
    else:
        position = list(df.columns).index("揀貨入數") + 1
        df.insert(position, "單位入數", units)

    main_df = df.loc[~df["_是否GCOR排除"]].copy()
    gcor_df = df.loc[df["_是否GCOR排除"]].copy()
    main_boxtype = main_df["BOXTYPE"].map(normalize_text)
    gcor_boxtype = gcor_df["BOXTYPE"].map(normalize_text)

    main_b0 = main_df.loc[main_boxtype.eq("0"), "單位入數"].map(to_number).dropna().sum()
    main_b1_count = int(main_boxtype.eq("1").sum())
    main_b1_qty = main_df.loc[main_boxtype.eq("1"), "packqty"].map(to_number).dropna().sum()
    gcor_units = gcor_df["單位入數"].map(to_number).dropna().sum()
    gcor_b0 = gcor_df.loc[gcor_boxtype.eq("0"), "單位入數"].map(to_number).dropna().sum()
    gcor_b1_count = int(gcor_boxtype.eq("1").sum())
    gcor_b1_qty = gcor_df.loc[gcor_boxtype.eq("1"), "packqty"].map(to_number).dropna().sum()
    final_b0 = main_b0 + gcor_units

    stats = {
        "original_count": original_count,
        "after_delete_count": len(df),
        "deleted_count": deleted_count,
        "sku_matched_count": matched,
        "sku_unmatched_count": unmatched,
        "main_boxtype_0_unit_sum": main_b0,
        "main_boxtype_1_count": main_b1_count,
        "main_boxtype_1_packqty_sum": main_b1_qty,
        "gcor_count": len(gcor_df),
        "gcor_unit_sum": gcor_units,
        "gcor_boxtype_0_unit_sum": gcor_b0,
        "gcor_boxtype_1_count": gcor_b1_count,
        "gcor_boxtype_1_packqty_sum": gcor_b1_qty,
        "final_boxtype_0_unit_sum": final_b0,
    }
    summary_rows = [
        ["統計項目", "數量"],
        ["主統計 BOXTYPE=0 單位入數加總", main_b0],
        ["主統計 BOXTYPE=1 筆數", main_b1_count],
        ["主統計 BOXTYPE=1 packqty加總", main_b1_qty],
        ["GCOR排除資料 單位入數加總", gcor_units],
        ["BOXTYPE=0 單位入數加總+GCOR排除資料 單位入數加總", final_b0],
        ["SKU比對成功筆數", matched],
        ["SKU比對不到筆數", unmatched],
    ]
    gcor_summary_rows = [
        ["統計項目", "數量"],
        ["GCOR且packqty>=100資料筆數", len(gcor_df)],
        ["GCOR排除資料 單位入數加總", gcor_units],
        ["BOXTYPE=0 單位入數加總", gcor_b0],
        ["BOXTYPE=1 筆數", gcor_b1_count],
        ["BOXTYPE=1 packqty 加總數量", gcor_b1_qty],
    ]
    return (
        df.drop(columns="_是否GCOR排除"),
        gcor_df.drop(columns="_是否GCOR排除"),
        summary_rows,
        gcor_summary_rows,
        stats,
    )


def _append_dataframe(ws, df: pd.DataFrame) -> None:
    ws.append(list(df.columns))
    for row in df.itertuples(index=False, name=None):
        ws.append([clean_illegal_excel_text(value) for value in row])


def _format_workbook(workbook: Workbook) -> None:
    header_fill = PatternFill("solid", fgColor=GREEN)
    header_font = Font(color="FFFFFF", bold=True)
    for ws in workbook.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        for index, cells in enumerate(ws.columns, start=1):
            length = max((len(str(cell.value)) for cell in cells if cell.value is not None), default=0)
            ws.column_dimensions[get_column_letter(index)].width = min(length + 2, 50)


def make_detail_excel(df_output, gcor_detail_df, summary_rows, gcor_summary_rows) -> bytes:
    workbook = Workbook()
    ws = workbook.active
    ws.title = "處理後資料"
    _append_dataframe(ws, df_output)
    ws = workbook.create_sheet("統計總表")
    for row in summary_rows:
        ws.append(row)
    ws = workbook.create_sheet("GCOR排除統計")
    for row in gcor_summary_rows:
        ws.append(row)
    ws = workbook.create_sheet("GCOR排除資料明細")
    _append_dataframe(ws, gcor_detail_df)
    _format_workbook(workbook)
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def empty_total_stats() -> dict[str, float]:
    return {key: 0 for key in (
        "original_count", "after_delete_count", "deleted_count", "sku_matched_count",
        "sku_unmatched_count", "main_boxtype_0_unit_sum", "main_boxtype_1_count",
        "main_boxtype_1_packqty_sum", "gcor_count", "gcor_unit_sum",
        "gcor_boxtype_0_unit_sum", "gcor_boxtype_1_count",
        "gcor_boxtype_1_packqty_sum", "final_boxtype_0_unit_sum",
    )}


def make_batch_summary(batch_rows: list[dict], total: dict[str, float]) -> bytes:
    headers = [
        "檔名", "原始資料筆數", "刪除FT03~FT09筆數", "刪除後保留筆數", "SKU比對成功筆數",
        "SKU比對不到筆數", "主統計_BOXTYPE=0單位入數加總", "主統計_BOXTYPE=1筆數",
        "主統計_BOXTYPE=1packqty加總", "GCOR排除資料筆數", "GCOR排除資料_單位入數加總",
        "BOXTYPE=0單位入數加總+GCOR排除資料單位入數加總", "GCOR_BOXTYPE=0單位入數加總",
        "GCOR_BOXTYPE=1筆數", "GCOR_BOXTYPE=1packqty加總",
    ]
    keys = [
        "file_name", "original_count", "deleted_count", "after_delete_count", "sku_matched_count",
        "sku_unmatched_count", "main_boxtype_0_unit_sum", "main_boxtype_1_count",
        "main_boxtype_1_packqty_sum", "gcor_count", "gcor_unit_sum", "final_boxtype_0_unit_sum",
        "gcor_boxtype_0_unit_sum", "gcor_boxtype_1_count", "gcor_boxtype_1_packqty_sum",
    ]
    workbook = Workbook()
    ws = workbook.active
    ws.title = "各檔案統計"
    ws.append(headers)
    for item in batch_rows:
        ws.append([item[key] for key in keys])
    ws = workbook.create_sheet("全部檔案總計")
    ws.append(["統計項目", "數量"])
    labels = dict(zip(keys[1:], headers[1:]))
    for key in keys[1:]:
        ws.append([labels[key], total[key]])
    _format_workbook(workbook)
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def _fmt(value) -> str:
    try:
        number = float(value)
        return f"{int(number):,}" if number.is_integer() else f"{number:,.2f}"
    except (TypeError, ValueError):
        return "0"


def run_app() -> None:
    _setup_page()
    st.markdown('<div class="df-title">📦 月出貨量與產力</div>', unsafe_allow_html=True)
    st.markdown('<div class="df-subtitle">上傳商品主檔與月出貨 TXT／CSV，系統會自動清理、比對並彙總。</div>', unsafe_allow_html=True)

    left, right = st.columns(2)
    with left:
        master_file = st.file_uploader("① 商品主檔", type=["xlsx", "xls"], help="需包含「商品代號」及「揀貨入數」欄位")
    with right:
        data_files = st.file_uploader("② 月出貨資料（可多選）", type=["txt", "csv"], accept_multiple_files=True)

    if not master_file or not data_files:
        st.info("請先上傳 1 份商品主檔，以及至少 1 份 TXT／CSV 出貨資料。")
        return

    if st.button("開始計算", type="primary", use_container_width=True):
        try:
            mapping, sheet_name = load_master_mapping(master_file)
        except Exception as exc:
            st.error(f"商品主檔讀取失敗：{exc}")
            return

        batch_rows: list[dict] = []
        failed: list[dict] = []
        detail_files: list[tuple[str, bytes]] = []
        total = empty_total_stats()
        progress = st.progress(0, text="開始處理…")

        for index, uploaded in enumerate(data_files, start=1):
            try:
                df = read_txt_smart(uploaded)
                output_df, gcor_df, summary, gcor_summary, stats = process_dataframe(df, mapping)
                stem = re.sub(r"\.(txt|csv)$", "", uploaded.name, flags=re.IGNORECASE)
                output_name = f"{stem}_月出貨量與產力.xlsx"
                detail_files.append((output_name, make_detail_excel(output_df, gcor_df, summary, gcor_summary)))
                row = {"file_name": uploaded.name, **stats}
                batch_rows.append(row)
                for key in total:
                    total[key] += stats[key]
            except Exception as exc:
                failed.append({"檔名": uploaded.name, "錯誤原因": str(exc)})
            progress.progress(index / len(data_files), text=f"正在處理 {index}/{len(data_files)}：{uploaded.name}")
        progress.empty()

        if not batch_rows:
            st.error("所有檔案均處理失敗，請檢查檔案格式與欄位名稱。")
            st.dataframe(pd.DataFrame(failed), use_container_width=True, hide_index=True)
            return

        summary_bytes = make_batch_summary(batch_rows, total)
        zip_output = io.BytesIO()
        with zipfile.ZipFile(zip_output, "w", zipfile.ZIP_DEFLATED) as archive:
            archive.writestr("多檔案彙總統計.xlsx", summary_bytes)
            for filename, content in detail_files:
                archive.writestr(filename, content)

        st.session_state["monthly_shipping_result"] = {
            "batch_rows": batch_rows,
            "failed": failed,
            "total": total,
            "summary": summary_bytes,
            "zip": zip_output.getvalue(),
            "sheet": sheet_name,
            "mapping_count": len(mapping),
        }

    result = st.session_state.get("monthly_shipping_result")
    if not result:
        return

    st.success(f"處理完成：成功 {len(result['batch_rows'])} 個、失敗 {len(result['failed'])} 個；商品主檔使用「{result['sheet']}」分頁。")
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("BOXTYPE=0 出貨單位", _fmt(result["total"]["final_boxtype_0_unit_sum"]))
    c2.metric("BOXTYPE=1 出貨數量", _fmt(result["total"]["main_boxtype_1_packqty_sum"]))
    c3.metric("SKU 比對成功", _fmt(result["total"]["sku_matched_count"]))
    c4.metric("已排除 FT03～FT09", _fmt(result["total"]["deleted_count"]))

    display = pd.DataFrame(result["batch_rows"]).rename(columns={
        "file_name": "檔名", "original_count": "原始筆數", "deleted_count": "排除FT筆數",
        "after_delete_count": "保留筆數", "sku_matched_count": "SKU比對成功",
        "sku_unmatched_count": "SKU比對不到", "final_boxtype_0_unit_sum": "BOXTYPE=0出貨單位",
        "main_boxtype_1_packqty_sum": "BOXTYPE=1出貨數量",
    })
    visible = ["檔名", "原始筆數", "排除FT筆數", "保留筆數", "SKU比對成功", "SKU比對不到", "BOXTYPE=0出貨單位", "BOXTYPE=1出貨數量"]
    st.dataframe(display[visible], use_container_width=True, hide_index=True)

    d1, d2 = st.columns(2)
    d1.download_button("下載彙總統計 Excel", result["summary"], "多檔案彙總統計.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
    d2.download_button("下載全部結果 ZIP", result["zip"], "月出貨量與產力_全部結果.zip", "application/zip", type="primary", use_container_width=True)

    if result["failed"]:
        with st.expander("查看處理失敗檔案", expanded=True):
            st.dataframe(pd.DataFrame(result["failed"]), use_container_width=True, hide_index=True)


run_app()
