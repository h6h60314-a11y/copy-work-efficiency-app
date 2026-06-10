from io import BytesIO

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from common_ui import inject_logistics_theme, set_page, card_open, card_close

st.set_page_config(page_title="拉單明細", page_icon="📄", layout="wide")
inject_logistics_theme()
set_page("拉單明細", "多檔拉單明細合併整理，最後只產出一份 Excel。")

KEEP_COLUMNS = [
    "貨主",
    "商品",
    "數量",
    "儲位",
    "成箱箱號",
    "計量單位",
    "計量單位數量",
    "出貨單位",
    "出貨入數",
    "原始配庫存量",
    "貨主訂單",
    "門市代號",
    "門市名稱",
]


def clean_header(value):
    if pd.isna(value):
        return value
    return str(value).strip()


def adjust_one_sheet(df, sheet_name, file_name):
    df = df.copy()
    df.columns = [clean_header(col) for col in df.columns]

    missing_columns = [col for col in KEEP_COLUMNS if col not in df.columns]

    if missing_columns:
        raise ValueError(
            f"檔案「{file_name}」工作表「{sheet_name}」缺少欄位：{', '.join(missing_columns)}"
        )

    return df[KEEP_COLUMNS]


def autosize_excel(excel_bytes):
    bio = BytesIO(excel_bytes)
    wb = load_workbook(bio)

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"

        for column_cells in ws.columns:
            max_length = 0
            col_letter = get_column_letter(column_cells[0].column)

            for cell in column_cells:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))

            ws.column_dimensions[col_letter].width = min(max_length + 2, 28)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def process_files(uploaded_files):
    merged_sheets = {}

    for uploaded_file in uploaded_files:
        sheets = pd.read_excel(uploaded_file, sheet_name=None, dtype=str)

        for sheet_name, df in sheets.items():
            adjusted_df = adjust_one_sheet(df, sheet_name, uploaded_file.name)

            if sheet_name not in merged_sheets:
                merged_sheets[sheet_name] = []

            merged_sheets[sheet_name].append(adjusted_df)

    output = BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for sheet_name, dfs in merged_sheets.items():
            final_df = pd.concat(dfs, ignore_index=True)
            final_df.to_excel(writer, sheet_name=sheet_name, index=False)

    return autosize_excel(output.getvalue())


card_open("📄 拉單明細整理")

st.write("請上傳一個或多個 Excel 檔，系統會合併後只產出一份整理後的 Excel。")

uploaded_files = st.file_uploader(
    "上傳拉單明細 Excel",
    type=["xlsx", "xls"],
    accept_multiple_files=True,
)

if uploaded_files:
    st.info(f"已上傳 {len(uploaded_files)} 個檔案")

    for file in uploaded_files:
        st.write(f"✅ {file.name}")

    if st.button("開始整理", use_container_width=True):
        try:
            output_bytes = process_files(uploaded_files)

            st.success("拉單明細整理完成")

            st.download_button(
                label="下載拉單明細整理.xlsx",
                data=output_bytes,
                file_name="拉單明細整理.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

        except Exception as e:
            st.error("處理失敗")
            st.write(str(e))

else:
    st.info("請先上傳 Excel 檔案。")

card_close()
